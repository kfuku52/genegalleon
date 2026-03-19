#!/usr/bin/env python3

import argparse
import bz2
from collections import defaultdict, deque
from concurrent.futures import ThreadPoolExecutor, as_completed
import csv
from http.client import IncompleteRead, RemoteDisconnected
import gzip
import hashlib
import io
import json
import os
from pathlib import Path
import re
import shutil
import sqlite3
import socket
import subprocess
import sys
import tarfile
import threading
import time
import zipfile
from urllib.parse import parse_qs, quote, unquote, urljoin, urlparse
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - exercised in runtime environments without openpyxl
    load_workbook = None

try:
    from Bio import SeqIO
except Exception:  # pragma: no cover - exercised in runtime environments without biopython
    SeqIO = None


COMMON_REPLACEMENTS = (
    ("evm_27.model.", ""),
    ("evm.model.", ""),
    ("Oropetium_20150105_", ""),
)
GENE_GROUPING_MODES = ("strict", "rescue_overlap")
RESCUE_SHARED_JUNCTION_MIN_SHORTER_OVERLAP = 0.70
RESCUE_SHARED_JUNCTION_MIN_LONGER_OVERLAP = 0.40
RESCUE_SAME_TERMINAL_MIN_SHORTER_OVERLAP = 0.90
RESCUE_SAME_TERMINAL_MIN_LONGER_OVERLAP = 0.60

FASTA_EXTENSIONS = (
    ".fa",
    ".fas",
    ".fasta",
    ".fna",
    ".fa.bz2",
    ".fas.bz2",
    ".fasta.bz2",
    ".fna.bz2",
    ".fa.gz",
    ".fas.gz",
    ".fasta.gz",
    ".fna.gz",
    ".fa.tar.gz",
    ".fas.tar.gz",
    ".fasta.tar.gz",
    ".fna.tar.gz",
    ".fa.tar.bz2",
    ".fas.tar.bz2",
    ".fasta.tar.bz2",
    ".fna.tar.bz2",
)

FASTA_ARCHIVE_EXTENSIONS = (
    ".fa.tar.gz",
    ".fas.tar.gz",
    ".fasta.tar.gz",
    ".fna.tar.gz",
    ".fa.tar.bz2",
    ".fas.tar.bz2",
    ".fasta.tar.bz2",
    ".fna.tar.bz2",
)

GFF_EXTENSIONS = (
    ".gff",
    ".gff3",
    ".gtf",
    ".gff.gz",
    ".gff3.gz",
    ".gtf.gz",
)

GENBANK_EXTENSIONS = (
    ".gb",
    ".gbk",
    ".gbff",
    ".genbank",
    ".gb.gz",
    ".gbk.gz",
    ".gbff.gz",
    ".genbank.gz",
)

ENSEMBL_CDS_PATTERN = re.compile(r"(?:^|[._-])cds(?:[._-]|$)", re.IGNORECASE)
INVALID_ID_CHARS = re.compile(r"[%/\+:;&\^\$#@!~=\'\"`*\(\)\{\}\[\]\|\?\s]+")
NCBI_ASSEMBLY_ACCESSION_PATTERN = re.compile(r"^GC[AF]_[0-9]+\.[0-9]+$", re.IGNORECASE)
ENSEMBL_GENE_ID_PATTERN = re.compile(r"^ENS[A-Z0-9]*G[0-9]+(?:\.[0-9]+)?$", re.IGNORECASE)
COGE_ID_HINT_PATTERN = re.compile(r"^coge[:_].+", re.IGNORECASE)
CNGB_ID_HINT_PATTERN = re.compile(r"^cngb[:_].+", re.IGNORECASE)
GWH_ID_HINT_PATTERN = re.compile(r"^gwh[:_].+", re.IGNORECASE)
ENSEMBL_ID_HINT_PATTERN = re.compile(r"^ensembl[:_].+", re.IGNORECASE)
FERNBASE_ID_HINT_PATTERN = re.compile(r"^fernbase[:_].+", re.IGNORECASE)
VEUPATHDB_ID_HINT_PATTERN = re.compile(r"^veupathdb[:_].+", re.IGNORECASE)
DICTYBASE_ID_HINT_PATTERN = re.compile(r"^dictybase[:_].+", re.IGNORECASE)
INSECTBASE_ID_HINT_PATTERN = re.compile(r"^insectbase[:_].+", re.IGNORECASE)
INSECTBASE_IBG_ID_PATTERN = re.compile(r"^IBG_[0-9]+$", re.IGNORECASE)
COGE_GID_PATTERN = re.compile(r"^[0-9]+$")
CNGB_ASSEMBLY_ACCESSION_PATTERN = re.compile(r"^CNA[0-9]+$", re.IGNORECASE)
GWH_ASSEMBLY_ACCESSION_PATTERN = re.compile(r"^GWH[A-Z0-9]+(?:\.[0-9]+)?$", re.IGNORECASE)
GWH_ASSEMBLY_ACCESSION_SEARCH_PATTERN = re.compile(r"(GWH[A-Z0-9]+(?:\.[0-9]+)?)", re.IGNORECASE)

DEFAULT_INPUT_RELATIVE_DIRS = {
    "ensembl": Path("Ensembl") / "original_files",
    "ensemblplants": Path("20230216_EnsemblPlants") / "original_files",
    "phycocosm": Path("PhycoCosm") / "species_wise_original",
    "phytozome": Path("Phytozome") / "species_wise_original",
    "ncbi": Path("NCBI_Genome") / "species_wise_original",
    "refseq": Path("NCBI_RefSeq") / "species_wise_original",
    "genbank": Path("NCBI_GenBank") / "species_wise_original",
    "coge": Path("CoGe") / "species_wise_original",
    "cngb": Path("CNGB") / "species_wise_original",
    "gwh": Path("GWH") / "species_wise_original",
    "flybase": Path("FlyBase") / "species_wise_original",
    "wormbase": Path("WormBase") / "species_wise_original",
    "vectorbase": Path("VectorBase") / "species_wise_original",
    "fernbase": Path("FernBase") / "species_wise_original",
    "veupathdb": Path("VEuPathDB") / "species_wise_original",
    "dictybase": Path("dictyBase") / "species_wise_original",
    "insectbase": Path("InsectBase") / "species_wise_original",
    "direct": Path("Direct") / "species_wise_original",
    "local": Path("Local") / "species_wise_original",
}

PROVIDERS = (
    "ensembl",
    "ensemblplants",
    "phycocosm",
    "phytozome",
    "ncbi",
    "refseq",
    "genbank",
    "coge",
    "cngb",
    "gwh",
    "flybase",
    "wormbase",
    "vectorbase",
    "fernbase",
    "veupathdb",
    "dictybase",
    "insectbase",
    "direct",
    "local",
)
DOWNLOAD_MANIFEST_SUPPORTED_PROVIDERS = (
    "ensembl",
    "ensemblplants",
    "ncbi",
    "refseq",
    "genbank",
    "coge",
    "cngb",
    "gwh",
    "flybase",
    "wormbase",
    "vectorbase",
    "fernbase",
    "veupathdb",
    "dictybase",
    "insectbase",
    "direct",
    "local",
)
DEFAULT_DOWNLOAD_LOCK_STALE_SECONDS = 900
DEFAULT_DOWNLOAD_LOCK_HEARTBEAT_SECONDS = 60
DEFAULT_DOWNLOAD_LOCK_ACQUIRE_TIMEOUT_SECONDS = 3600
DEFAULT_DOWNLOAD_LOCK_POLL_SECONDS = 5.0
SHARED_DOWNLOAD_LOCK_FORMAT = "shared-lock-v2"
DEFAULT_NCBI_EUTILS_BASE_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils"
DEFAULT_NCBI_FTP_BASE_URL = "https://ftp.ncbi.nlm.nih.gov"
DEFAULT_NCBI_DATASETS_BASE_URL = "https://api.ncbi.nlm.nih.gov/datasets/v2"
TRANSIENT_HTTP_STATUS_CODES = frozenset((408, 425, 429, 500, 502, 503, 504))
DEFAULT_COGE_API_BASE_URL = "https://genomevolution.org/coge/api/v1"
DEFAULT_COGE_WEB_BASE_URL = "https://genomevolution.org/coge"
DEFAULT_CNGB_CNSA_BASE_URL = "https://db.cngb.org/cnsa/ajax"
DEFAULT_GWH_DOWNLOAD_BASE_URL = "https://download.cncb.ac.cn/gwh"
DEFAULT_VEUPATHDB_SERVICE_BASE_URL = "https://veupathdb.org/veupathdb/service"
DEFAULT_INSECTBASE_API_BASE_URL = "https://www.insect-genome.com/api/genome"
NCBI_DATASETS_INCLUDE_BY_LABEL = {
    "CDS": "CDS_FASTA",
    "GFF": "GENOME_GFF",
    "GBFF": "GENOME_GBFF",
    "GENOME": "GENOME_FASTA",
}
DOWNLOAD_LABELS = ("CDS", "GFF", "GENOME")
RESOLVED_MANIFEST_PREFERRED_COLUMNS = (
    "provider",
    "id",
    "species_key",
    "cds_url",
    "gff_url",
    "gbff_url",
    "genome_url",
    "cds_archive_member",
    "gff_archive_member",
    "gbff_archive_member",
    "genome_archive_member",
    "cds_filename",
    "gff_filename",
    "gbff_filename",
    "genome_filename",
)
ENSEMBL_DEFAULT_ID_URL_TEMPLATES = {
    "CDS": "https://ftp.ensembl.org/pub/current_fasta/{id_lower}/cds/",
    "GFF": "https://ftp.ensembl.org/pub/current_gff3/{id_lower}/",
    "GENOME": "https://ftp.ensembl.org/pub/current_fasta/{id_lower}/dna/",
}
ENSEMBLPLANTS_DEFAULT_ID_URL_TEMPLATES = {
    "CDS": "https://ftp.ensemblgenomes.ebi.ac.uk/pub/plants/current/fasta/{id_lower}/cds/",
    "GFF": "https://ftp.ensemblgenomes.ebi.ac.uk/pub/plants/current/gff3/{id_lower}/",
    "GENOME": "https://ftp.ensemblgenomes.ebi.ac.uk/pub/plants/current/fasta/{id_lower}/dna/",
}
PROVIDER_DEFAULT_ID_PAGE_URL_TEMPLATES = {
    "phycocosm": (
        "https://genome.jgi.doe.gov/portal/pages/dynamicOrganismDownload.jsf?organism={id}",
    ),
    "phytozome": (
        "https://phytozome-next.jgi.doe.gov/download/file/{id}",
        "https://phytozome-next.jgi.doe.gov/info/{id}",
        "https://data.jgi.doe.gov/refine-download/phytozome?genome_id={id}",
    ),
    "coge": (
        "https://genomevolution.org/coge/GenomeInfo.pl?gid={id}",
    ),
    "cngb": (
        "https://db.cngb.org/data_resources/project/{id}",
    ),
    "gwh": (
        "https://download.cncb.ac.cn/gwh/",
    ),
    "fernbase": (
        "https://fernbase.org/ftp/{id}/",
    ),
    "veupathdb": (
        "https://veupathdb.org/",
    ),
    "dictybase": (
        "https://dictybase.org/Downloads/",
        "https://dictybase.org/download/",
    ),
    "insectbase": (
        "https://www.insect-genome.com/genome/{id}",
    ),
}
PROVIDER_DEFAULT_MAX_CONCURRENT_DOWNLOADS = {
    # NCBI E-utilities explicitly documents request-rate limits.
    # We keep concurrent file downloads conservative and separately enforce
    # E-utilities request throttling below.
    "ncbi": 2,
    "refseq": 2,
    "genbank": 2,
    # No explicit numeric concurrency guidance was found for these sources;
    # keep defaults conservative to reduce ban/abuse risk.
    "ensembl": 2,
    "ensemblplants": 2,
    "phycocosm": 1,
    "phytozome": 1,
    "coge": 1,
    "cngb": 1,
    "gwh": 1,
    "flybase": 2,
    "wormbase": 2,
    "vectorbase": 2,
    "fernbase": 2,
    "veupathdb": 1,
    "dictybase": 1,
    "insectbase": 2,
    "direct": 2,
    "local": 1,
}
DEFAULT_GLOBAL_DOWNLOAD_WORKERS = 1
DEFAULT_NCBI_EUTILS_RPS_NO_API_KEY = 3
DEFAULT_NCBI_EUTILS_RPS_WITH_API_KEY = 10
_ncbi_eutils_rate_lock = threading.Lock()
_ncbi_eutils_request_times = deque()
SPECIES_SUMMARY_COLUMNS = (
    "updated_utc",
    "run_started_utc",
    "provider",
    "species_key",
    "species_prefix",
    "taxid",
    "nuclear_genetic_code_id",
    "nuclear_genetic_code_name",
    "mitochondrial_genetic_code_id",
    "mitochondrial_genetic_code_name",
    "plastid_genetic_code_id",
    "plastid_genetic_code_name",
    "cds_input_path",
    "gff_input_path",
    "genome_input_path",
    "cds_output_path",
    "gff_output_path",
    "genome_output_path",
    "cds_status",
    "gff_status",
    "genome_status",
    "cds_sequences_before",
    "cds_sequences_after",
    "cds_first_sequence_name",
    "aggregated_cds_removed",
    "gene_grouping_mode",
    "overwrite",
    "dry_run",
)
PLASTID_GENETIC_CODE_LINEAGE_DEFAULTS = {
    # NCBI taxonomy/taxdump exposes nuclear and mitochondrial codes directly, but
    # not plastid codes. Use the standard plastid code for well-established
    # plastid-bearing clades as a conservative best-effort fallback.
    "33090": "11",    # Viridiplantae
    "2763": "11",     # Rhodophyta
    "2830": "11",     # Haptophyta
    "3027": "11",     # Cryptophyceae
    "2696291": "11",  # Ochrophyta
}


def blank_species_taxonomy_metadata():
    return {
        "taxid": "",
        "nuclear_genetic_code_id": "",
        "nuclear_genetic_code_name": "",
        "mitochondrial_genetic_code_id": "",
        "mitochondrial_genetic_code_name": "",
        "plastid_genetic_code_id": "",
        "plastid_genetic_code_name": "",
    }


class SpeciesTaxonomyMetadataResolver:
    def __init__(self, taxonomy_dbfile="", taxonomy_taxdumpfile=""):
        self.taxonomy_dbfile = str(taxonomy_dbfile or "").strip()
        self.taxonomy_taxdumpfile = str(taxonomy_taxdumpfile or "").strip()
        self._conn = None
        self._nodes = None
        self._genetic_codes = None
        self._cache = {}

    @classmethod
    def from_environment(cls):
        return cls(
            taxonomy_dbfile=os.environ.get("GG_TAXONOMY_DBFILE", ""),
            taxonomy_taxdumpfile=os.environ.get("GG_TAXONOMY_TAXDUMPFILE", ""),
        )

    def _ensure_conn(self):
        if self._conn is not None:
            return self._conn
        if self.taxonomy_dbfile == "":
            return None
        db_path = Path(self.taxonomy_dbfile).expanduser()
        if not db_path.exists():
            return None
        self._conn = sqlite3.connect(str(db_path))
        return self._conn

    def _ensure_taxdump(self):
        if self._nodes is not None and self._genetic_codes is not None:
            return
        self._nodes = {}
        self._genetic_codes = {}
        if self.taxonomy_taxdumpfile == "":
            return
        taxdump_path = Path(self.taxonomy_taxdumpfile).expanduser()
        if not taxdump_path.exists():
            return
        with tarfile.open(taxdump_path, "r:gz") as archive:
            with archive.extractfile("gencode.dmp") as handle:
                for raw in handle:
                    parts = [part.strip() for part in raw.decode("utf-8").split("|")]
                    if len(parts) < 3 or parts[0] == "":
                        continue
                    self._genetic_codes[parts[0]] = {
                        "abbr": parts[1],
                        "name": parts[2],
                    }
            with archive.extractfile("nodes.dmp") as handle:
                for raw in handle:
                    parts = [part.strip() for part in raw.decode("utf-8").split("|")]
                    if len(parts) < 9 or parts[0] == "":
                        continue
                    self._nodes[parts[0]] = {
                        "genetic_code_id": parts[6],
                        "mitochondrial_genetic_code_id": parts[8],
                    }

    def _lookup_species_row(self, species_name):
        conn = self._ensure_conn()
        if conn is None:
            return None
        cur = conn.cursor()
        row = cur.execute(
            "SELECT taxid, spname, rank, track FROM species WHERE spname = ? COLLATE NOCASE",
            (species_name,),
        ).fetchone()
        if row is None:
            row = cur.execute(
                """
                SELECT s.taxid, s.spname, s.rank, s.track
                FROM synonym sy
                JOIN species s ON sy.taxid = s.taxid
                WHERE sy.spname = ? COLLATE NOCASE
                """,
                (species_name,),
            ).fetchone()
        return row

    def _code_name(self, code_id):
        if code_id == "":
            return ""
        self._ensure_taxdump()
        return str(self._genetic_codes.get(str(code_id), {}).get("name", ""))

    def _infer_plastid_code_id(self, lineage_taxids):
        lineage_taxid_set = {str(taxid).strip() for taxid in lineage_taxids if str(taxid).strip() != ""}
        for lineage_taxid, code_id in PLASTID_GENETIC_CODE_LINEAGE_DEFAULTS.items():
            if lineage_taxid in lineage_taxid_set:
                return code_id
        return ""

    def resolve(self, species_name):
        normalized_name = str(species_name or "").strip().replace("_", " ")
        if normalized_name == "":
            return blank_species_taxonomy_metadata()
        if normalized_name in self._cache:
            return dict(self._cache[normalized_name])

        metadata = blank_species_taxonomy_metadata()
        row = self._lookup_species_row(normalized_name)
        if row is None:
            self._cache[normalized_name] = dict(metadata)
            return metadata

        taxid, _spname, _rank, track = row
        taxid_str = str(taxid)
        metadata["taxid"] = taxid_str

        self._ensure_taxdump()
        node = self._nodes.get(taxid_str, {})
        nuclear_code_id = str(node.get("genetic_code_id", "") or "")
        mitochondrial_code_id = str(node.get("mitochondrial_genetic_code_id", "") or "")
        metadata["nuclear_genetic_code_id"] = nuclear_code_id
        metadata["nuclear_genetic_code_name"] = self._code_name(nuclear_code_id)
        metadata["mitochondrial_genetic_code_id"] = mitochondrial_code_id
        metadata["mitochondrial_genetic_code_name"] = self._code_name(mitochondrial_code_id)

        lineage_taxids = [part.strip() for part in str(track or "").split(",") if part.strip() != ""]
        plastid_code_id = self._infer_plastid_code_id(lineage_taxids)
        metadata["plastid_genetic_code_id"] = plastid_code_id
        metadata["plastid_genetic_code_name"] = self._code_name(plastid_code_id)

        self._cache[normalized_name] = dict(metadata)
        return metadata


def output_gzip_compresslevel():
    raw = os.environ.get("GG_INPUT_GZIP_LEVEL", "").strip()
    if raw == "":
        return 9
    try:
        level = int(raw)
    except ValueError:
        raise ValueError("GG_INPUT_GZIP_LEVEL must be an integer between 1 and 9.")
    if level < 1 or level > 9:
        raise ValueError("GG_INPUT_GZIP_LEVEL must be between 1 and 9.")
    return level


def output_compression_threads():
    raw = os.environ.get("GG_TASK_CPUS", "").strip()
    if raw == "":
        return 1
    try:
        value = int(raw)
    except ValueError:
        return 1
    return max(1, value)


def build_arg_parser():
    parser = argparse.ArgumentParser(
        description=(
            "Format CDS/GFF inputs for genegalleon from provider-specific raw files. "
            "Accepts CDS-only inputs, GBFF inputs, or GFF plus genome FASTA inputs. "
            "When CDS is missing but both GFF and genome FASTA are present, CDS is derived automatically. "
            "Current targets are based on legacy data_formatting*.sh scripts."
        )
    )
    parser.add_argument(
        "--provider",
        choices=("all",) + PROVIDERS,
        required=True,
        help="Input provider type. Use 'all' with --input-dir pointing to a provider-root directory.",
    )
    parser.add_argument(
        "--input-dir",
        default="",
        help=(
            "Provider input directory. For ensembl/ensemblplants: original_files/. "
            "For phycocosm/phytozome/ncbi/coge/cngb/gwh/flybase/wormbase/vectorbase/fernbase/veupathdb/dictybase/insectbase/direct/local: species_wise_original/. "
            "Legacy aliases refseq/genbank are treated as ncbi. "
            "For --provider all, this must be the shared root containing all provider subdirectories."
        ),
    )
    parser.add_argument(
        "--species-cds-dir",
        default="workspace/output/input_generation/species_cds",
        help="Output directory for formatted species CDS FASTA files.",
    )
    parser.add_argument(
        "--species-gff-dir",
        default="workspace/output/input_generation/species_gff",
        help="Output directory for formatted species GFF files.",
    )
    parser.add_argument(
        "--species-genome-dir",
        default="workspace/output/input_generation/species_genome",
        help="Output directory for formatted species genome FASTA files.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing outputs.",
    )
    parser.add_argument(
        "--download-manifest",
        default="",
        help=(
            "Optional XLSX/TSV/CSV manifest for input download. "
            "Required columns: provider,id (in this order for XLSX templates). "
            "provider,id are required on every row; species_key is optional. "
            "CDS/GFF/genome can be set directly with cds_url/gff_url/genome_url or resolved from id "
            "(ncbi supports GCF/GCA/NCBI-URL auto-resolution; "
            "other supported providers support id-based template/index inference). "
            "Supported providers for --download-manifest: "
            "ensembl, ensemblplants, ncbi, coge, cngb, gwh, flybase, wormbase, vectorbase, fernbase, veupathdb, dictybase, insectbase, direct, local "
            "(legacy aliases refseq/genbank are treated as ncbi). "
            "provider=direct requires explicit urls or an index-style id URL. "
            "cds_url may be provided by itself, or may be empty when both gff_url and genome_url are available; "
            "genegalleon will derive CDS during formatting. "
            "provider=local reads local files/directories (for example local phytozome files). "
            "Use --input-dir for direct local phycocosm/phytozome formatting. "
            "Optional columns: species_key,cds_filename,gff_filename,genome_filename,"
            "cds_url_template,gff_url_template,genome_url_template,"
            "local_cds_path,local_gff_path,local_genome_path."
        ),
    )
    parser.add_argument(
        "--resolved-manifest-output",
        default="workspace/output/input_generation/download_plan.resolved.tsv",
        help=(
            "Output TSV path for resolved download-manifest rows. "
            "Rows selected for download are written with provider/species_key/URL/filename fields filled. "
            "Set empty string to disable."
        ),
    )
    parser.add_argument(
        "--download-dir",
        default="workspace/output/input_generation/tmp/input_download_cache",
        help=(
            "Directory for raw downloaded provider files. "
            "This can be reused as --input-dir for formatting."
        ),
    )
    parser.add_argument(
        "--download-only",
        action="store_true",
        help="Download raw files from manifest and exit without formatting.",
    )
    parser.add_argument(
        "--http-header",
        action="append",
        default=[],
        help="Additional HTTP header for download requests. Repeatable. Format: 'Key: Value'.",
    )
    parser.add_argument(
        "--auth-bearer-token-env",
        default="",
        help="Environment variable name containing bearer token for download Authorization header.",
    )
    parser.add_argument(
        "--download-timeout",
        type=float,
        default=120.0,
        help="Timeout (seconds) per download request.",
    )
    parser.add_argument(
        "--jobs",
        type=int,
        default=0,
        help=(
            "Maximum parallel download workers. "
            "If <=0, uses GG_TASK_CPUS when available, otherwise legacy NSLOTS, otherwise 1."
        ),
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Plan actions without writing downloaded or formatted files.",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Exit with error if any species lacks a usable source bundle (CDS, GBFF, or GFF plus genome).",
    )
    parser.add_argument(
        "--gene-grouping-mode",
        choices=GENE_GROUPING_MODES,
        default="rescue_overlap",
        help=(
            "How to group transcript/CDS records into gene-level representatives. "
            "'strict' uses annotation hierarchy and identifier normalization only. "
            "'rescue_overlap' also merges highly similar overlapping GFF-derived transcripts "
            "when annotation gene IDs appear inconsistent."
        ),
    )
    parser.add_argument(
        "--stats-output",
        default="",
        help=(
            "Optional JSON output path for run stats. "
            "Includes CDS counts before/after processing and first CDS sequence name."
        ),
    )
    parser.add_argument(
        "--species-summary-output",
        default="workspace/output/input_generation/gg_input_generation_species.tsv",
        help=(
            "Per-species TSV summary output path. "
            "Successful formatting results are persisted incrementally and retained across runs "
            "while referenced species output files exist."
        ),
    )
    return parser


def open_text(path, mode):
    if path.name.endswith(".gz"):
        kwargs = {"encoding": "utf-8"}
        if any(flag in mode for flag in ("w", "a", "x")):
            kwargs["compresslevel"] = output_gzip_compresslevel()
        return gzip.open(path, mode, **kwargs)
    if path.name.endswith(".bz2"):
        return bz2.open(path, mode, encoding="utf-8")
    return open(path, mode, encoding="utf-8")


def make_temporary_output_path(output_path):
    base_name = output_path.name
    suffix = output_path.suffix
    if suffix != "" and base_name.endswith(suffix):
        base_name = base_name[: -len(suffix)]
    return output_path.parent / ".{}.tmp.{}.{}{}".format(
        base_name,
        os.getpid(),
        time.time_ns(),
        suffix,
    )


def write_text_output_via_command(output_path, writer, command_builder, output_via_stdout):
    tmp_output = make_temporary_output_path(output_path)
    stderr_text = ""
    proc = None
    stdout_handle = None
    try:
        stdout_target = subprocess.DEVNULL
        if output_via_stdout:
            stdout_handle = open(tmp_output, "wb")
            stdout_target = stdout_handle
        command = command_builder(tmp_output)
        proc = subprocess.Popen(
            command,
            stdin=subprocess.PIPE,
            stdout=stdout_target,
            stderr=subprocess.PIPE,
            text=True,
            encoding="utf-8",
        )
        try:
            if proc.stdin is None:
                raise RuntimeError("Failed to open stdin for: {}".format(" ".join(command)))
            writer(proc.stdin)
            proc.stdin.close()
            proc.stdin = None
            if proc.stderr is not None:
                stderr_text = proc.stderr.read()
            return_code = proc.wait()
        except Exception:
            if proc.stdin is not None:
                try:
                    proc.stdin.close()
                except Exception:
                    pass
            proc.kill()
            proc.wait()
            if proc.stderr is not None:
                try:
                    stderr_text = proc.stderr.read()
                except Exception:
                    stderr_text = ""
            raise
        finally:
            if stdout_handle is not None:
                stdout_handle.close()
                stdout_handle = None

        if return_code != 0:
            raise RuntimeError(
                "Command failed while writing '{}': {}".format(
                    output_path,
                    stderr_text.strip() or "exit code {}".format(return_code),
                )
            )
        if not tmp_output.exists() or tmp_output.stat().st_size == 0:
            raise RuntimeError("Command produced no output for '{}'".format(output_path))
        tmp_output.replace(output_path)
    finally:
        if proc is not None and proc.stderr is not None:
            proc.stderr.close()
        if stdout_handle is not None:
            stdout_handle.close()
        if tmp_output.exists():
            tmp_output.unlink()


def write_fasta_records_gzip(output_path, records):
    seqkit_path = shutil.which("seqkit")
    if seqkit_path is None:
        with open_text(output_path, "wt") as handle:
            for record_id, sequence in records:
                write_fasta_record(handle, record_id, sequence)
        return

    def writer(handle):
        for record_id, sequence in records:
            write_fasta_record(handle, record_id, sequence)

    write_text_output_via_command(
        output_path,
        writer,
        lambda tmp_output: [
            seqkit_path,
            "seq",
            "--threads",
            str(output_compression_threads()),
            "-o",
            str(tmp_output),
            "-",
        ],
        output_via_stdout=False,
    )


def write_gff_gzip(input_path, output_path):
    pigz_path = shutil.which("pigz")
    line_count = 0

    if pigz_path is None:
        with open_text(input_path, "rt") as fin, open_text(output_path, "wt") as fout:
            for line in fin:
                fout.write(apply_common_replacements(line))
                line_count += 1
        return line_count

    def writer(handle):
        nonlocal line_count
        with open_text(input_path, "rt") as fin:
            for line in fin:
                handle.write(apply_common_replacements(line))
                line_count += 1

    write_text_output_via_command(
        output_path,
        writer,
        lambda _tmp_output: [
            pigz_path,
            "-p",
            str(output_compression_threads()),
            "-c",
        ],
        output_via_stdout=True,
    )
    return line_count


def write_gff_lines_gzip(output_path, lines):
    pigz_path = shutil.which("pigz")
    line_count = 0

    if pigz_path is None:
        with open_text(output_path, "wt") as fout:
            for line in lines:
                fout.write(apply_common_replacements(line))
                line_count += 1
        return line_count

    def writer(handle):
        nonlocal line_count
        for line in lines:
            handle.write(apply_common_replacements(line))
            line_count += 1

    write_text_output_via_command(
        output_path,
        writer,
        lambda _tmp_output: [
            pigz_path,
            "-p",
            str(output_compression_threads()),
            "-c",
        ],
        output_via_stdout=True,
    )
    return line_count


def detect_manifest_delimiter(path):
    with open(path, "rt", encoding="utf-8") as handle:
        first_line = handle.readline()
    if "\t" in first_line:
        return "\t"
    return ","


def read_download_manifest_xlsx(path):
    if load_workbook is None:
        raise ValueError("openpyxl is required to read .xlsx download manifests.")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        row_iter = sheet.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if header_row is None:
            return []
        headers = []
        for value in header_row:
            text = str(value).strip() if value is not None else ""
            headers.append(text)
        while len(headers) > 0 and headers[-1] == "":
            headers.pop()
        if len(headers) == 0:
            return []
        if len(headers) < 2 or headers[0] != "provider" or headers[1] != "id":
            raise ValueError(
                "XLSX download manifest must have 'provider' and 'id' as the first two columns."
            )

        rows = []
        for raw_values in row_iter:
            values = list(raw_values or ())
            if len(values) < len(headers):
                values.extend([None] * (len(headers) - len(values)))
            row = {}
            has_nonempty = False
            for idx, key in enumerate(headers):
                if key == "":
                    continue
                cell = values[idx] if idx < len(values) else None
                text = str(cell).strip() if cell is not None else ""
                if text != "":
                    has_nonempty = True
                row[key] = text
            if has_nonempty:
                rows.append(row)
        return rows
    finally:
        workbook.close()


def resolved_manifest_fieldnames(rows):
    discovered = []
    for row in rows:
        for key in row.keys():
            if key not in discovered:
                discovered.append(key)
    ordered = []
    for key in RESOLVED_MANIFEST_PREFERRED_COLUMNS:
        ordered.append(key)
    for key in discovered:
        if key not in ordered:
            ordered.append(key)
    return ordered


def build_resolved_manifest_row(
    raw_row,
    fieldnames,
    provider,
    source_id,
    species_key,
    cds_url,
    gff_url,
    gbff_url,
    genome_url,
    cds_archive_member,
    gff_archive_member,
    gbff_archive_member,
    genome_archive_member,
    cds_filename,
    gff_filename,
    gbff_filename,
    genome_filename,
):
    row = {}
    for key in fieldnames:
        row[key] = str(raw_row.get(key) or "")
    row["provider"] = provider
    row["id"] = source_id
    row["species_key"] = species_key
    row["cds_url"] = cds_url
    row["gff_url"] = gff_url
    row["gbff_url"] = gbff_url
    row["genome_url"] = genome_url
    row["cds_archive_member"] = cds_archive_member
    row["gff_archive_member"] = gff_archive_member
    row["gbff_archive_member"] = gbff_archive_member
    row["genome_archive_member"] = genome_archive_member
    row["cds_filename"] = cds_filename
    row["gff_filename"] = gff_filename
    row["gbff_filename"] = gbff_filename
    row["genome_filename"] = genome_filename
    return row


def write_resolved_manifest_tsv(output_path, fieldnames, rows):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "wt", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: str(row.get(key) or "") for key in fieldnames})


def provider_raw_dir(provider, download_root, species_key):
    if provider in ("ensembl", "ensemblplants"):
        return download_root / DEFAULT_INPUT_RELATIVE_DIRS[provider]
    if provider in (
        "phycocosm",
        "phytozome",
        "ncbi",
        "refseq",
        "genbank",
        "coge",
        "cngb",
        "gwh",
        "flybase",
        "wormbase",
        "vectorbase",
        "fernbase",
        "veupathdb",
        "dictybase",
        "insectbase",
        "direct",
        "local",
    ):
        return download_root / DEFAULT_INPUT_RELATIVE_DIRS[provider] / species_key
    raise ValueError("Unknown provider: {}".format(provider))


def infer_provider_from_id(source_id):
    if source_id == "":
        return ""
    lowered = source_id.lower()
    if NCBI_ASSEMBLY_ACCESSION_PATTERN.match(source_id):
        return "ncbi"
    if "ncbi.nlm.nih.gov/datasets/genome/" in lowered:
        return "ncbi"
    if ENSEMBL_ID_HINT_PATTERN.match(source_id):
        return "ensembl"
    if FERNBASE_ID_HINT_PATTERN.match(source_id):
        return "fernbase"
    if VEUPATHDB_ID_HINT_PATTERN.match(source_id):
        return "veupathdb"
    if DICTYBASE_ID_HINT_PATTERN.match(source_id):
        return "dictybase"
    if INSECTBASE_ID_HINT_PATTERN.match(source_id):
        return "insectbase"
    if INSECTBASE_IBG_ID_PATTERN.match(source_id):
        return "insectbase"
    if "ftp.ensembl.org" in lowered or "ensembl.org/pub/current_" in lowered:
        return "ensembl"
    if COGE_ID_HINT_PATTERN.match(source_id):
        return "coge"
    if GWH_ID_HINT_PATTERN.match(source_id):
        return "gwh"
    if CNGB_ID_HINT_PATTERN.match(source_id):
        return "cngb"
    if GWH_ASSEMBLY_ACCESSION_PATTERN.match(source_id):
        return "gwh"
    if CNGB_ASSEMBLY_ACCESSION_PATTERN.match(source_id):
        return "cngb"
    if "genomevolution.org" in lowered:
        return "coge"
    if "/gwh/" in lowered:
        return "gwh"
    if "cngb.org" in lowered or "cncb.ac.cn" in lowered:
        return "cngb"
    if "flybase.org" in lowered:
        return "flybase"
    if "wormbase.org" in lowered:
        return "wormbase"
    if "vectorbase.org" in lowered:
        return "vectorbase"
    if "fernbase.org" in lowered:
        return "fernbase"
    if "veupathdb.org" in lowered:
        return "veupathdb"
    if "dictybase.org" in lowered:
        return "dictybase"
    if "insect-genome.com" in lowered:
        return "insectbase"
    return ""


def strip_provider_prefix(source_id, provider):
    text = str(source_id or "").strip()
    prefix = provider.lower() + ":"
    if text.lower().startswith(prefix):
        return text[len(prefix) :]
    return text


def extract_ncbi_accession_from_source_id(source_id):
    text = str(source_id or "").strip()
    if NCBI_ASSEMBLY_ACCESSION_PATTERN.match(text):
        return text
    match = re.search(r"(GC[AF]_[0-9]+\.[0-9]+)", text, flags=re.IGNORECASE)
    if match is None:
        return ""
    return match.group(1)


def provider_env_prefix(provider):
    normalized = re.sub(r"[^A-Za-z0-9]+", "_", provider.upper())
    return "GG_{}".format(normalized)


def is_url_like(text):
    parsed = urlparse(str(text or ""))
    return parsed.scheme in ("http", "https", "ftp", "file")


def render_id_url_template(template, provider, source_id, species_key):
    source_clean = str(source_id or "").strip()
    id_value = strip_provider_prefix(source_clean, provider)
    if id_value == "":
        id_value = source_clean
    mapping = {
        "id": id_value,
        "id_raw": source_clean,
        "id_lower": id_value.lower(),
        "species_key": species_key,
        "provider": provider,
    }
    return template.format(**mapping)


def fetch_text_with_headers(url, timeout, headers):
    req_headers = dict(headers)
    if "User-Agent" not in req_headers:
        req_headers["User-Agent"] = "genegalleon-input-generation"
    request = Request(url, headers=req_headers)
    with urlopen(request, timeout=timeout) as response:
        payload = response.read()
    try:
        return payload.decode("utf-8")
    except UnicodeDecodeError:
        return payload.decode("utf-8", errors="replace")


def parse_links_from_document(base_url, text):
    links = []
    seen = set()
    for href in re.findall(r"""href\s*=\s*["']([^"']+)["']""", text, flags=re.IGNORECASE):
        candidate = href.strip()
        if candidate == "" or candidate.startswith("javascript:") or candidate.startswith("#"):
            continue
        absolute = urljoin(base_url, candidate)
        if absolute in seen:
            continue
        seen.add(absolute)
        links.append(absolute)
    for raw in re.findall(r"""https?://[^\s"'<>]+""", text):
        candidate = raw.strip()
        if candidate in seen:
            continue
        seen.add(candidate)
        links.append(candidate)
    return links


FERNBASE_GFF_EXCLUDE_PATTERN = re.compile(
    r"(?:^|[._-])(te|teanno|repeat|repeats|transpos(?:on)?)(?:[._-]|$)",
    re.IGNORECASE,
)


def is_probable_cds_filename(provider, name):
    lower = name.lower()
    if provider == "fernbase" and (lower.endswith(".cds") or lower.endswith(".cds.gz")):
        return True
    if not is_fasta_filename(name):
        return False
    return any(marker in lower for marker in ("cds", "transcript", "mrna", "cdna", "genecatalog_cds"))


def is_probable_cds_url(provider, url):
    lower = url.lower()
    return is_probable_cds_filename(provider, lower)


def is_probable_gff_url(url):
    return is_gff_filename(url.lower())


def is_probable_genome_url(provider, url):
    return is_probable_genome_filename(provider, urlparse(url).path.lower())


def provider_candidate_sort_key(provider, label, name):
    lower = str(name or "").lower()
    label_upper = str(label or "").upper()
    if provider != "fernbase":
        if label_upper == "GENOME":
            return (".chromosome." in lower, lower)
        return (lower,)
    if label_upper == "CDS":
        return (
            0 if "highconfidence" in lower else 1,
            1 if "lowconfidence" in lower else 0,
            0 if "cds" in lower else 1,
            1 if any(marker in lower for marker in ("transcript", "mrna", "cdna")) else 0,
            lower,
        )
    if label_upper == "GFF":
        return (
            1 if FERNBASE_GFF_EXCLUDE_PATTERN.search(lower) else 0,
            0 if "highconfidence" in lower else 1,
            1 if "lowconfidence" in lower else 0,
            lower,
        )
    if label_upper == "GENOME":
        return (
            1 if "chloroplast" in lower else 0,
            0 if any(marker in lower for marker in ("genome", "assembly", "chr", "chromosome")) else 1,
            lower,
        )
    return (lower,)


def select_best_url_for_label(provider, label, candidates):
    if len(candidates) == 0:
        return ""
    filtered = []
    for candidate in candidates:
        path_lower = urlparse(candidate).path.lower()
        if label == "CDS":
            if is_probable_cds_url(provider, path_lower):
                filtered.append(candidate)
            continue
        if label == "GFF":
            if is_probable_gff_url(path_lower):
                filtered.append(candidate)
            continue
        if label == "GENOME":
            if is_probable_genome_url(provider, candidate):
                filtered.append(candidate)
            continue
    if len(filtered) == 0:
        return ""
    preferred = sorted(
        filtered,
        key=lambda x: provider_candidate_sort_key(provider, label, urlparse(x).path.split("/")[-1]),
    )
    return preferred[0]


def resolve_urls_from_index_url(provider, index_url, timeout, headers):
    text = fetch_text_with_headers(index_url, timeout, headers)
    links = parse_links_from_document(index_url, text)
    return {
        "cds_url": select_best_url_for_label(provider, "CDS", links),
        "gff_url": select_best_url_for_label(provider, "GFF", links),
        "genome_url": select_best_url_for_label(provider, "GENOME", links),
    }


def fetch_json_with_headers(url, timeout, headers):
    text = fetch_text_with_headers(url, timeout, headers)
    return json.loads(text)


def resolve_coge_api_base_url():
    return os.environ.get("GG_COGE_API_BASE_URL", DEFAULT_COGE_API_BASE_URL).rstrip("/")


def resolve_coge_web_base_url():
    return os.environ.get("GG_COGE_WEB_BASE_URL", DEFAULT_COGE_WEB_BASE_URL).rstrip("/")


def resolve_cngb_cnsa_base_url():
    return os.environ.get("GG_CNGB_CNSA_BASE_URL", DEFAULT_CNGB_CNSA_BASE_URL).rstrip("/")


def resolve_gwh_download_base_url():
    return os.environ.get("GG_GWH_DOWNLOAD_BASE_URL", DEFAULT_GWH_DOWNLOAD_BASE_URL).rstrip("/")


def resolve_veupathdb_service_base_url():
    return os.environ.get("GG_VEUPATHDB_SERVICE_BASE_URL", DEFAULT_VEUPATHDB_SERVICE_BASE_URL).rstrip("/")


def resolve_insectbase_api_base_url():
    return os.environ.get("GG_INSECTBASE_API_BASE_URL", DEFAULT_INSECTBASE_API_BASE_URL).rstrip("/")


def parse_species_key_candidate(text):
    cleaned = re.sub(r"\s*\(.*?\)\s*", " ", str(text or "")).strip()
    tokens = re.findall(r"[A-Za-z0-9]+", cleaned)
    if len(tokens) >= 2:
        return "{}_{}".format(tokens[0], tokens[1])
    if len(tokens) == 1:
        return tokens[0]
    return ""


def source_id_candidates(provider, source_id, species_key):
    candidates = []

    def add(value):
        text = str(value or "").strip()
        if text == "":
            return
        if text in candidates:
            return
        candidates.append(text)

    source_clean = str(source_id or "").strip()
    stripped = strip_provider_prefix(source_clean, provider)
    add(source_clean)
    add(stripped)

    if "_" in stripped:
        add(stripped.rsplit("_", 1)[-1])
    if species_key != "":
        add(species_key)
        if "_" in species_key:
            add(species_key.rsplit("_", 1)[-1])

    if is_url_like(stripped):
        parsed = urlparse(stripped)
        query = parse_qs(parsed.query)
        if provider == "coge":
            for gid in query.get("gid", []):
                add(gid)
        if provider == "cngb":
            for qval in query.get("q", []):
                add(qval)
            parts = [token for token in parsed.path.split("/") if token != ""]
            for idx, token in enumerate(parts[:-1]):
                if token.lower() == "assembly":
                    add(parts[idx + 1])
        if provider == "gwh":
            for qvals in query.values():
                for qval in qvals:
                    for match in GWH_ASSEMBLY_ACCESSION_SEARCH_PATTERN.findall(qval):
                        add(match)
            for part in parsed.path.split("/"):
                for match in GWH_ASSEMBLY_ACCESSION_SEARCH_PATTERN.findall(unquote(part)):
                    add(match)

    if provider == "phytozome":
        match = re.search(r"_([0-9]+)_", stripped)
        if match is not None:
            add(match.group(1))
    if provider == "gwh":
        for match in GWH_ASSEMBLY_ACCESSION_SEARCH_PATTERN.findall(stripped):
            add(match)

    return candidates


def normalize_manifest_source_id(provider, source_id):
    text = str(source_id or "").strip()
    if text == "":
        return ""
    if provider in ("local", "direct"):
        return text
    if is_url_like(text):
        return text
    match = re.match(r"^(\S+)\s+\(.*\)$", text)
    if match is not None:
        return match.group(1)
    return text


def extract_insectbase_ibg_id_candidate(source_id):
    text = str(source_id or "").strip()
    if text == "":
        return ""
    stripped = strip_provider_prefix(text, "insectbase")
    for candidate in (text, stripped):
        if INSECTBASE_IBG_ID_PATTERN.match(candidate):
            return candidate.upper()
        match = re.search(r"(IBG_[0-9]+)", candidate, flags=re.IGNORECASE)
        if match is not None:
            return match.group(1).upper()
    if is_url_like(stripped):
        parts = [unquote(part) for part in urlparse(stripped).path.split("/") if part]
        for part in reversed(parts):
            if INSECTBASE_IBG_ID_PATTERN.match(part):
                return part.upper()
    return ""


def extract_coge_gid_candidate(source_id):
    candidates = source_id_candidates("coge", source_id, species_key="")
    for candidate in candidates:
        stripped = strip_provider_prefix(candidate, "coge")
        if COGE_GID_PATTERN.match(stripped):
            return stripped
    return ""


def normalize_lookup_text(text):
    return re.sub(r"[^a-z0-9]+", "", str(text or "").lower())


def pick_best_coge_genome(genomes, source_id):
    query = strip_provider_prefix(source_id, "coge")
    query_normalized = normalize_lookup_text(query)
    selected = None
    selected_rank = None
    for genome in genomes:
        if not isinstance(genome, dict):
            continue
        gid = str(genome.get("id", "") or "").strip()
        if not gid.isdigit():
            continue
        name = str(genome.get("name", "") or "")
        info = str(genome.get("info", "") or "")
        organism = str((genome.get("organism") or {}).get("name", "") or "")
        searchable = [name, info, organism]
        exact = any(query_normalized != "" and normalize_lookup_text(text) == query_normalized for text in searchable)
        contains = any(query_normalized != "" and query_normalized in normalize_lookup_text(text) for text in searchable)
        deleted = bool(genome.get("deleted"))
        certified = bool(genome.get("certified"))
        rank = (
            0 if deleted else 1,
            1 if exact else 0,
            1 if contains else 0,
            1 if certified else 0,
            int(gid),
        )
        if selected is None or rank > selected_rank:
            selected = genome
            selected_rank = rank
    return selected


def resolve_coge_genome_from_source_id(source_id, timeout, headers):
    gid = extract_coge_gid_candidate(source_id)
    if gid == "":
        raise ValueError("CoGe id must be genome_id (numeric gid). got '{}'".format(source_id))
    return gid, {"id": int(gid)}


def resolve_coge_download_urls_from_id(source_id, species_key, timeout, headers):
    gid, genome_info = resolve_coge_genome_from_source_id(source_id, timeout, headers)
    web_base = resolve_coge_web_base_url()
    api_base = resolve_coge_api_base_url()

    gff_meta_url = (
        "{}/GenomeInfo.pl?fname=get_gff&gid={}&id_type=name&cds=0&annos=0&nu=0&upa=0&chr=".format(web_base, gid)
    )
    gff_meta = fetch_json_with_headers(gff_meta_url, timeout, headers)
    gff_candidates = [str(item).strip() for item in gff_meta.get("files", []) if str(item).strip() != ""]
    gff_url = gff_candidates[0] if len(gff_candidates) > 0 else ""
    gff_filename = str(gff_meta.get("file", "") or "").strip()
    if gff_url == "" and gff_filename != "":
        gff_url = "{}/api/v1/downloads/?gid={}&filename={}".format(
            web_base, gid, quote(gff_filename, safe="")
        )
    if gff_url == "":
        raise ValueError("CoGe did not return a downloadable GFF URL for gid {}".format(gid))

    organism_name = str((genome_info.get("organism") or {}).get("name", "") or "")
    inferred_species = parse_species_key_candidate(organism_name)
    if inferred_species == "" and gff_filename != "":
        prefix = re.sub(r"[.]gid[0-9]+.*$", "", gff_filename, flags=re.IGNORECASE)
        inferred_species = parse_species_key_candidate(prefix.replace("_", " "))
    if inferred_species == "":
        inferred_species = species_key
    if inferred_species == "":
        inferred_species = "CoGe_gid{}".format(gid)
    prefix = sanitize_identifier(inferred_species)

    if gff_filename == "":
        gff_filename = "{}.coge.gid{}.gff".format(prefix, gid)
    elif not is_gff_filename(gff_filename):
        gff_filename = "{}.coge.gid{}.gff".format(prefix, gid)

    return {
        "species_key": inferred_species,
        "cds_url": "{}/get_seqs_for_feattype_for_genome.pl?ftid=3;dsgid={};".format(web_base, gid),
        "gff_url": gff_url,
        "genome_url": "{}/genomes/{}/sequence".format(api_base, gid),
        "cds_filename": "{}.coge.gid{}.cds.fa".format(prefix, gid),
        "gff_filename": gff_filename,
        "genome_filename": "{}.coge.gid{}.genome.fa".format(prefix, gid),
    }


def resolve_cngb_summary_from_id(cngb_id, timeout, headers):
    base = resolve_cngb_cnsa_base_url()
    summary_url = "{}/assembly/public_view/?q={}".format(base, quote(cngb_id, safe=""))
    payload = fetch_json_with_headers(summary_url, timeout, headers)
    if payload.get("code") != 0:
        error_message = str((payload.get("error") or {}).get("content", "") or "").strip()
        if error_message == "":
            error_message = "unknown CNGB API error"
        raise ValueError("CNGB summary lookup failed for '{}': {}".format(cngb_id, error_message))
    summary = payload.get("data", {}).get("summary_data", {})
    if not isinstance(summary, dict):
        raise ValueError("CNGB summary response was malformed for '{}'".format(cngb_id))
    return summary


def resolve_cngb_download_urls_from_id(source_id, timeout, headers):
    raw_candidates = source_id_candidates("cngb", source_id, species_key="")
    cngb_id = strip_provider_prefix(source_id, "cngb")
    if cngb_id == "":
        cngb_id = source_id

    accession = ""
    for candidate in raw_candidates:
        accession = extract_ncbi_accession_from_source_id(candidate)
        if accession != "":
            break

    summary = {}
    if accession == "":
        summary = resolve_cngb_summary_from_id(cngb_id, timeout, headers)
        for key in (
            "refseq_assembly_accession",
            "genbank_assembly_accession",
            "accession_id",
            "external_accession_id",
        ):
            accession = extract_ncbi_accession_from_source_id(summary.get(key, ""))
            if accession != "":
                break

    if accession == "":
        raise ValueError(
            "CNGB id '{}' did not map to a downloadable NCBI assembly accession".format(source_id)
        )

    resolved = resolve_ncbi_download_urls_from_id(accession, timeout, ncbi_source="auto")
    if resolved.get("species_key", "") == "":
        organism_name = str((summary.get("organism") or {}).get("name", "") or "")
        species_candidate = parse_species_key_candidate(organism_name)
        if species_candidate != "":
            resolved["species_key"] = species_candidate
    return resolved


_veupathdb_records_cache = {}
_veupathdb_records_lock = threading.Lock()


def veupathdb_source_id_from_url(url):
    path = urlparse(str(url or "")).path
    parts = [unquote(part) for part in path.split("/") if part]
    for idx, token in enumerate(parts[:-1]):
        if token == "Current_Release" and idx + 1 < len(parts):
            return parts[idx + 1]
    return ""


def infer_veupathdb_cds_url(attributes):
    protein_url = str(attributes.get("URLproteinFasta", "") or "").strip()
    if protein_url != "":
        return re.sub(
            r"_AnnotatedProteins[.]fasta$",
            "_AnnotatedCDSs.fasta",
            protein_url,
            flags=re.IGNORECASE,
        )
    genome_url = str(attributes.get("URLGenomeFasta", "") or "").strip()
    if genome_url == "":
        return ""
    parsed = urlparse(genome_url)
    directory = parsed.path.rsplit("/", 1)[0]
    basename = parsed.path.rsplit("/", 1)[-1]
    stem = re.sub(r"_Genome[.]fasta$", "", basename, flags=re.IGNORECASE)
    if stem == basename:
        return ""
    return parsed._replace(path="{}/{}_AnnotatedCDSs.fasta".format(directory, stem)).geturl()


def fetch_veupathdb_records(timeout, headers):
    service_base = resolve_veupathdb_service_base_url()
    with _veupathdb_records_lock:
        cached = _veupathdb_records_cache.get(service_base)
    if cached is not None:
        return cached
    config = {
        "attributes": ["primary_key", "species", "URLGenomeFasta", "URLgff", "URLproteinFasta", "project_id"],
        "tables": [],
        "attributeFormat": "text",
    }
    service_url = (
        "{}/record-types/organism/searches/GenomeDataTypes/reports/standard?reportConfig={}".format(
            service_base,
            quote(json.dumps(config, separators=(",", ":")), safe=""),
        )
    )
    payload = fetch_json_with_headers(service_url, timeout, headers)
    records = list(payload.get("records", []))
    with _veupathdb_records_lock:
        _veupathdb_records_cache[service_base] = records
    return records


def resolve_veupathdb_download_urls_from_id(source_id, species_key, timeout, headers):
    candidates = source_id_candidates("veupathdb", source_id, species_key)
    if len(candidates) == 0:
        candidates = [str(source_id or "").strip()]
    candidate_keys = [normalize_lookup_text(value) for value in candidates if normalize_lookup_text(value) != ""]
    if len(candidate_keys) == 0:
        raise ValueError("VEuPathDB id is empty")

    best_record = None
    best_rank = None
    for record in fetch_veupathdb_records(timeout, headers):
        attributes = record.get("attributes", {})
        genome_url = str(attributes.get("URLGenomeFasta", "") or "").strip()
        gff_url = str(attributes.get("URLgff", "") or "").strip()
        cds_url = infer_veupathdb_cds_url(attributes)
        derived_source_id = veupathdb_source_id_from_url(genome_url) or veupathdb_source_id_from_url(gff_url)
        normalized_source_id = normalize_lookup_text(derived_source_id)
        normalized_primary = normalize_lookup_text(attributes.get("primary_key", ""))
        normalized_species = normalize_lookup_text(attributes.get("species", ""))
        normalized_project = normalize_lookup_text(attributes.get("project_id", ""))
        source_id_exact = any(candidate == normalized_source_id for candidate in candidate_keys if normalized_source_id != "")
        primary_exact = any(candidate == normalized_primary for candidate in candidate_keys if normalized_primary != "")
        species_exact = any(candidate == normalized_species for candidate in candidate_keys if normalized_species != "")
        project_exact = any(candidate == normalized_project for candidate in candidate_keys if normalized_project != "")
        source_id_contains = any(
            candidate in normalized_source_id or normalized_source_id in candidate
            for candidate in candidate_keys
            if normalized_source_id != ""
        )
        primary_contains = any(
            candidate in normalized_primary or normalized_primary in candidate
            for candidate in candidate_keys
            if normalized_primary != ""
        )
        species_contains = any(
            candidate in normalized_species or normalized_species in candidate
            for candidate in candidate_keys
            if normalized_species != ""
        )
        rank = (
            1 if genome_url != "" else 0,
            1 if gff_url != "" else 0,
            1 if cds_url != "" else 0,
            1 if source_id_exact else 0,
            1 if primary_exact else 0,
            1 if species_exact else 0,
            1 if project_exact else 0,
            1 if source_id_contains else 0,
            1 if primary_contains else 0,
            1 if species_contains else 0,
            -len(normalized_source_id),
            normalized_source_id,
        )
        if best_record is None or rank > best_rank:
            best_record = {
                "record": record,
                "derived_source_id": derived_source_id,
                "cds_url": cds_url,
            }
            best_rank = rank

    if best_record is None or max(best_rank[3:10]) == 0:
        raise ValueError("VEuPathDB id '{}' was not found in GenomeDataTypes".format(source_id))

    attributes = best_record["record"].get("attributes", {})
    genome_url = str(attributes.get("URLGenomeFasta", "") or "").strip()
    gff_url = str(attributes.get("URLgff", "") or "").strip()
    cds_url = str(best_record.get("cds_url", "") or "").strip()
    if gff_url == "" or genome_url == "":
        raise ValueError("VEuPathDB id '{}' did not resolve to GFF/genome URLs".format(source_id))

    inferred_species_key = parse_species_key_candidate(attributes.get("species", ""))
    if inferred_species_key == "":
        inferred_species_key = str(species_key or "").strip()
    if inferred_species_key == "":
        inferred_species_key = sanitize_identifier(best_record.get("derived_source_id", "") or str(source_id or ""))
    prefix = sanitize_identifier(inferred_species_key)
    source_token = best_record.get("derived_source_id", "") or strip_provider_prefix(source_id, "veupathdb")
    source_token = sanitize_identifier(source_token)

    return {
        "species_key": inferred_species_key,
        "cds_url": cds_url,
        "gff_url": gff_url,
        "genome_url": genome_url,
        "cds_filename": "{}.veupathdb.{}.cds.fa".format(prefix, source_token),
        "gff_filename": "{}.veupathdb.{}.gene.gff3".format(prefix, source_token),
        "genome_filename": "{}.veupathdb.{}.genome.fa".format(prefix, source_token),
    }


def resolve_insectbase_download_urls_from_id(source_id, species_key, timeout, headers):
    ibg_id = extract_insectbase_ibg_id_candidate(source_id)
    if ibg_id == "":
        raise ValueError(
            "InsectBase id '{}' did not contain an IBG_* genome identifier".format(source_id)
        )

    api_base = resolve_insectbase_api_base_url()
    detail_url = "{}/genomes/{}/".format(api_base, quote(ibg_id, safe=""))
    record = fetch_json_with_headers(detail_url, timeout, headers)
    species_name = str(record.get("species", "") or "").strip()
    if species_name == "":
        raise ValueError("InsectBase id '{}' did not resolve to a species name".format(source_id))

    species_token = re.sub(r"\s+", "_", species_name)
    inferred_species_key = str(species_key or "").strip()
    if inferred_species_key == "":
        inferred_species_key = parse_species_key_candidate(species_name)
    if inferred_species_key == "":
        inferred_species_key = species_token

    api_parts = urlparse(api_base)
    site_root = "{}://{}".format(api_parts.scheme, api_parts.netloc)
    base_data_url = "{}/data/genome/{}/{}".format(
        site_root.rstrip("/"),
        quote(species_token, safe="._-"),
        quote(species_token, safe="._-"),
    )

    return {
        "species_key": inferred_species_key,
        "cds_url": base_data_url + ".cds.fa",
        "gff_url": base_data_url + ".gff3",
        "genome_url": base_data_url + ".genome.fa.tar.bz2",
        "cds_filename": species_token + ".cds.fa",
        "gff_filename": species_token + ".gff3",
        "genome_filename": species_token + ".genome.fa.tar.bz2",
    }


def extract_gwh_accession_candidate(source_id):
    candidates = source_id_candidates("gwh", source_id, species_key="")
    for candidate in candidates:
        stripped = strip_provider_prefix(candidate, "gwh")
        match = GWH_ASSEMBLY_ACCESSION_SEARCH_PATTERN.search(stripped)
        if match is None:
            continue
        accession = match.group(1).upper()
        if GWH_ASSEMBLY_ACCESSION_PATTERN.match(accession):
            return accession
    return ""


def gwh_accession_stem(accession):
    return str(accession or "").strip().upper().split(".", 1)[0]


def parse_last_path_token(url):
    path = urlparse(str(url or "")).path.rstrip("/")
    if path == "":
        return ""
    return unquote(path.split("/")[-1])


def resolve_gwh_folder_url_from_id(source_id, timeout, headers):
    source_clean = str(source_id or "").strip()
    accession = extract_gwh_accession_candidate(source_clean)
    if accession == "":
        raise ValueError("GWH id '{}' did not contain a GWH accession".format(source_id))

    if is_url_like(source_clean) and "/gwh/" in source_clean.lower():
        parsed = urlparse(source_clean)
        if parsed.path.endswith("/"):
            return source_clean, accession
        return parsed._replace(path=parsed.path.rsplit("/", 1)[0] + "/").geturl(), accession

    root_url = resolve_gwh_download_base_url().rstrip("/") + "/"
    root_text = fetch_text_with_headers(root_url, timeout, headers)
    category_urls = []
    for link in parse_links_from_document(root_url, root_text):
        if not urlparse(link).path.endswith("/"):
            continue
        token = parse_last_path_token(link)
        if token == "":
            continue
        category_urls.append(link)
    if len(category_urls) == 0:
        raise ValueError("GWH root '{}' did not expose category directories".format(root_url))

    accession_stem = gwh_accession_stem(accession)
    best_url = ""
    best_rank = None
    last_error = None
    for category_url in sorted(
        set(category_urls),
        key=lambda url: (0 if parse_last_path_token(url).lower() == "plants" else 1, parse_last_path_token(url).lower()),
    ):
        try:
            category_text = fetch_text_with_headers(category_url, timeout, headers)
        except Exception as exc:
            last_error = exc
            continue
        for link in parse_links_from_document(category_url, category_text):
            if not urlparse(link).path.endswith("/"):
                continue
            folder_name = parse_last_path_token(link)
            if folder_name == "":
                continue
            folder_lower = folder_name.lower()
            has_full = accession.lower() in folder_lower
            has_stem = accession_stem.lower() in folder_lower
            if not has_full and not has_stem:
                continue
            rank = (
                1 if has_full else 0,
                1 if folder_lower.endswith(accession.lower()) else 0,
                1 if parse_last_path_token(category_url).lower() == "plants" else 0,
                -len(folder_name),
                folder_lower,
            )
            if best_url == "" or rank > best_rank:
                best_url = link
                best_rank = rank
    if best_url != "":
        return best_url, accession
    if last_error is not None:
        raise ValueError("GWH accession '{}' did not resolve: {}".format(accession, last_error))
    raise ValueError("GWH accession '{}' was not found under {}".format(accession, root_url))


def resolve_gwh_download_urls_from_id(source_id, species_key, timeout, headers):
    folder_url, accession = resolve_gwh_folder_url_from_id(source_id, timeout, headers)
    resolved = resolve_urls_from_index_url("gwh", folder_url, timeout, headers)
    if resolved.get("gff_url", "") == "" or resolved.get("genome_url", "") == "":
        raise ValueError(
            "GWH accession '{}' did not resolve to downloadable GFF/genome URLs from {}".format(
                accession, folder_url
            )
        )

    folder_name = parse_last_path_token(folder_url)
    inferred_species_key = str(species_key or "").strip()
    if inferred_species_key == "":
        inferred_species_key = parse_species_key_candidate(folder_name.replace("_", " "))
    if inferred_species_key == "":
        inferred_species_key = sanitize_identifier(accession)
    return {
        "species_key": inferred_species_key,
        "cds_url": resolved["cds_url"],
        "gff_url": resolved["gff_url"],
        "genome_url": resolved["genome_url"],
        "cds_filename": Path(urlparse(resolved["cds_url"]).path).name,
        "gff_filename": Path(urlparse(resolved["gff_url"]).path).name,
        "genome_filename": Path(urlparse(resolved["genome_url"]).path).name,
    }


def fernbase_release_sort_key(name):
    lower = str(name or "").lower()
    version_tokens = tuple(int(token) for token in re.findall(r"[0-9]+", lower))
    has_version = 1 if re.search(r"(?:^|[_-])(?:asm[_-])?v?[0-9]+", lower) else 0
    return (has_version, version_tokens, lower)


def infer_fernbase_species_key(source_id, species_key):
    explicit = str(species_key or "").strip()
    if explicit != "":
        return explicit
    raw = strip_provider_prefix(source_id, "fernbase")
    if raw == "":
        raw = str(source_id or "").strip()
    if is_url_like(raw):
        parts = [unquote(part) for part in urlparse(raw).path.split("/") if part]
        if "ftp" in parts:
            ftp_index = parts.index("ftp")
            parts = parts[ftp_index + 1 :]
        if len(parts) > 0:
            return parts[0]
        return ""
    return raw.split("/", 1)[0]


def resolve_fernbase_download_urls_from_id(source_id, species_key, timeout, headers):
    source_clean = str(source_id or "").strip()
    root_template = os.environ.get("GG_FERNBASE_ID_URL_TEMPLATE", "").strip()
    if root_template == "":
        root_template = "https://fernbase.org/ftp/{id}/"

    index_urls = []
    if is_url_like(source_clean):
        index_urls.append(source_clean)
    id_candidates = source_id_candidates("fernbase", source_id, species_key)
    if len(id_candidates) == 0:
        id_candidates = [source_clean]
    for candidate in id_candidates:
        rendered = render_id_url_template(root_template, "fernbase", candidate, species_key)
        if rendered != "":
            index_urls.append(rendered)

    deduped_index_urls = []
    seen_index_urls = set()
    for index_url in index_urls:
        normalized = str(index_url or "").strip()
        if normalized == "":
            continue
        if normalized in seen_index_urls:
            continue
        seen_index_urls.add(normalized)
        deduped_index_urls.append(normalized)

    inferred_species_key = infer_fernbase_species_key(source_id, species_key)
    last_error = None
    for index_url in deduped_index_urls:
        try:
            resolved = resolve_urls_from_index_url("fernbase", index_url, timeout, headers)
        except Exception as exc:
            last_error = exc
            continue
        if resolved.get("gff_url", "") != "" and resolved.get("genome_url", "") != "":
            if inferred_species_key != "":
                resolved["species_key"] = inferred_species_key
            return resolved

        try:
            index_text = fetch_text_with_headers(index_url, timeout, headers)
        except Exception as exc:
            last_error = exc
            continue

        subdir_candidates = []
        for link in parse_links_from_document(index_url, index_text):
            path = urlparse(link).path
            if not path.endswith("/"):
                continue
            subdir_name = unquote(path.rstrip("/").split("/")[-1])
            if subdir_name in ("", "..", ".", inferred_species_key, "chloroplast_genome"):
                continue
            if not re.fullmatch(r"[A-Za-z0-9_.-]+", subdir_name):
                continue
            subdir_candidates.append((fernbase_release_sort_key(subdir_name), link))

        for _sort_key, subdir_url in sorted(subdir_candidates, reverse=True):
            try:
                resolved = resolve_urls_from_index_url("fernbase", subdir_url, timeout, headers)
            except Exception as exc:
                last_error = exc
                continue
            if resolved.get("gff_url", "") != "" and resolved.get("genome_url", "") != "":
                if inferred_species_key != "":
                    resolved["species_key"] = inferred_species_key
                return resolved

    if last_error is not None:
        raise ValueError("FernBase id '{}' did not resolve: {}".format(source_id, last_error))
    raise ValueError("FernBase id '{}' did not resolve to downloadable GFF/genome URLs".format(source_id))


def resolve_provider_specific_download_urls_from_id(provider, source_id, species_key, timeout, headers):
    if provider == "coge":
        return resolve_coge_download_urls_from_id(source_id, species_key, timeout, headers)
    if provider == "cngb":
        return resolve_cngb_download_urls_from_id(source_id, timeout, headers)
    if provider == "gwh":
        return resolve_gwh_download_urls_from_id(source_id, species_key, timeout, headers)
    if provider == "fernbase":
        return resolve_fernbase_download_urls_from_id(source_id, species_key, timeout, headers)
    if provider == "veupathdb":
        return resolve_veupathdb_download_urls_from_id(source_id, species_key, timeout, headers)
    if provider == "insectbase":
        return resolve_insectbase_download_urls_from_id(source_id, species_key, timeout, headers)
    return None


def resolve_non_ncbi_download_urls_from_id(provider, source_id, species_key, timeout, headers):
    if provider not in DOWNLOAD_MANIFEST_SUPPORTED_PROVIDERS:
        raise ValueError(
            "provider '{}' is not supported for --download-manifest (use --input-dir for local formatting)".format(
                provider
            )
        )

    resolved = {}
    provider_specific_error = None

    env_prefix = provider_env_prefix(provider)
    label_to_key = {"CDS": "cds_url", "GFF": "gff_url", "GENOME": "genome_url"}
    id_candidates = source_id_candidates(provider, source_id, species_key)
    if len(id_candidates) == 0:
        id_candidates = [str(source_id or "").strip()]

    for label in DOWNLOAD_LABELS:
        template_env = "{}_{}_URL_TEMPLATE".format(env_prefix, label)
        template = os.environ.get(template_env, "").strip()
        if template == "":
            if provider == "ensembl":
                template = ENSEMBL_DEFAULT_ID_URL_TEMPLATES.get(label, "")
            if provider == "ensemblplants":
                template = ENSEMBLPLANTS_DEFAULT_ID_URL_TEMPLATES.get(label, "")
        if template == "":
            continue
        for candidate in id_candidates:
            url = render_id_url_template(template, provider, candidate, species_key)
            if url.endswith("/"):
                try:
                    discovered = resolve_urls_from_index_url(provider, url, timeout, headers)
                    inferred_url = discovered.get(label_to_key[label], "")
                    if inferred_url != "":
                        resolved[label_to_key[label]] = inferred_url
                        break
                except Exception:
                    continue
            else:
                resolved[label_to_key[label]] = url
                break

    if manifest_has_usable_source_bundle(
        resolved.get("cds_url", ""),
        resolved.get("gff_url", ""),
        resolved.get("gbff_url", ""),
        resolved.get("genome_url", ""),
    ):
        return resolved

    try:
        provider_specific = resolve_provider_specific_download_urls_from_id(
            provider=provider,
            source_id=source_id,
            species_key=species_key,
            timeout=timeout,
            headers=headers,
        )
    except Exception as exc:
        provider_specific = None
        provider_specific_error = exc
    if provider_specific is not None:
        for key in (
            "cds_url",
            "gff_url",
            "gbff_url",
            "genome_url",
            "cds_filename",
            "gff_filename",
            "gbff_filename",
            "genome_filename",
            "species_key",
        ):
            if resolved.get(key, "") == "" and str(provider_specific.get(key, "") or "").strip() != "":
                resolved[key] = str(provider_specific[key]).strip()
        if manifest_has_usable_source_bundle(
            resolved.get("cds_url", ""),
            resolved.get("gff_url", ""),
            resolved.get("gbff_url", ""),
            resolved.get("genome_url", ""),
        ):
            return resolved

    source_clean = str(source_id or "").strip()
    if is_url_like(source_clean):
        discovered = resolve_urls_from_index_url(provider, source_clean, timeout, headers)
        for key in ("cds_url", "gff_url", "gbff_url", "genome_url"):
            if resolved.get(key, "") == "" and discovered.get(key, "") != "":
                resolved[key] = discovered[key]

    id_page_template_env = "{}_ID_URL_TEMPLATE".format(env_prefix)
    id_page_template = os.environ.get(id_page_template_env, "").strip()
    page_templates = []
    if id_page_template != "":
        page_templates.append(id_page_template)
    page_templates.extend(PROVIDER_DEFAULT_ID_PAGE_URL_TEMPLATES.get(provider, ()))

    for page_template in page_templates:
        for candidate in id_candidates:
            page_url = render_id_url_template(page_template, provider, candidate, species_key)
            try:
                discovered = resolve_urls_from_index_url(provider, page_url, timeout, headers)
            except Exception:
                continue
            for key in ("cds_url", "gff_url", "gbff_url", "genome_url"):
                if resolved.get(key, "") == "" and discovered.get(key, "") != "":
                    resolved[key] = discovered[key]
            if manifest_has_usable_source_bundle(
                resolved.get("cds_url", ""),
                resolved.get("gff_url", ""),
                resolved.get("gbff_url", ""),
                resolved.get("genome_url", ""),
            ):
                break
        if manifest_has_usable_source_bundle(
            resolved.get("cds_url", ""),
            resolved.get("gff_url", ""),
            resolved.get("gbff_url", ""),
            resolved.get("genome_url", ""),
        ):
            break

    if not manifest_has_usable_source_bundle(
        resolved.get("cds_url", ""),
        resolved.get("gff_url", ""),
        resolved.get("gbff_url", ""),
        resolved.get("genome_url", ""),
    ):
        detail = ""
        if provider_specific_error is not None:
            detail = " (provider-specific resolver: {})".format(provider_specific_error)
        raise ValueError(
            "could not infer a usable source bundle (CDS, GBFF, or GFF plus genome) from id '{}' for provider '{}'{}".format(
                source_id, provider, detail
            )
        )
    return resolved


def default_download_filename(provider, species_key, label, url, archive_member=""):
    archive_member_text = str(archive_member or "").strip()
    if archive_member_text != "":
        base = Path(archive_member_text).name
    else:
        parsed = urlparse(url)
        base = Path(parsed.path).name
    if base == "":
        if label in ("cds", "genome"):
            ext = ".fa.gz"
        elif label == "gbff":
            ext = ".gbff.gz"
        else:
            ext = ".gff3.gz"
        base = "{}.{}{}".format(species_key, label, ext)
    if provider in ("ensembl", "ensemblplants"):
        if not base.startswith(species_key + "."):
            base = "{}.{}".format(species_key, base)
    return base


def parse_http_headers(http_header_values, auth_bearer_token_env):
    headers = {}
    for raw in http_header_values:
        text = raw.strip()
        if text == "":
            continue
        if ":" not in text:
            raise ValueError("Invalid --http-header value (missing ':'): {}".format(raw))
        key, value = text.split(":", 1)
        key = key.strip()
        value = value.strip()
        if key == "":
            raise ValueError("Invalid --http-header value (empty key): {}".format(raw))
        headers[key] = value
    if auth_bearer_token_env != "":
        token = os.environ.get(auth_bearer_token_env, "").strip()
        if token == "":
            raise ValueError(
                "Environment variable for bearer token is empty or undefined: {}".format(
                    auth_bearer_token_env
                )
            )
        headers["Authorization"] = "Bearer {}".format(token)
    return headers


def parse_positive_int(value, fallback):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return int(fallback)
    if parsed < 1:
        return int(fallback)
    return parsed


def resolve_parallel_jobs(requested_jobs):
    jobs = parse_positive_int(requested_jobs, 0)
    if jobs > 0:
        return jobs
    env_slots = os.environ.get("GG_TASK_CPUS", "").strip()
    if env_slots == "":
        env_slots = os.environ.get("NSLOTS", "").strip()
    if env_slots != "":
        return parse_positive_int(env_slots, DEFAULT_GLOBAL_DOWNLOAD_WORKERS)
    return DEFAULT_GLOBAL_DOWNLOAD_WORKERS


def resolve_provider_download_limits(global_jobs):
    limits = {}
    for provider in PROVIDERS:
        default_limit = PROVIDER_DEFAULT_MAX_CONCURRENT_DOWNLOADS.get(provider, 1)
        env_name = "GG_INPUT_MAX_CONCURRENT_DOWNLOADS_{}".format(provider.upper())
        env_value = os.environ.get(env_name, "").strip()
        limit = parse_positive_int(env_value, default_limit)
        limits[provider] = max(1, min(int(global_jobs), int(limit)))
    return limits


def resolve_ncbi_api_key():
    return os.environ.get("GG_NCBI_API_KEY", "").strip()


def resolve_ncbi_eutils_max_rps():
    fallback = DEFAULT_NCBI_EUTILS_RPS_NO_API_KEY
    if resolve_ncbi_api_key() != "":
        fallback = DEFAULT_NCBI_EUTILS_RPS_WITH_API_KEY
    env_value = os.environ.get("GG_NCBI_EUTILS_MAX_RPS", "").strip()
    return parse_positive_int(env_value, fallback)


def throttle_ncbi_eutils_request():
    max_rps = resolve_ncbi_eutils_max_rps()
    if max_rps < 1:
        return
    window_seconds = 1.0
    while True:
        with _ncbi_eutils_rate_lock:
            now = time.time()
            while len(_ncbi_eutils_request_times) > 0 and (now - _ncbi_eutils_request_times[0]) >= window_seconds:
                _ncbi_eutils_request_times.popleft()
            if len(_ncbi_eutils_request_times) < max_rps:
                _ncbi_eutils_request_times.append(now)
                return
            oldest = _ncbi_eutils_request_times[0]
            sleep_seconds = window_seconds - (now - oldest)
        if sleep_seconds < 0.001:
            sleep_seconds = 0.001
        time.sleep(sleep_seconds)


def is_transient_network_error(exc):
    pending = [exc]
    seen = set()
    while len(pending) > 0:
        current = pending.pop()
        if current is None:
            continue
        current_id = id(current)
        if current_id in seen:
            continue
        seen.add(current_id)

        if isinstance(current, HTTPError):
            if int(getattr(current, "code", 0)) in TRANSIENT_HTTP_STATUS_CODES:
                return True
            continue

        if isinstance(
            current,
            (
                RemoteDisconnected,
                IncompleteRead,
                TimeoutError,
                socket.timeout,
                ConnectionResetError,
                ConnectionAbortedError,
                BrokenPipeError,
                EOFError,
            ),
        ):
            return True

        if isinstance(current, OSError):
            if getattr(current, "errno", None) in (104, 110, 111, 113):
                return True

        if isinstance(current, URLError):
            reason = getattr(current, "reason", None)
            if isinstance(reason, str):
                lowered = reason.lower()
                if (
                    "timed out" in lowered
                    or "temporarily unavailable" in lowered
                    or "connection reset" in lowered
                    or "remote end closed connection" in lowered
                ):
                    return True
            elif isinstance(reason, BaseException):
                pending.append(reason)

        cause = getattr(current, "__cause__", None)
        context = getattr(current, "__context__", None)
        if isinstance(cause, BaseException):
            pending.append(cause)
        if isinstance(context, BaseException):
            pending.append(context)
    return False


def fetch_json(url, timeout):
    request = Request(url, headers={"User-Agent": "genegalleon-input-generation"})
    last_error = None
    for attempt in range(1, 4):
        try:
            with urlopen(request, timeout=timeout) as response:
                payload = response.read()
            return json.loads(payload.decode("utf-8"))
        except Exception as exc:
            last_error = exc
            if attempt >= 3 or not is_transient_network_error(exc):
                raise
            time.sleep(float(attempt))
    raise last_error


def resolve_ncbi_eutils_base_url():
    return os.environ.get("GG_NCBI_EUTILS_BASE_URL", DEFAULT_NCBI_EUTILS_BASE_URL).rstrip("/")


def resolve_ncbi_ftp_base_url():
    return os.environ.get("GG_NCBI_FTP_BASE_URL", DEFAULT_NCBI_FTP_BASE_URL).rstrip("/")


def resolve_ncbi_datasets_base_url():
    return os.environ.get("GG_NCBI_DATASETS_BASE_URL", DEFAULT_NCBI_DATASETS_BASE_URL).rstrip("/")


def normalize_ncbi_ftp_path(ftppath):
    if ftppath == "":
        return ""
    ftp_base = resolve_ncbi_ftp_base_url()
    if ftppath.startswith("ftp://ftp.ncbi.nlm.nih.gov"):
        return ftp_base + ftppath[len("ftp://ftp.ncbi.nlm.nih.gov"):]
    if ftppath.startswith("https://ftp.ncbi.nlm.nih.gov"):
        return ftp_base + ftppath[len("https://ftp.ncbi.nlm.nih.gov"):]
    if ftppath.startswith("http://ftp.ncbi.nlm.nih.gov"):
        return ftp_base + ftppath[len("http://ftp.ncbi.nlm.nih.gov"):]
    if ftppath.startswith("ftp://"):
        return "https://" + ftppath[len("ftp://"):]
    return ftppath


def infer_ncbi_species_key_from_doc(doc, fallback):
    candidates = []
    speciesname = str(doc.get("speciesname", "") or "").strip()
    organism = str(doc.get("organism", "") or "").strip()
    if speciesname != "":
        candidates.append(speciesname)
    if organism != "":
        candidates.append(organism)
    for raw in candidates:
        text = re.sub(r"\s*\(.*?\)\s*", " ", raw).strip()
        tokens = re.findall(r"[A-Za-z0-9]+", text)
        if len(tokens) >= 2:
            return "{}_{}".format(tokens[0], tokens[1])
        if len(tokens) == 1:
            return tokens[0]
    return fallback


def resolve_ncbi_download_urls_from_id(source_id, timeout, ncbi_source="auto"):
    accession = extract_ncbi_accession_from_source_id(source_id)
    if accession == "":
        raise ValueError(
            "ncbi id must include an assembly accession like GCF_000001405.40 or GCA_000001405.29: {}".format(
                source_id
            )
        )

    eutils_base = resolve_ncbi_eutils_base_url()
    term = quote("{}[Assembly Accession]".format(accession), safe="")
    api_key = resolve_ncbi_api_key()
    api_key_query = ""
    if api_key != "":
        api_key_query = "&api_key={}".format(quote(api_key, safe=""))
    esearch_url = "{}/esearch.fcgi?db=assembly&term={}&retmode=json{}".format(eutils_base, term, api_key_query)
    throttle_ncbi_eutils_request()
    esearch_data = fetch_json(esearch_url, timeout=timeout)
    uid_list = esearch_data.get("esearchresult", {}).get("idlist", [])
    if len(uid_list) == 0:
        raise ValueError("NCBI assembly was not found for id: {}".format(source_id))
    uid = uid_list[0]

    esummary_url = "{}/esummary.fcgi?db=assembly&id={}&retmode=json{}".format(
        eutils_base,
        quote(uid, safe=""),
        api_key_query,
    )
    throttle_ncbi_eutils_request()
    esummary_data = fetch_json(esummary_url, timeout=timeout)
    doc = esummary_data.get("result", {}).get(uid, {})

    ftppath_refseq = str(doc.get("ftppath_refseq", "") or "").strip()
    ftppath_genbank = str(doc.get("ftppath_genbank", "") or "").strip()
    selected_source = "auto"
    if ncbi_source == "refseq":
        selected_source = "refseq"
        ftp_dir = ftppath_refseq if ftppath_refseq != "" else ftppath_genbank
        if ftppath_refseq == "" and ftppath_genbank != "":
            selected_source = "genbank"
    elif ncbi_source == "genbank":
        selected_source = "genbank"
        ftp_dir = ftppath_genbank if ftppath_genbank != "" else ftppath_refseq
        if ftppath_genbank == "" and ftppath_refseq != "":
            selected_source = "refseq"
    else:
        ftp_dir = ftppath_refseq if ftppath_refseq != "" else ftppath_genbank
        if ftppath_refseq != "":
            selected_source = "refseq"
        elif ftppath_genbank != "":
            selected_source = "genbank"
    if ftp_dir == "":
        raise ValueError("NCBI FTP path was not found in assembly summary for id: {}".format(source_id))

    normalized_ftp_dir = normalize_ncbi_ftp_path(ftp_dir).rstrip("/")
    assembly_dir_name = Path(urlparse(ftp_dir).path.rstrip("/")).name
    if assembly_dir_name == "":
        raise ValueError("Could not parse assembly directory name from: {}".format(ftp_dir))

    cds_filename = "{}_cds_from_genomic.fna.gz".format(assembly_dir_name)
    gff_filename = "{}_genomic.gff.gz".format(assembly_dir_name)
    gbff_filename = "{}_genomic.gbff.gz".format(assembly_dir_name)
    genome_filename = "{}_genomic.fna.gz".format(assembly_dir_name)
    species_key = infer_ncbi_species_key_from_doc(doc, accession)
    return {
        "species_key": species_key,
        "cds_url": "{}/{}".format(normalized_ftp_dir, cds_filename),
        "gff_url": "{}/{}".format(normalized_ftp_dir, gff_filename),
        "gbff_url": "{}/{}".format(normalized_ftp_dir, gbff_filename),
        "genome_url": "{}/{}".format(normalized_ftp_dir, genome_filename),
        "cds_filename": cds_filename,
        "gff_filename": gff_filename,
        "gbff_filename": gbff_filename,
        "genome_filename": genome_filename,
        "ncbi_source_db": selected_source,
    }


def resolve_download_urls_from_templates(provider, source_id, species_key, row):
    cds_template = str(row.get("cds_url_template", "") or "").strip()
    gff_template = str(row.get("gff_url_template", "") or "").strip()
    genome_template = str(row.get("genome_url_template", "") or "").strip()
    if cds_template == "" and gff_template == "" and genome_template == "":
        return None
    mapping = {"id": source_id, "species_key": species_key, "provider": provider}
    resolved = {}
    try:
        if cds_template != "":
            resolved["cds_url"] = cds_template.format(**mapping)
        if gff_template != "":
            resolved["gff_url"] = gff_template.format(**mapping)
        if genome_template != "":
            resolved["genome_url"] = genome_template.format(**mapping)
    except KeyError as exc:
        raise ValueError("template placeholder is not supported: {}".format(exc))
    return resolved


def pick_ncbi_datasets_member_name(member_names, label):
    if label == "CDS":
        suffixes = ("/cds_from_genomic.fna", "_cds_from_genomic.fna")
    elif label == "GFF":
        suffixes = ("/genomic.gff", "_genomic.gff", "/genomic.gff3", "_genomic.gff3")
    elif label == "GBFF":
        suffixes = ("/genomic.gbff", "_genomic.gbff", "/genomic.gbff3", "_genomic.gbff3")
    elif label == "GENOME":
        suffixes = ("/genomic.fna", "_genomic.fna", "/genomic.fa", "_genomic.fa")
    else:
        raise ValueError("Unsupported NCBI datasets label: {}".format(label))

    matches = [name for name in member_names if any(name.endswith(suffix) for suffix in suffixes)]
    if len(matches) == 0:
        return ""
    return sorted(matches)[0]


def write_download_payload(destination, payload_bytes):
    tmp = Path(str(destination) + ".tmp.{}".format(os.getpid()))
    try:
        if destination.name.lower().endswith(".gz"):
            with gzip.open(tmp, "wb") as out:
                out.write(payload_bytes)
        else:
            with open(tmp, "wb") as out:
                out.write(payload_bytes)
        tmp.replace(destination)
    except Exception:
        try:
            tmp.unlink()
        except FileNotFoundError:
            pass
        except OSError:
            pass
        raise


def download_ncbi_datasets_file_from_id(
    source_id,
    label,
    destination,
    headers,
    timeout,
    dry_run,
    overwrite,
    lock_stale_seconds,
    warnings,
    lock_context,
):
    include_annotation_type = NCBI_DATASETS_INCLUDE_BY_LABEL.get(label, "")
    if include_annotation_type == "":
        raise ValueError("No datasets include_annotation_type is defined for label: {}".format(label))
    if dry_run:
        return False

    lock_path = Path(str(destination) + ".lock")
    heartbeat_state = acquire_download_lock(lock_path, lock_stale_seconds, warnings, lock_context)
    tmp_zip = Path(str(destination) + ".datasets.tmp.{}".format(os.getpid()))
    try:
        if destination.exists() and destination.stat().st_size > 0 and not overwrite:
            return False

        base = resolve_ncbi_datasets_base_url()
        datasets_url = "{}/genome/accession/{}/download?include_annotation_type={}".format(
            base,
            quote(source_id, safe=""),
            quote(include_annotation_type, safe=""),
        )

        req_headers = dict(headers)
        if "User-Agent" not in req_headers:
            req_headers["User-Agent"] = "genegalleon-input-generation"
        if "Accept" not in req_headers:
            req_headers["Accept"] = "application/zip"
        if "Accept-Encoding" not in req_headers:
            req_headers["Accept-Encoding"] = "identity"

        last_error = None
        for attempt in range(1, 4):
            try:
                request = Request(datasets_url, headers=req_headers)
                with urlopen(request, timeout=timeout) as response, open(tmp_zip, "wb") as out:
                    while True:
                        chunk = response.read(1024 * 1024)
                        if not chunk:
                            break
                        out.write(chunk)

                with zipfile.ZipFile(tmp_zip) as archive:
                    member_name = pick_ncbi_datasets_member_name(archive.namelist(), label)
                    if member_name == "":
                        raise ValueError(
                            "datasets archive did not contain expected {} member for id {}".format(label, source_id)
                        )
                    payload = archive.read(member_name)
                write_download_payload(destination, payload)
                last_error = None
                break
            except Exception as exc:
                last_error = exc
                try:
                    tmp_zip.unlink()
                except FileNotFoundError:
                    pass
                except OSError:
                    pass
                if attempt < 3:
                    time.sleep(float(attempt))
                    continue
        if last_error is not None:
            raise last_error
    except Exception:
        try:
            tmp_zip.unlink()
        except FileNotFoundError:
            pass
        except OSError:
            pass
        raise
    finally:
        try:
            tmp_zip.unlink()
        except FileNotFoundError:
            pass
        except OSError:
            pass
        release_download_lock(lock_path, heartbeat_state)
    return True


def resolve_download_lock_stale_seconds():
    raw = os.environ.get("GG_DOWNLOAD_LOCK_STALE_SECONDS", "").strip()
    if raw == "":
        return DEFAULT_DOWNLOAD_LOCK_STALE_SECONDS
    try:
        value = int(raw)
    except ValueError:
        return DEFAULT_DOWNLOAD_LOCK_STALE_SECONDS
    if value < 1:
        return 1
    return value


def resolve_download_lock_heartbeat_seconds():
    raw = os.environ.get("GG_DOWNLOAD_LOCK_HEARTBEAT_SECONDS", "").strip()
    if raw == "":
        return DEFAULT_DOWNLOAD_LOCK_HEARTBEAT_SECONDS
    try:
        value = int(raw)
    except ValueError:
        return DEFAULT_DOWNLOAD_LOCK_HEARTBEAT_SECONDS
    if value < 1:
        return 1
    return value


def resolve_download_lock_acquire_timeout_seconds():
    raw = os.environ.get("GG_DOWNLOAD_LOCK_ACQUIRE_TIMEOUT_SECONDS", "").strip()
    if raw == "":
        return DEFAULT_DOWNLOAD_LOCK_ACQUIRE_TIMEOUT_SECONDS
    try:
        value = int(raw)
    except ValueError:
        return DEFAULT_DOWNLOAD_LOCK_ACQUIRE_TIMEOUT_SECONDS
    if value < 1:
        return 1
    return value


def resolve_download_lock_poll_seconds():
    raw = os.environ.get("GG_DOWNLOAD_LOCK_POLL_SECONDS", "").strip()
    if raw == "":
        return DEFAULT_DOWNLOAD_LOCK_POLL_SECONDS
    try:
        value = float(raw)
    except ValueError:
        return DEFAULT_DOWNLOAD_LOCK_POLL_SECONDS
    if value <= 0:
        return 0.1
    return value


def lock_pid_is_alive(pid):
    try:
        os.kill(pid, 0)
        return True
    except OSError:
        return False


def lock_hostname():
    return socket.gethostname()


def lock_boot_id():
    boot_id_path = Path("/proc/sys/kernel/random/boot_id")
    try:
        if boot_id_path.is_file():
            return boot_id_path.read_text(encoding="utf-8").strip()
    except OSError:
        pass
    try:
        completed = subprocess.run(
            ["sysctl", "-n", "kern.bootsessionuuid"],
            capture_output=True,
            text=True,
            check=False,
        )
    except OSError:
        return ""
    if completed.returncode != 0:
        return ""
    return completed.stdout.strip()


def read_lock_metadata(lock_path):
    metadata = {
        "format": "",
        "pid": None,
        "hostname": "",
        "boot_id": "",
        "created_at": "",
    }
    try:
        raw = lock_path.read_text(encoding="utf-8").strip()
    except OSError:
        return metadata
    if raw == "":
        return metadata
    try:
        payload = json.loads(raw)
    except json.JSONDecodeError:
        return metadata
    if not isinstance(payload, dict):
        return metadata
    metadata["format"] = str(payload.get("format", "") or "")
    pid = payload.get("pid")
    if isinstance(pid, int):
        metadata["pid"] = pid
    else:
        try:
            metadata["pid"] = int(str(pid).strip())
        except (TypeError, ValueError):
            metadata["pid"] = None
    metadata["hostname"] = str(payload.get("hostname", "") or "")
    metadata["boot_id"] = str(payload.get("boot_id", "") or "")
    metadata["created_at"] = str(payload.get("created_at", "") or "")
    return metadata


def read_lock_stat(lock_path):
    try:
        return lock_path.stat()
    except OSError:
        return None


def format_lock_owner_summary(lock_path, metadata=None, stat_result=None):
    if metadata is None:
        metadata = read_lock_metadata(lock_path)
    if stat_result is None:
        stat_result = read_lock_stat(lock_path)
    if stat_result is None:
        age_text = "unknown"
    else:
        age = int(time.time() - stat_result.st_mtime)
        if age < 0:
            age = 0
        age_text = "{}s".format(age)
    return "host={}, pid={}, created_at={}, boot_id={}, heartbeat_age={}, lock={}".format(
        metadata.get("hostname", "") or "unknown",
        metadata.get("pid", None) if metadata.get("pid", None) is not None else "unknown",
        metadata.get("created_at", "") or "unknown",
        metadata.get("boot_id", "") or "unknown",
        age_text,
        lock_path,
    )


def stale_lock_reason(lock_path, stale_seconds, metadata=None, stat_result=None):
    if not lock_path.exists():
        return ""
    if metadata is None:
        metadata = read_lock_metadata(lock_path)
    if stat_result is None:
        stat_result = read_lock_stat(lock_path)
    current_boot_id = lock_boot_id()
    is_same_host_boot = (
        metadata.get("hostname", "") != ""
        and metadata.get("boot_id", "") != ""
        and current_boot_id != ""
        and metadata.get("hostname", "") == lock_hostname()
        and metadata.get("boot_id", "") == current_boot_id
    )
    pid = metadata.get("pid", None)
    if is_same_host_boot and pid is not None and not lock_pid_is_alive(pid):
        return "same_host_same_boot_dead_pid"
    if stat_result is None:
        return ""
    age = int(time.time() - stat_result.st_mtime)
    if age < 0:
        age = 0
    if age < stale_seconds:
        return ""
    return "heartbeat_timeout"


def remove_lock_if_unchanged(lock_path, stat_result):
    current_stat = read_lock_stat(lock_path)
    if current_stat is None or stat_result is None:
        return False
    if current_stat.st_dev != stat_result.st_dev or current_stat.st_ino != stat_result.st_ino:
        return False
    try:
        lock_path.unlink()
    except FileNotFoundError:
        return False
    except OSError:
        return False
    return True


def start_download_lock_heartbeat(lock_path):
    stop_event = threading.Event()
    interval_seconds = resolve_download_lock_heartbeat_seconds()

    def _heartbeat():
        while not stop_event.wait(interval_seconds):
            try:
                os.utime(lock_path, None)
            except FileNotFoundError:
                return
            except OSError:
                continue

    thread = threading.Thread(
        target=_heartbeat,
        name="download-lock-heartbeat-{}".format(lock_path.name),
        daemon=True,
    )
    thread.start()
    return stop_event, thread


def acquire_download_lock(lock_path, stale_seconds, warnings, lock_context):
    timeout_seconds = resolve_download_lock_acquire_timeout_seconds()
    poll_seconds = resolve_download_lock_poll_seconds()
    wait_started = time.monotonic()
    wait_logged = False
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    while True:
        payload = {
            "format": SHARED_DOWNLOAD_LOCK_FORMAT,
            "pid": os.getpid(),
            "hostname": lock_hostname(),
            "boot_id": lock_boot_id(),
            "created_at": time.time(),
        }
        try:
            fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        except FileExistsError:
            metadata = read_lock_metadata(lock_path)
            stat_result = read_lock_stat(lock_path)
            reason = stale_lock_reason(lock_path, stale_seconds, metadata, stat_result)
            if reason != "":
                owner_summary = format_lock_owner_summary(lock_path, metadata, stat_result)
                if remove_lock_if_unchanged(lock_path, stat_result):
                    warnings.append(
                        "[download-lock] recovered stale lock for {}: {} ({}; {})".format(
                            lock_context, lock_path, reason, owner_summary
                        )
                    )
                    continue
            owner_summary = format_lock_owner_summary(lock_path, metadata, stat_result)
            if not wait_logged:
                warnings.append(
                    "[download-lock] waiting for shared lock for {} ({})".format(
                        lock_context, owner_summary
                    )
                )
                wait_logged = True
            if time.monotonic() - wait_started >= float(timeout_seconds):
                raise RuntimeError(
                    "[download-lock] timed out waiting for shared lock for {} ({})".format(
                        lock_context, owner_summary
                    )
                )
            time.sleep(float(poll_seconds))
            continue
        try:
            with os.fdopen(fd, "w", encoding="utf-8") as handle:
                json.dump(payload, handle, separators=(",", ":"))
                handle.write("\n")
        except Exception:
            try:
                lock_path.unlink()
            except OSError:
                pass
            raise
        finally:
            try:
                os.close(fd)
            except OSError:
                pass
        return start_download_lock_heartbeat(lock_path)


def release_download_lock(lock_path, heartbeat_state=None):
    if heartbeat_state is not None:
        stop_event, thread = heartbeat_state
        stop_event.set()
        thread.join(timeout=1.0)
    try:
        lock_path.unlink()
    except FileNotFoundError:
        return
    except OSError:
        return


def download_url_to_file(
    url,
    destination,
    headers,
    timeout,
    dry_run,
    overwrite,
    lock_stale_seconds,
    warnings,
    lock_context,
    archive_member="",
):
    if dry_run:
        return False
    archive_member_text = str(archive_member or "").strip()
    archive_cache_path = None
    archive_tmp = None
    if archive_member_text == "":
        lock_path = Path(str(destination) + ".lock")
    else:
        archive_name = Path(unquote(urlparse(url).path)).name
        if archive_name == "":
            archive_name = "archive.bin"
        archive_hash = hashlib.sha256(str(url).encode("utf-8")).hexdigest()[:16]
        archive_cache_dir = destination.parent / ".archive_cache"
        archive_cache_dir.mkdir(parents=True, exist_ok=True)
        archive_cache_path = archive_cache_dir / "{}__{}".format(archive_hash, archive_name)
        lock_path = Path(str(archive_cache_path) + ".lock")
    heartbeat_state = acquire_download_lock(lock_path, lock_stale_seconds, warnings, lock_context)
    tmp = Path(str(destination) + ".tmp.{}".format(os.getpid()))
    try:
        if destination.exists() and destination.stat().st_size > 0 and not overwrite:
            return False
        if archive_member_text == "":
            request = Request(url, headers=headers)
            with urlopen(request, timeout=timeout) as response, open(tmp, "wb") as out:
                while True:
                    chunk = response.read(1024 * 1024)
                    if not chunk:
                        break
                    out.write(chunk)
            tmp.replace(destination)
        else:
            archive_tmp = Path(str(archive_cache_path) + ".tmp.{}".format(os.getpid()))
            if overwrite or not archive_cache_path.exists() or archive_cache_path.stat().st_size == 0:
                request = Request(url, headers=headers)
                with urlopen(request, timeout=timeout) as response, open(archive_tmp, "wb") as out:
                    while True:
                        chunk = response.read(1024 * 1024)
                        if not chunk:
                            break
                        out.write(chunk)
                archive_tmp.replace(archive_cache_path)
            if zipfile.is_zipfile(archive_cache_path):
                with zipfile.ZipFile(archive_cache_path) as archive:
                    payload = archive.read(archive_member_text)
            else:
                with tarfile.open(archive_cache_path, "r:*") as archive:
                    extracted = archive.extractfile(archive_member_text)
                    if extracted is None:
                        raise KeyError(archive_member_text)
                    payload = extracted.read()
            with open(tmp, "wb") as out:
                out.write(payload)
            tmp.replace(destination)
    except Exception:
        try:
            tmp.unlink()
        except FileNotFoundError:
            pass
        except OSError:
            pass
        if archive_tmp is not None:
            try:
                archive_tmp.unlink()
            except FileNotFoundError:
                pass
            except OSError:
                pass
        raise
    finally:
        release_download_lock(lock_path, heartbeat_state)
    return True


def read_download_manifest(path):
    if path.suffix.lower() == ".xlsx":
        return read_download_manifest_xlsx(path)
    delimiter = detect_manifest_delimiter(path)
    with open(path, "rt", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        rows = [row for row in reader]
    return rows


def resolve_local_reference_path(reference, manifest_parent_dir):
    text = str(reference or "").strip()
    if text == "":
        return None
    if is_url_like(text):
        parsed = urlparse(text)
        if parsed.scheme.lower() != "file":
            raise ValueError("expected local file path or file:// URL, got '{}'".format(reference))
        path = Path(unquote(parsed.path)).expanduser()
        return path.resolve()
    path = Path(text).expanduser()
    if not path.is_absolute():
        path = (manifest_parent_dir / path).resolve()
    else:
        path = path.resolve()
    return path


def resolve_local_source_id_with_label_fallback(source_id, manifest_parent_dir):
    text = str(source_id or "").strip()
    if text == "":
        return text
    try:
        direct_path = resolve_local_reference_path(text, manifest_parent_dir)
    except Exception:
        direct_path = None
    if direct_path is not None and direct_path.exists():
        return text

    # Accept spreadsheet label format like:
    #   /path/to/species_dir (Species name)
    # only when the stripped path actually exists.
    match = re.match(r"^(.*\S)\s+\([^()]*\)$", text)
    if match is None:
        return text
    stripped = match.group(1).strip()
    if stripped == "":
        return text
    try:
        stripped_path = resolve_local_reference_path(stripped, manifest_parent_dir)
    except Exception:
        return text
    if stripped_path is not None and stripped_path.exists():
        return stripped
    return text


def local_reference_to_file_url(reference, manifest_parent_dir):
    path = resolve_local_reference_path(reference, manifest_parent_dir)
    if path is None:
        return ""
    return path.as_uri()


def resolve_local_manifest_row(provider, source_id, species_key, row, manifest_parent_dir, warnings, line_number):
    cds_url = local_reference_to_file_url(row.get("cds_url", ""), manifest_parent_dir)
    gff_url = local_reference_to_file_url(row.get("gff_url", ""), manifest_parent_dir)
    gbff_url = local_reference_to_file_url(row.get("gbff_url", ""), manifest_parent_dir)
    genome_url = local_reference_to_file_url(row.get("genome_url", ""), manifest_parent_dir)
    cds_filename = str(row.get("cds_filename") or "").strip()
    gff_filename = str(row.get("gff_filename") or "").strip()
    gbff_filename = str(row.get("gbff_filename") or "").strip()
    genome_filename = str(row.get("genome_filename") or "").strip()

    local_cds_path = str(row.get("local_cds_path") or "").strip()
    local_gff_path = str(row.get("local_gff_path") or "").strip()
    local_gbff_path = str(row.get("local_gbff_path") or "").strip()
    local_genome_path = str(row.get("local_genome_path") or "").strip()
    if cds_url == "" and local_cds_path != "":
        cds_url = local_reference_to_file_url(local_cds_path, manifest_parent_dir)
    if gff_url == "" and local_gff_path != "":
        gff_url = local_reference_to_file_url(local_gff_path, manifest_parent_dir)
    if gbff_url == "" and local_gbff_path != "":
        gbff_url = local_reference_to_file_url(local_gbff_path, manifest_parent_dir)
    if genome_url == "" and local_genome_path != "":
        genome_url = local_reference_to_file_url(local_genome_path, manifest_parent_dir)

    source_id_text = str(source_id or "").strip()
    source_id_resolved = resolve_local_source_id_with_label_fallback(source_id_text, manifest_parent_dir)
    if source_id_resolved != source_id_text:
        warnings.append(
            "Manifest line {}: provider=local id looked like a labeled choice. "
            "Using '{}'".format(line_number, source_id_resolved)
        )
    source_dir = resolve_local_reference_path(source_id_resolved, manifest_parent_dir)
    if species_key == "":
        if source_dir is not None and source_dir.exists() and source_dir.is_dir():
            species_key = source_dir.name
        else:
            species_key = source_id_resolved

    if (not manifest_has_usable_source_bundle(cds_url, gff_url, gbff_url, genome_url)) and source_dir is not None and source_dir.exists() and source_dir.is_dir():
        files = [path for path in source_dir.iterdir() if path.is_file()]
        cds_matches = [
            path
            for path in files
            if is_fasta_filename(path.name)
            and any(marker in path.name.lower() for marker in ("cds", "transcript", "mrna", "cdna"))
        ]
        if len(cds_matches) == 0:
            cds_matches = [
                path
                for path in files
                if is_fasta_filename(path.name) and not is_probable_genome_filename(provider, path.name)
            ]
        gff_matches = [path for path in files if is_gff_filename(path.name)]
        gbff_matches = [path for path in files if is_gbff_filename(path.name)]
        genome_matches = [path for path in files if is_probable_genome_filename(provider, path.name)]

        cds_path = pick_single_file(cds_matches, provider, species_key, "CDS", warnings)
        gff_path = pick_single_file(gff_matches, provider, species_key, "GFF", warnings)
        gbff_path = pick_single_file(gbff_matches, provider, species_key, "GBFF", warnings)
        genome_path = pick_single_file(genome_matches, provider, species_key, "genome", warnings)

        if cds_url == "" and cds_path is not None:
            cds_url = cds_path.resolve().as_uri()
            if cds_filename == "":
                cds_filename = cds_path.name
        if gff_url == "" and gff_path is not None:
            gff_url = gff_path.resolve().as_uri()
            if gff_filename == "":
                gff_filename = gff_path.name
        if gbff_url == "" and gbff_path is not None:
            gbff_url = gbff_path.resolve().as_uri()
            if gbff_filename == "":
                gbff_filename = gbff_path.name
        if genome_url == "" and genome_path is not None:
            genome_url = genome_path.resolve().as_uri()
            if genome_filename == "":
                genome_filename = genome_path.name

    if not manifest_has_usable_source_bundle(cds_url, gff_url, gbff_url, genome_url):
        raise ValueError(
            "provider=local requires CDS, GBFF, or GFF plus genome paths/URLs, or a species directory id. "
            "line {} id='{}'".format(line_number, source_id_resolved)
        )

    return {
        "species_key": species_key,
        "cds_url": cds_url,
        "gff_url": gff_url,
        "gbff_url": gbff_url,
        "genome_url": genome_url,
        "cds_filename": cds_filename,
        "gff_filename": gff_filename,
        "gbff_filename": gbff_filename,
        "genome_filename": genome_filename,
    }


def execute_download_target_job(
    job,
    headers,
    timeout,
    overwrite,
    lock_stale_seconds,
    provider_semaphores,
):
    provider = job["provider"]
    source_id = job["source_id"]
    species_key = job["species_key"]
    label = job["label"]
    url = job["url"]
    target = job["target"]
    archive_member = str(job.get("archive_member") or "").strip()
    local_warnings = []
    local_errors = []
    downloaded = 0
    failed = []

    sem = provider_semaphores.get(provider)
    if sem is None:
        sem = threading.Semaphore(1)

    with sem:
        if target.exists() and target.stat().st_size > 0 and not overwrite:
            local_warnings.append(
                "[download:{}] {} {} already exists. Skipping: {}".format(provider, species_key, label, target)
            )
            return {"warnings": local_warnings, "errors": local_errors, "downloaded": downloaded, "failed": failed}

        try:
            did_download = download_url_to_file(
                url,
                target,
                headers=headers,
                timeout=timeout,
                dry_run=False,
                overwrite=overwrite,
                lock_stale_seconds=lock_stale_seconds,
                warnings=local_warnings,
                lock_context="[download:{}] {} {}".format(provider, species_key, label),
                archive_member=archive_member,
            )
            if did_download:
                downloaded += 1
            elif target.exists() and target.stat().st_size > 0 and not overwrite:
                local_warnings.append(
                    "[download:{}] {} {} already exists after lock. Skipping: {}".format(
                        provider, species_key, label, target
                    )
                )
        except Exception as exc:
            fallback_exc = None
            if provider in ("ncbi", "refseq", "genbank") and source_id != "":
                try:
                    did_download = download_ncbi_datasets_file_from_id(
                        source_id=source_id,
                        label=label,
                        destination=target,
                        headers=headers,
                        timeout=timeout,
                        dry_run=False,
                        overwrite=overwrite,
                        lock_stale_seconds=lock_stale_seconds,
                        warnings=local_warnings,
                        lock_context="[download:{}] {} {} datasets".format(provider, species_key, label),
                    )
                    if did_download:
                        downloaded += 1
                        local_warnings.append(
                            "[download:{}] {} {} fallback via NCBI Datasets API for id '{}'".format(
                                provider, species_key, label, source_id
                            )
                        )
                        return {"warnings": local_warnings, "errors": local_errors, "downloaded": downloaded, "failed": failed}
                except Exception as fallback_error:
                    fallback_exc = fallback_error
            if fallback_exc is None:
                failed.append(
                    {
                        "row_id": job.get("row_id"),
                        "message": "[download:{}] failed {} {} from {} -> {} ({})".format(
                            provider, species_key, label, url, target, exc
                        ),
                    }
                )
            else:
                failed.append(
                    {
                        "row_id": job.get("row_id"),
                        "message": "[download:{}] failed {} {} from {} -> {} ({}) ; fallback datasets failed ({})".format(
                            provider, species_key, label, url, target, exc, fallback_exc
                        ),
                    }
                )
    return {"warnings": local_warnings, "errors": local_errors, "downloaded": downloaded, "failed": failed}


def download_from_manifest(
    manifest_path,
    download_root,
    provider_filter,
    overwrite,
    headers,
    timeout,
    dry_run,
    jobs,
    resolved_manifest_output_path=None,
):
    rows = read_download_manifest(manifest_path)
    warnings = []
    errors = []
    processed = 0
    downloaded = 0
    planned = 0
    resolved_rows = []
    resolved_fieldnames = resolved_manifest_fieldnames(rows)
    lock_stale_seconds = resolve_download_lock_stale_seconds()
    download_jobs = []
    failed_downloads = []
    row_target_paths = {}

    if len(rows) == 0:
        errors.append("Download manifest is empty: {}".format(manifest_path))
        return {
            "warnings": warnings,
            "errors": errors,
            "processed": processed,
            "downloaded": downloaded,
            "planned": planned,
        }
    manifest_parent_dir = manifest_path.parent
    header_cols = set(rows[0].keys())
    if "provider" not in header_cols or "id" not in header_cols:
        errors.append("Download manifest must contain required columns provider,id: {}".format(manifest_path))
        return {
            "warnings": warnings,
            "errors": errors,
            "processed": processed,
            "downloaded": downloaded,
            "planned": planned,
        }

    for i, row in enumerate(rows, start=2):
        provider = (row.get("provider") or "").strip().lower()
        source_id_raw = (row.get("id") or "").strip()
        source_id = normalize_manifest_source_id(provider, source_id_raw)
        species_key = (row.get("species_key") or "").strip()
        cds_url = (row.get("cds_url") or "").strip()
        gff_url = (row.get("gff_url") or "").strip()
        gbff_url = (row.get("gbff_url") or "").strip()
        genome_url = (row.get("genome_url") or "").strip()
        cds_archive_member = (row.get("cds_archive_member") or "").strip()
        gff_archive_member = (row.get("gff_archive_member") or "").strip()
        gbff_archive_member = (row.get("gbff_archive_member") or "").strip()
        genome_archive_member = (row.get("genome_archive_member") or "").strip()
        cds_filename = (row.get("cds_filename") or "").strip()
        gff_filename = (row.get("gff_filename") or "").strip()
        gbff_filename = (row.get("gbff_filename") or "").strip()
        genome_filename = (row.get("genome_filename") or "").strip()
        resolved_ncbi = None

        if provider_filter != "all" and provider != provider_filter:
            continue
        processed += 1

        if provider == "":
            errors.append("Manifest line {}: provider is required".format(i))
            continue
        if source_id == "":
            errors.append("Manifest line {}: id is required".format(i))
            continue
        if provider not in PROVIDERS:
            errors.append("Manifest line {}: unsupported provider '{}'".format(i, provider))
            continue
        if provider == "coge":
            coge_gid = extract_coge_gid_candidate(source_id)
            if coge_gid == "":
                errors.append(
                    "Manifest line {}: provider=coge requires numeric genome_id (gid) in id column".format(i)
                )
                continue
            source_id = coge_gid
        if provider not in DOWNLOAD_MANIFEST_SUPPORTED_PROVIDERS:
            errors.append(
                "Manifest line {}: provider '{}' is not supported for --download-manifest "
                "(use --input-dir for local formatting)".format(i, provider)
            )
            continue

        if provider == "local":
            try:
                resolved_local = resolve_local_manifest_row(
                    provider=provider,
                    source_id=source_id,
                    species_key=species_key,
                    row=row,
                    manifest_parent_dir=manifest_parent_dir,
                    warnings=warnings,
                    line_number=i,
                )
                species_key = str(resolved_local.get("species_key") or "").strip()
                cds_url = str(resolved_local.get("cds_url") or "").strip()
                gff_url = str(resolved_local.get("gff_url") or "").strip()
                gbff_url = str(resolved_local.get("gbff_url") or "").strip()
                genome_url = str(resolved_local.get("genome_url") or "").strip()
                if cds_filename == "":
                    cds_filename = str(resolved_local.get("cds_filename") or "").strip()
                if gff_filename == "":
                    gff_filename = str(resolved_local.get("gff_filename") or "").strip()
                if gbff_filename == "":
                    gbff_filename = str(resolved_local.get("gbff_filename") or "").strip()
                if genome_filename == "":
                    genome_filename = str(resolved_local.get("genome_filename") or "").strip()
            except Exception as exc:
                errors.append(
                    "Manifest line {}: failed to resolve local paths from id '{}' (provider={}): {}".format(
                        i, source_id, provider, exc
                    )
                )
                continue
        elif provider in ("ncbi", "refseq", "genbank"):
            accession_candidate = extract_ncbi_accession_from_source_id(source_id)
            requires_ncbi_resolution = not manifest_has_usable_source_bundle(cds_url, gff_url, gbff_url, genome_url)
            should_enrich_from_ncbi = accession_candidate != "" and (species_key == "" or genome_url == "")
            if requires_ncbi_resolution or should_enrich_from_ncbi:
                try:
                    preferred_ncbi_source = "auto"
                    if provider in ("refseq", "genbank"):
                        preferred_ncbi_source = provider
                    resolved_ncbi = resolve_ncbi_download_urls_from_id(
                        source_id,
                        timeout=float(timeout),
                        ncbi_source=preferred_ncbi_source,
                    )
                except Exception as exc:
                    if requires_ncbi_resolution:
                        errors.append(
                            "Manifest line {}: failed to resolve id '{}' (provider={}): {}".format(
                                i, source_id, provider, exc
                            )
                        )
                        continue
                    warnings.append(
                        "Manifest line {}: skipped optional NCBI enrichment from id '{}' "
                        "(provider={}): {}".format(i, source_id, provider, exc)
                    )
                    resolved_ncbi = None

        if species_key == "" and resolved_ncbi is not None:
            species_key = (resolved_ncbi.get("species_key") or "").strip()

        if provider != "local" and (
            not manifest_has_usable_source_bundle(cds_url, gff_url, gbff_url, genome_url)
        ):
            try:
                resolved_urls = None
                if provider in ("ncbi", "refseq", "genbank"):
                    resolved_urls = resolved_ncbi
                else:
                    resolved_urls = resolve_download_urls_from_templates(provider, source_id, species_key, row)
                    if resolved_urls is None and (
                        not manifest_has_usable_source_bundle(cds_url, gff_url, gbff_url, genome_url)
                    ):
                        resolved_urls = resolve_non_ncbi_download_urls_from_id(
                            provider=provider,
                            source_id=source_id,
                            species_key=species_key,
                            timeout=float(timeout),
                            headers=headers,
                        )
                if resolved_urls is not None:
                    if cds_url == "":
                        cds_url = (resolved_urls.get("cds_url") or "").strip()
                    if gff_url == "":
                        gff_url = (resolved_urls.get("gff_url") or "").strip()
                    if gbff_url == "":
                        gbff_url = (resolved_urls.get("gbff_url") or "").strip()
                    if genome_url == "":
                        genome_url = (resolved_urls.get("genome_url") or "").strip()
                    if cds_filename == "":
                        cds_filename = (resolved_urls.get("cds_filename") or "").strip()
                    if gff_filename == "":
                        gff_filename = (resolved_urls.get("gff_filename") or "").strip()
                    if gbff_filename == "":
                        gbff_filename = (resolved_urls.get("gbff_filename") or "").strip()
                    if genome_filename == "":
                        genome_filename = (resolved_urls.get("genome_filename") or "").strip()
                    if species_key == "":
                        species_key = (resolved_urls.get("species_key") or "").strip()
            except Exception as exc:
                errors.append(
                    "Manifest line {}: failed to resolve urls from id '{}' (provider={}): {}".format(
                        i, source_id, provider, exc
                    )
                )
                continue

        if species_key == "":
            species_key = source_id
        if species_key == "":
            errors.append("Manifest line {}: species_key/id is empty".format(i))
            continue

        if not manifest_has_usable_source_bundle(cds_url, gff_url, gbff_url, genome_url):
            errors.append(
                "Manifest line {}: CDS, GBFF, or GFF plus genome are required for {} "
                "(set urls directly or provide resolvable id)".format(
                    i, species_key
                )
            )
            continue

        raw_dir = provider_raw_dir(provider, download_root, species_key)
        raw_dir.mkdir(parents=True, exist_ok=True)

        if cds_filename == "":
            cds_filename = default_download_filename(provider, species_key, "cds", cds_url, cds_archive_member)
        if gff_url != "" and gff_filename == "":
            gff_filename = default_download_filename(provider, species_key, "gff", gff_url, gff_archive_member)
        if gbff_url != "" and gbff_filename == "":
            gbff_filename = default_download_filename(provider, species_key, "gbff", gbff_url, gbff_archive_member)
        if genome_url != "" and genome_filename == "":
            genome_filename = default_download_filename(
                provider,
                species_key,
                "genome",
                genome_url,
                genome_archive_member,
            )

        resolved_rows.append(
            build_resolved_manifest_row(
                raw_row=row,
                fieldnames=resolved_fieldnames,
                provider=provider,
                source_id=source_id,
                species_key=species_key,
                cds_url=cds_url,
                gff_url=gff_url,
                gbff_url=gbff_url,
                genome_url=genome_url,
                cds_archive_member=cds_archive_member,
                gff_archive_member=gff_archive_member,
                gbff_archive_member=gbff_archive_member,
                genome_archive_member=genome_archive_member,
                cds_filename=cds_filename,
                gff_filename=gff_filename,
                gbff_filename=gbff_filename,
                genome_filename=genome_filename,
            )
        )

        targets = []
        if cds_url != "":
            targets.append(("CDS", cds_url, raw_dir / cds_filename, cds_archive_member))
        if gff_url != "":
            targets.append(("GFF", gff_url, raw_dir / gff_filename, gff_archive_member))
        if gbff_url != "":
            targets.append(("GBFF", gbff_url, raw_dir / gbff_filename, gbff_archive_member))
        if genome_url != "":
            targets.append(("GENOME", genome_url, raw_dir / genome_filename, genome_archive_member))
        row_target_paths[i] = {
            "provider": provider,
            "species_key": species_key,
            "paths": {label: target for label, _url, target, _archive_member in targets},
        }
        for label, url, target, archive_member in targets:
            if target.exists() and target.stat().st_size > 0 and not overwrite:
                warnings.append(
                    "[download:{}] {} {} already exists. Skipping: {}".format(
                        provider, species_key, label, target
                    )
                )
                continue
            if dry_run:
                planned += 1
                continue
            download_jobs.append(
                {
                    "row_id": i,
                    "provider": provider,
                    "source_id": source_id,
                    "species_key": species_key,
                    "label": label,
                    "url": url,
                    "target": target,
                    "archive_member": archive_member,
                }
            )

    if len(download_jobs) > 0:
        max_workers = max(1, int(jobs))
        provider_limits = resolve_provider_download_limits(max_workers)
        provider_semaphores = {
            provider_name: threading.Semaphore(provider_limits.get(provider_name, 1))
            for provider_name in PROVIDERS
        }

        with ThreadPoolExecutor(max_workers=max_workers) as pool:
            futures = [
                pool.submit(
                    execute_download_target_job,
                    job,
                    headers,
                    timeout,
                    overwrite,
                    lock_stale_seconds,
                    provider_semaphores,
                )
                for job in download_jobs
            ]
            for future in as_completed(futures):
                try:
                    result = future.result()
                except Exception as exc:
                    errors.append("Unhandled download worker error: {}".format(exc))
                    continue
                warnings.extend(result.get("warnings", []))
                errors.extend(result.get("errors", []))
                downloaded += int(result.get("downloaded", 0))
                failed_downloads.extend(result.get("failed", []))

    for failure in failed_downloads:
        row_info = row_target_paths.get(failure.get("row_id"), {})
        paths = row_info.get("paths", {})
        cds_ok = paths.get("CDS") is not None and paths["CDS"].exists() and paths["CDS"].stat().st_size > 0
        gff_ok = paths.get("GFF") is not None and paths["GFF"].exists() and paths["GFF"].stat().st_size > 0
        gbff_ok = paths.get("GBFF") is not None and paths["GBFF"].exists() and paths["GBFF"].stat().st_size > 0
        genome_ok = paths.get("GENOME") is not None and paths["GENOME"].exists() and paths["GENOME"].stat().st_size > 0
        if cds_ok or gbff_ok or (gff_ok and genome_ok):
            warnings.append(
                "{} ; continuing because a usable source bundle is available".format(
                    failure.get("message", "")
                )
            )
        else:
            errors.append(failure.get("message", "download failed"))

    if processed == 0:
        warnings.append("No manifest rows matched --provider {}.".format(provider_filter))

    if resolved_manifest_output_path is not None:
        try:
            write_resolved_manifest_tsv(resolved_manifest_output_path, resolved_fieldnames, resolved_rows)
        except Exception as exc:
            errors.append(
                "Failed to write resolved manifest TSV '{}': {}".format(
                    resolved_manifest_output_path, exc
                )
            )

    return {
        "warnings": warnings,
        "errors": errors,
        "processed": processed,
        "downloaded": downloaded,
        "planned": planned,
        "resolved_manifest_output": str(resolved_manifest_output_path) if resolved_manifest_output_path is not None else "",
    }


def is_fasta_filename(name):
    lower = name.lower()
    return any(lower.endswith(ext) for ext in FASTA_EXTENSIONS)


def is_gff_filename(name):
    lower = name.lower()
    return any(lower.endswith(ext) for ext in GFF_EXTENSIONS)


def is_gbff_filename(name):
    lower = name.lower()
    return any(lower.endswith(ext) for ext in GENBANK_EXTENSIONS)


def manifest_has_annotation_source(gff_url, gbff_url):
    return str(gff_url or "").strip() != "" or str(gbff_url or "").strip() != ""


def manifest_has_cds_source(cds_url, genome_url, gbff_url):
    return (
        str(cds_url or "").strip() != ""
        or str(genome_url or "").strip() != ""
        or str(gbff_url or "").strip() != ""
    )


def manifest_has_usable_source_bundle(cds_url, gff_url, gbff_url, genome_url):
    if str(cds_url or "").strip() != "":
        return True
    if str(gbff_url or "").strip() != "":
        return True
    return str(gff_url or "").strip() != "" and str(genome_url or "").strip() != ""


def task_has_usable_source_bundle(cds_path, gff_path, gbff_path, genome_path):
    return cds_path is not None or gbff_path is not None or (gff_path is not None and genome_path is not None)


def is_probable_genome_filename(provider, name):
    if not is_fasta_filename(name):
        return False
    lower = name.lower()
    non_genome_markers = (
        "cds",
        "mrna",
        "trna",
        "rrna",
        "ncrna",
        "rna_from_genomic",
        "transcript",
        "cdna",
        "pep",
        "protein",
    )
    if any(marker in lower for marker in non_genome_markers):
        return False
    if provider == "fernbase":
        # FernBase often exposes the assembly as a plain ".fa"/".fasta" filename.
        return True
    if provider in ("ncbi", "refseq", "genbank"):
        return "genomic" in lower
    if provider in ("ensembl", "ensemblplants"):
        return any(marker in lower for marker in ("dna", "genome", "toplevel", "primary_assembly", "chromosome"))
    return any(marker in lower for marker in ("genome", "assembly", "genomic", "dna", "chromosome", "scaffold"))


def first_token(text):
    parts = text.split()
    if len(parts) == 0:
        return ""
    return parts[0]


def extract_header_tag_value(header, tag):
    pattern = r"\[{}=([^\]]+)\]".format(re.escape(tag))
    match = re.search(pattern, header)
    if match is not None:
        return match.group(1).strip()
    plain_pattern = r"(?:^|\s){}=([^\s]+)".format(re.escape(tag))
    match = re.search(plain_pattern, header)
    if match is not None:
        return match.group(1).strip()
    return ""


def extract_ncbi_gene_id_from_header(header):
    db_xref = extract_header_tag_value(header, "db_xref")
    if db_xref == "":
        return ""
    for token in db_xref.split(","):
        text = token.strip()
        if text.startswith("GeneID:"):
            return text[len("GeneID:") :]
    return ""


def extract_ncbi_ensembl_gene_id_from_header(header):
    db_xref = extract_header_tag_value(header, "db_xref")
    if db_xref == "":
        return ""
    for token in db_xref.split(","):
        text = token.strip()
        if not text.startswith("Ensembl:"):
            continue
        candidate = text[len("Ensembl:") :].strip()
        if ENSEMBL_GENE_ID_PATTERN.match(candidate):
            return candidate
    return ""


def species_prefix_from_value(value):
    tokens = value.split("_")
    if len(tokens) < 2:
        return ""
    return "{}_{}".format(tokens[0], tokens[1])


def normalize_output_basename(source_name, species_prefix):
    if source_name.startswith(species_prefix + "_"):
        return source_name
    dotted_prefix = species_prefix + "."
    if source_name.startswith(dotted_prefix):
        return species_prefix + "_" + source_name[len(dotted_prefix):]
    return species_prefix + "_" + source_name


def strip_suffix_case_insensitive(text, suffixes):
    lower = text.lower()
    for suffix in sorted(suffixes, key=len, reverse=True):
        if lower.endswith(suffix):
            return text[: len(text) - len(suffix)]
    return text


def normalize_cds_output_basename(source_name, species_prefix):
    normalized = normalize_output_basename(source_name, species_prefix)
    stem = strip_suffix_case_insensitive(normalized, FASTA_EXTENSIONS)
    return stem + ".fa.gz"


def normalize_genome_output_basename(source_name, species_prefix):
    normalized = normalize_output_basename(source_name, species_prefix)
    stem = strip_suffix_case_insensitive(normalized, FASTA_EXTENSIONS)
    return stem + ".fa.gz"


def normalize_gff_output_basename(source_name, species_prefix):
    normalized = normalize_output_basename(source_name, species_prefix)
    stem = strip_suffix_case_insensitive(normalized, GFF_EXTENSIONS)
    return stem + ".gff.gz"


def apply_common_replacements(text):
    out = text
    for old, new in COMMON_REPLACEMENTS:
        out = out.replace(old, new)
    return out


def sanitize_identifier(identifier):
    out = identifier.replace("−", "-")
    out = apply_common_replacements(out)
    out = INVALID_ID_CHARS.sub("_", out)
    out = re.sub(r"_+", "_", out)
    out = out.strip("_")
    if out == "":
        out = "unnamed"
    return out


def pad_to_codon_length(seq):
    remainder = len(seq) % 3
    if remainder == 0:
        return seq
    return seq + ("N" * (3 - remainder))


def reverse_complement(seq):
    return seq.translate(str.maketrans("ACGTNacgtn", "TGCANtgcan"))[::-1]


def normalize_gff_attribute_value(raw_value):
    value = str(raw_value or "").strip()
    if len(value) >= 2 and value[0] == value[-1] and value[0] in ("'", '"'):
        value = value[1:-1]
    else:
        if value.startswith(("'", '"')):
            value = value[1:]
        if value.endswith(("'", '"')):
            value = value[:-1]
    return unquote(value.strip())


def parse_gff_attributes(attr_text):
    attrs = defaultdict(list)
    for raw_field in str(attr_text or "").split(";"):
        field = raw_field.strip()
        if field == "":
            continue
        equal_pos = field.find("=")
        space_pos = field.find(" ")
        if equal_pos != -1 and (space_pos == -1 or equal_pos < space_pos):
            key = field[:equal_pos].strip()
            raw_value = field[equal_pos + 1 :]
        elif space_pos > 0:
            key = field[:space_pos].strip()
            raw_value = field[space_pos + 1 :]
        else:
            continue
        if key == "":
            continue
        for raw_item in str(raw_value or "").split(","):
            value = normalize_gff_attribute_value(raw_item)
            if value != "":
                attrs[key].append(value)
    return {key: tuple(values) for key, values in attrs.items()}


def choose_first_gff_attribute(attrs, keys):
    for key in keys:
        values = attrs.get(key, ())
        if len(values) > 0 and str(values[0]).strip() != "":
            return str(values[0]).strip()
    return ""


def load_genome_sequences(path):
    sequences = {}
    for header, sequence in iter_fasta_records(path):
        full_name = apply_common_replacements(str(header or "").strip())
        token = first_token(full_name)
        seq = re.sub(r"\s+", "", str(sequence or "")).upper()
        if token != "" and token not in sequences:
            sequences[token] = seq
        if full_name != "" and full_name not in sequences:
            sequences[full_name] = seq
    return sequences


def iter_genome_records_from_gbff(path):
    for record in iter_genbank_records(path):
        record_id = extract_seqrecord_id(record)
        sequence = re.sub(r"\s+", "", str(record.seq or "")).upper()
        if sequence == "":
            continue
        yield record_id, sequence


def resolve_feature_gene_token(feature_id, feature_records, provider, cache, active):
    feature_text = str(feature_id or "").strip()
    if feature_text == "":
        return ""
    if feature_text in cache:
        return cache[feature_text]
    if feature_text in active:
        collapsed = collapse_transcript_suffix(provider, feature_text)
        cache[feature_text] = collapsed
        return collapsed

    record = feature_records.get(feature_text)
    if record is None:
        collapsed = collapse_transcript_suffix(provider, feature_text)
        cache[feature_text] = collapsed if collapsed != "" else feature_text
        return cache[feature_text]

    active.add(feature_text)
    attrs = record.get("attrs", {})
    direct = choose_first_gff_attribute(attrs, ("gene", "gene_id", "locus_tag", "geneName"))
    if direct == "":
        if record.get("feature_type") == "gene":
            direct = choose_first_gff_attribute(attrs, ("Name", "ID"))
        else:
            for parent_id in record.get("parents", ()):
                direct = resolve_feature_gene_token(parent_id, feature_records, provider, cache, active)
                if direct != "":
                    break
    if direct == "":
        direct = choose_first_gff_attribute(attrs, ("Name", "ID"))
    if direct == "":
        direct = collapse_transcript_suffix(provider, feature_text)
    active.remove(feature_text)
    cache[feature_text] = str(direct or "").strip()
    return cache[feature_text]


def merge_coordinate_intervals(intervals):
    ordered = sorted(
        (int(start), int(end))
        for start, end in intervals
        if int(start) <= int(end)
    )
    merged = []
    for start, end in ordered:
        if len(merged) == 0 or start > merged[-1][1] + 1:
            merged.append([start, end])
            continue
        if end > merged[-1][1]:
            merged[-1][1] = end
    return [(start, end) for start, end in merged]


def compute_interval_overlap_bases(left_intervals, right_intervals):
    overlap = 0
    left_index = 0
    right_index = 0
    while left_index < len(left_intervals) and right_index < len(right_intervals):
        left_start, left_end = left_intervals[left_index]
        right_start, right_end = right_intervals[right_index]
        start = max(left_start, right_start)
        end = min(left_end, right_end)
        if start <= end:
            overlap += end - start + 1
        if left_end <= right_end:
            left_index += 1
        else:
            right_index += 1
    return overlap


def transcript_feature_gene_token(features):
    counts = defaultdict(int)
    for feature in features:
        token = str(feature.get("gene_token", "") or "").strip()
        if token != "":
            counts[token] += 1
    if len(counts) == 0:
        return ""
    return sorted(counts.items(), key=lambda item: (-item[1], item[0]))[0][0]


def build_transcript_grouping_entry(provider, transcript_id, features):
    if len(features) == 0:
        return {
            "transcript_id": transcript_id,
            "gene_token": "",
            "seqid": "",
            "strand": "+",
            "merged_parts": (),
            "junctions": frozenset(),
            "cds_length": 0,
            "span_start": 0,
            "span_end": 0,
            "terminal_coord": 0,
            "rescue_ineligible": True,
            "collapsed_transcript_id": collapse_transcript_suffix(provider, transcript_id),
        }
    seqid = str(features[0].get("seqid", "") or "").strip()
    strand = str(features[0].get("strand", "") or "").strip() or "+"
    rescue_ineligible = False
    intervals = []
    for feature in features:
        feature_seqid = str(feature.get("seqid", "") or "").strip()
        feature_strand = str(feature.get("strand", "") or "").strip() or "+"
        if feature_seqid != seqid or feature_strand != strand:
            rescue_ineligible = True
        intervals.append((int(feature["start"]), int(feature["end"])))
    merged_parts = merge_coordinate_intervals(intervals)
    if len(merged_parts) == 0:
        rescue_ineligible = True
        span_start = 0
        span_end = 0
        terminal_coord = 0
        cds_length = 0
        junctions = frozenset()
    else:
        span_start = merged_parts[0][0]
        span_end = merged_parts[-1][1]
        terminal_coord = span_end if strand == "+" else span_start
        cds_length = sum(end - start + 1 for start, end in merged_parts)
        junctions = frozenset(
            (merged_parts[index][1], merged_parts[index + 1][0])
            for index in range(len(merged_parts) - 1)
        )
    return {
        "transcript_id": transcript_id,
        "gene_token": transcript_feature_gene_token(features),
        "seqid": seqid,
        "strand": strand,
        "merged_parts": tuple(merged_parts),
        "junctions": junctions,
        "cds_length": cds_length,
        "span_start": span_start,
        "span_end": span_end,
        "terminal_coord": terminal_coord,
        "rescue_ineligible": rescue_ineligible,
        "collapsed_transcript_id": collapse_transcript_suffix(provider, transcript_id),
    }


def should_rescue_overlapping_transcripts(left, right):
    if left["rescue_ineligible"] or right["rescue_ineligible"]:
        return False
    if left["seqid"] != right["seqid"] or left["strand"] != right["strand"]:
        return False
    if left["cds_length"] <= 0 or right["cds_length"] <= 0:
        return False
    if left["span_end"] < right["span_start"] or right["span_end"] < left["span_start"]:
        return False
    overlap_bases = compute_interval_overlap_bases(left["merged_parts"], right["merged_parts"])
    if overlap_bases <= 0:
        return False
    shorter_overlap = overlap_bases / float(min(left["cds_length"], right["cds_length"]))
    longer_overlap = overlap_bases / float(max(left["cds_length"], right["cds_length"]))
    shared_junctions = len(left["junctions"].intersection(right["junctions"]))
    if (
        shared_junctions > 0
        and shorter_overlap >= RESCUE_SHARED_JUNCTION_MIN_SHORTER_OVERLAP
        and longer_overlap >= RESCUE_SHARED_JUNCTION_MIN_LONGER_OVERLAP
    ):
        return True
    if (
        left["terminal_coord"] == right["terminal_coord"]
        and shorter_overlap >= RESCUE_SAME_TERMINAL_MIN_SHORTER_OVERLAP
        and longer_overlap >= RESCUE_SAME_TERMINAL_MIN_LONGER_OVERLAP
    ):
        return True
    return False


def choose_rescue_cluster_gene_token(provider, entries):
    collapsed_transcript_ids = sorted(
        {
            str(entry.get("collapsed_transcript_id", "") or "").strip()
            for entry in entries
            if str(entry.get("collapsed_transcript_id", "") or "").strip() != ""
        }
    )
    if len(collapsed_transcript_ids) == 1:
        return collapsed_transcript_ids[0]
    gene_tokens = sorted(
        {
            str(entry.get("gene_token", "") or "").strip()
            for entry in entries
            if str(entry.get("gene_token", "") or "").strip() != ""
        }
    )
    if len(gene_tokens) > 0:
        return gene_tokens[0]
    if len(collapsed_transcript_ids) > 0:
        return collapsed_transcript_ids[0]
    transcript_ids = sorted(
        str(entry.get("transcript_id", "") or "").strip()
        for entry in entries
        if str(entry.get("transcript_id", "") or "").strip() != ""
    )
    if len(transcript_ids) > 0:
        fallback = collapse_transcript_suffix(provider, transcript_ids[0])
        return fallback if fallback != "" else transcript_ids[0]
    return ""


def build_rescued_gene_tokens_for_transcripts(task, cds_features_by_transcript):
    resolved = {
        transcript_id: transcript_feature_gene_token(features)
        for transcript_id, features in cds_features_by_transcript.items()
    }
    if gene_grouping_mode_for_task(task) != "rescue_overlap":
        return resolved

    provider = task["provider"]
    entries_by_id = {}
    buckets = defaultdict(list)
    for transcript_id, features in cds_features_by_transcript.items():
        entry = build_transcript_grouping_entry(provider, transcript_id, features)
        entries_by_id[transcript_id] = entry
        buckets[(entry["seqid"], entry["strand"])].append(entry)

    adjacency = defaultdict(set)
    for bucket_entries in buckets.values():
        ordered = sorted(
            bucket_entries,
            key=lambda item: (item["span_start"], item["span_end"], item["transcript_id"]),
        )
        for index, left in enumerate(ordered):
            for right in ordered[index + 1 :]:
                if right["span_start"] > left["span_end"]:
                    break
                if left["gene_token"] == right["gene_token"]:
                    continue
                if should_rescue_overlapping_transcripts(left, right):
                    adjacency[left["transcript_id"]].add(right["transcript_id"])
                    adjacency[right["transcript_id"]].add(left["transcript_id"])

    visited = set()
    for transcript_id in sorted(entries_by_id.keys()):
        if transcript_id in visited:
            continue
        stack = [transcript_id]
        component_ids = []
        while len(stack) > 0:
            current = stack.pop()
            if current in visited:
                continue
            visited.add(current)
            component_ids.append(current)
            for neighbor in sorted(adjacency.get(current, ())):
                if neighbor not in visited:
                    stack.append(neighbor)
        if len(component_ids) <= 1:
            continue
        component_entries = [entries_by_id[current_id] for current_id in sorted(component_ids)]
        cluster_gene_token = choose_rescue_cluster_gene_token(provider, component_entries)
        if cluster_gene_token == "":
            continue
        if not any(
            str(entry.get("gene_token", "") or "").strip() != cluster_gene_token
            for entry in component_entries
        ):
            continue
        for entry in component_entries:
            resolved[entry["transcript_id"]] = cluster_gene_token
    return resolved


def require_biopython_for_gbff(path):
    if SeqIO is None:
        raise RuntimeError(
            "Biopython is required to process GBFF/GenBank inputs: {}".format(path)
        )


def iter_genbank_records(path):
    require_biopython_for_gbff(path)
    with open_text(path, "rt") as handle:
        for record in SeqIO.parse(handle, "genbank"):
            yield record


def choose_first_feature_qualifier(feature, keys):
    qualifiers = getattr(feature, "qualifiers", {}) or {}
    for key in keys:
        values = qualifiers.get(key, ())
        if isinstance(values, str):
            values = (values,)
        for value in values:
            text = str(value or "").strip()
            if text != "":
                return text
    return ""


def extract_genbank_gene_id_from_dbxref(feature):
    qualifiers = getattr(feature, "qualifiers", {}) or {}
    values = qualifiers.get("db_xref", ())
    if isinstance(values, str):
        values = (values,)
    for raw in values:
        text = str(raw or "").strip()
        if text.startswith("GeneID:"):
            candidate = text.split(":", 1)[1].strip()
            if candidate != "":
                return "GeneID{}".format(candidate)
    return ""


def normalize_genbank_feature_location(feature):
    location = getattr(feature, "location", None)
    if location is None:
        return []
    if hasattr(location, "parts") and len(getattr(location, "parts", ())) > 0:
        parts = list(location.parts)
    else:
        parts = [location]
    normalized = []
    for part in parts:
        try:
            start = int(part.start) + 1
            end = int(part.end)
        except Exception:
            continue
        if start > end:
            start, end = end, start
        normalized.append((start, end))
    return normalized


def get_genbank_feature_strand(feature):
    strand = getattr(getattr(feature, "location", None), "strand", None)
    return "-" if strand == -1 else "+"


def infer_genbank_gene_token(feature, fallback_prefix, fallback_index):
    gene_token = choose_first_feature_qualifier(
        feature,
        ("locus_tag", "gene", "gene_synonym", "Name", "ID"),
    )
    if gene_token == "":
        gene_token = extract_genbank_gene_id_from_dbxref(feature)
    if gene_token == "":
        gene_token = "{}_gene{}".format(fallback_prefix, fallback_index)
    return sanitize_identifier(gene_token)


def infer_genbank_transcript_token(feature, gene_token, fallback_index):
    transcript_token = choose_first_feature_qualifier(
        feature,
        ("protein_id", "transcript_id", "mrna", "rna_id"),
    )
    if transcript_token == "":
        transcript_token = "{}.t{}".format(gene_token, fallback_index)
    return sanitize_identifier(transcript_token)


def parse_genbank_codon_start(feature):
    raw = choose_first_feature_qualifier(feature, ("codon_start",))
    if raw == "":
        return 1
    try:
        value = int(raw)
    except Exception:
        return 1
    if value not in (1, 2, 3):
        return 1
    return value


def extract_seqrecord_id(record):
    candidate = first_token(apply_common_replacements(str(getattr(record, "id", "") or "").strip()))
    if candidate != "":
        return candidate
    candidate = first_token(apply_common_replacements(str(getattr(record, "name", "") or "").strip()))
    if candidate != "":
        return candidate
    return "unnamed_record"


def format_gff_attributes(attributes):
    fields = []
    for key, value in attributes:
        text = str(value or "").strip()
        if text == "":
            continue
        fields.append("{}={}".format(key, quote(text, safe="._:-|")))
    return ";".join(fields) if len(fields) > 0 else "."


def compute_gff_phases(parts, strand, codon_start):
    if len(parts) == 0:
        return []
    transcript_order = list(parts)
    if strand == "-":
        transcript_order = list(reversed(transcript_order))
    phases_by_part = {}
    effective_bases = 0
    for index, part in enumerate(transcript_order):
        length = part["end"] - part["start"] + 1
        phase = codon_start - 1 if index == 0 else (3 - (effective_bases % 3)) % 3
        phases_by_part[(part["start"], part["end"])] = str(phase)
        effective_bases += max(0, length - phase)
    return [phases_by_part[(part["start"], part["end"])] for part in parts]


def iter_gbff_coding_entries(path):
    gene_entries = {}
    transcripts_by_gene = defaultdict(list)
    fallback_gene_index = 0
    transcript_counts_by_gene = defaultdict(int)

    for record in iter_genbank_records(path):
        seqid = extract_seqrecord_id(record)
        for feature in getattr(record, "features", ()):
            feature_type = str(getattr(feature, "type", "") or "").lower()
            parts = normalize_genbank_feature_location(feature)
            if len(parts) == 0:
                continue
            strand = get_genbank_feature_strand(feature)
            if feature_type == "gene":
                fallback_gene_index += 1
                gene_token = infer_genbank_gene_token(feature, seqid, fallback_gene_index)
                gene_key = (seqid, gene_token)
                gene_name = choose_first_feature_qualifier(feature, ("gene", "locus_tag"))
                start = min(start for start, _end in parts)
                end = max(end for _start, end in parts)
                existing = gene_entries.get(gene_key)
                if existing is None:
                    gene_entries[gene_key] = {
                        "seqid": seqid,
                        "gene_token": gene_token,
                        "gene_name": gene_name,
                        "start": start,
                        "end": end,
                        "strand": strand,
                    }
                else:
                    existing["start"] = min(existing["start"], start)
                    existing["end"] = max(existing["end"], end)
                    if existing["gene_name"] == "" and gene_name != "":
                        existing["gene_name"] = gene_name
                continue
            if feature_type != "cds":
                continue
            fallback_gene_index += 1
            gene_token = infer_genbank_gene_token(feature, seqid, fallback_gene_index)
            gene_key = (seqid, gene_token)
            gene_name = choose_first_feature_qualifier(feature, ("gene", "locus_tag"))
            transcript_counts_by_gene[gene_key] += 1
            transcript_token = infer_genbank_transcript_token(
                feature,
                gene_token,
                transcript_counts_by_gene[gene_key],
            )
            sequence = str(feature.extract(record.seq)).upper()
            codon_start = parse_genbank_codon_start(feature)
            if codon_start > 1:
                sequence = sequence[codon_start - 1 :]
            if sequence == "":
                continue
            start = min(start for start, _end in parts)
            end = max(end for _start, end in parts)
            existing = gene_entries.get(gene_key)
            if existing is None:
                gene_entries[gene_key] = {
                    "seqid": seqid,
                    "gene_token": gene_token,
                    "gene_name": gene_name,
                    "start": start,
                    "end": end,
                    "strand": strand,
                }
            else:
                existing["start"] = min(existing["start"], start)
                existing["end"] = max(existing["end"], end)
                if existing["gene_name"] == "" and gene_name != "":
                    existing["gene_name"] = gene_name
            transcript_entry = {
                "seqid": seqid,
                "gene_key": gene_key,
                "gene_token": gene_token,
                "gene_name": gene_name,
                "transcript_token": transcript_token,
                "start": start,
                "end": end,
                "strand": strand,
                "parts": [{"start": part_start, "end": part_end} for part_start, part_end in parts],
                "codon_start": codon_start,
                "sequence": sequence,
            }
            transcripts_by_gene[gene_key].append(transcript_entry)

    for gene_key, gene_entry in gene_entries.items():
        transcripts = transcripts_by_gene.get(gene_key, [])
        yield gene_entry, sorted(
            transcripts,
            key=lambda item: (item["start"], item["end"], item["transcript_token"]),
        )


def derive_cds_records_from_gff_and_genome(task):
    gff_path = task.get("gff_path")
    genome_path = task.get("genome_path")
    if gff_path is None or genome_path is None:
        raise ValueError("GFF+genome inputs are required to derive CDS for {}".format(task.get("species_key", "")))

    genome_sequences = load_genome_sequences(genome_path)
    feature_records = {}
    cds_features_by_transcript = defaultdict(list)
    utr_features_by_transcript = defaultdict(list)
    gene_cache = {}

    with open_text(gff_path, "rt") as handle:
        for raw_line in handle:
            line = raw_line.rstrip("\n\r")
            if line == "":
                continue
            if line.startswith("##FASTA"):
                break
            if line.startswith("#"):
                continue
            parts = line.split("\t")
            if len(parts) < 9:
                continue
            seqid, _source, feature_type, start_text, end_text, _score, strand, phase_text, attr_text = parts[:9]
            feature_type_lower = str(feature_type or "").strip().lower()
            attrs = parse_gff_attributes(attr_text)
            feature_id = choose_first_gff_attribute(attrs, ("ID", "transcript_id", "protein_id", "Name"))
            parents = attrs.get("Parent", ())
            if feature_id != "":
                feature_records[feature_id] = {
                    "feature_type": feature_type_lower,
                    "parents": parents,
                    "attrs": attrs,
                }
            if feature_type_lower in ("five_prime_utr", "three_prime_utr"):
                try:
                    start = int(start_text)
                    end = int(end_text)
                except Exception:
                    continue
                if start > end:
                    start, end = end, start
                for parent_id in parents:
                    if str(parent_id).strip() == "":
                        continue
                    utr_features_by_transcript[parent_id].append(
                        {
                            "seqid": str(seqid or "").strip(),
                            "start": start,
                            "end": end,
                            "strand": str(strand or "").strip() or "+",
                        }
                    )
                continue
            if feature_type_lower != "cds":
                continue
            try:
                start = int(start_text)
                end = int(end_text)
            except Exception:
                continue
            if start > end:
                start, end = end, start
            transcript_ids = [value for value in parents if str(value).strip() != ""]
            if len(transcript_ids) == 0:
                fallback_id = feature_id
                if fallback_id == "":
                    fallback_id = choose_first_gff_attribute(attrs, ("transcript_id", "protein_id", "Name"))
                if fallback_id == "":
                    fallback_id = "{}:{}-{}".format(seqid, start, end)
                transcript_ids = [fallback_id]
            explicit_gene = choose_first_gff_attribute(attrs, ("gene", "gene_id", "locus_tag", "geneName"))
            for transcript_id in transcript_ids:
                gene_token = explicit_gene
                if gene_token == "":
                    gene_token = resolve_feature_gene_token(
                        transcript_id,
                        feature_records,
                        task["provider"],
                        gene_cache,
                        set(),
                    )
                if gene_token == "":
                    gene_token = collapse_transcript_suffix(task["provider"], transcript_id)
                cds_features_by_transcript[transcript_id].append(
                    {
                        "seqid": str(seqid or "").strip(),
                        "start": start,
                        "end": end,
                        "strand": str(strand or "").strip() or "+",
                        "gene_token": gene_token,
                    }
                )

    rescued_gene_tokens = build_rescued_gene_tokens_for_transcripts(task, cds_features_by_transcript)
    for transcript_id in sorted(cds_features_by_transcript.keys()):
        features = cds_features_by_transcript[transcript_id]
        if len(features) == 0:
            continue
        utr_features = utr_features_by_transcript.get(transcript_id, ())
        strand = features[0]["strand"]
        transcript_gene_token = rescued_gene_tokens.get(transcript_id, "") or transcript_feature_gene_token(features)
        trimmed_features = []
        for feature in features:
            segments = [(feature["start"], feature["end"])]
            for utr in utr_features:
                if (
                    utr["seqid"] != feature["seqid"]
                    or utr["strand"] != feature["strand"]
                    or utr["end"] < feature["start"]
                    or feature["end"] < utr["start"]
                ):
                    continue
                next_segments = []
                for seg_start, seg_end in segments:
                    if utr["end"] < seg_start or seg_end < utr["start"]:
                        next_segments.append((seg_start, seg_end))
                        continue
                    if seg_start < utr["start"]:
                        next_segments.append((seg_start, utr["start"] - 1))
                    if utr["end"] < seg_end:
                        next_segments.append((utr["end"] + 1, seg_end))
                segments = [(seg_start, seg_end) for seg_start, seg_end in next_segments if seg_start <= seg_end]
                if len(segments) == 0:
                    break
            for seg_start, seg_end in segments:
                trimmed_features.append(
                    {
                        "seqid": feature["seqid"],
                        "start": seg_start,
                        "end": seg_end,
                        "strand": feature["strand"],
                        "gene_token": transcript_gene_token,
                    }
                )
        ordered = sorted(trimmed_features, key=lambda item: item["start"], reverse=(strand == "-"))
        pieces = []
        gene_token = ""
        for feature in ordered:
            if feature["strand"] != strand:
                raise ValueError(
                    "Mixed-strand CDS features for transcript '{}' in {}".format(transcript_id, gff_path)
                )
            seqid = feature["seqid"]
            genome_seq = genome_sequences.get(seqid, "")
            if genome_seq == "":
                raise ValueError(
                    "Genome FASTA for {} is missing sequence '{}' required by {}".format(
                        task.get("species_key", ""),
                        seqid,
                        gff_path,
                    )
                )
            start_idx = feature["start"] - 1
            end_idx = feature["end"]
            piece = genome_seq[start_idx:end_idx]
            # GFF3 phase annotates codon frame continuity; it does not indicate
            # bases to trim from the CDS feature when reconstructing the spliced
            # CDS sequence. We therefore concatenate the full CDS spans in
            # transcript order and leave any frame repair to the final padding
            # step used elsewhere in the pipeline.
            if strand == "-":
                piece = reverse_complement(piece)
            if piece != "":
                pieces.append(piece)
            if gene_token == "":
                gene_token = feature["gene_token"]
        sequence = "".join(pieces)
        if sequence == "":
            continue
        header = str(transcript_id or "").strip()
        if gene_token != "":
            header = "{} gene={}".format(header, gene_token)
        yield header, sequence


def derive_cds_records_from_gbff(task):
    gbff_path = task.get("gbff_path")
    if gbff_path is None:
        raise ValueError("GBFF input is required to derive CDS for {}".format(task.get("species_key", "")))
    for _gene_entry, transcripts in iter_gbff_coding_entries(gbff_path):
        for transcript in transcripts:
            header = transcript["transcript_token"]
            if transcript["gene_token"] != "":
                header = "{} gene={}".format(header, transcript["gene_token"])
            yield header, transcript["sequence"]


def iter_gff_lines_from_gbff(task):
    gbff_path = task.get("gbff_path")
    if gbff_path is None:
        raise ValueError("GBFF input is required to derive GFF for {}".format(task.get("species_key", "")))
    yield "##gff-version 3\n"
    for gene_entry, transcripts in sorted(
        iter_gbff_coding_entries(gbff_path),
        key=lambda item: (item[0]["seqid"], item[0]["start"], item[0]["gene_token"]),
    ):
        seqid = gene_entry["seqid"]
        strand = gene_entry["strand"]
        gene_id = "gene:{}".format(gene_entry["gene_token"])
        gene_attrs = [("ID", gene_id), ("Name", gene_entry["gene_name"] or gene_entry["gene_token"])]
        yield "\t".join(
            [
                seqid,
                "genbank_derived",
                "gene",
                str(gene_entry["start"]),
                str(gene_entry["end"]),
                ".",
                strand,
                ".",
                format_gff_attributes(gene_attrs),
            ]
        ) + "\n"
        for transcript in transcripts:
            transcript_id = "transcript:{}".format(transcript["transcript_token"])
            transcript_attrs = [
                ("ID", transcript_id),
                ("Parent", gene_id),
                ("Name", transcript["transcript_token"]),
            ]
            yield "\t".join(
                [
                    seqid,
                    "genbank_derived",
                    "mRNA",
                    str(transcript["start"]),
                    str(transcript["end"]),
                    ".",
                    transcript["strand"],
                    ".",
                    format_gff_attributes(transcript_attrs),
                ]
            ) + "\n"
            ordered_parts = sorted(transcript["parts"], key=lambda item: item["start"])
            phases = compute_gff_phases(ordered_parts, transcript["strand"], transcript["codon_start"])
            for index, part in enumerate(ordered_parts, start=1):
                cds_attrs = [
                    ("ID", "cds:{}.{}".format(transcript["transcript_token"], index)),
                    ("Parent", transcript_id),
                ]
                yield "\t".join(
                    [
                        seqid,
                        "genbank_derived",
                        "CDS",
                        str(part["start"]),
                        str(part["end"]),
                        ".",
                        transcript["strand"],
                        phases[index - 1],
                        format_gff_attributes(cds_attrs),
                    ]
                ) + "\n"


def _iter_fasta_records_from_handle(handle):
    header = None
    chunks = []
    for raw_line in handle:
        line = raw_line.rstrip("\n\r")
        if line == "":
            continue
        if line.startswith(">"):
            if header is not None:
                yield header, "".join(chunks)
            header = line[1:]
            chunks = []
            continue
        chunks.append(re.sub(r"\s+", "", line))
    if header is not None:
        yield header, "".join(chunks)


def iter_fasta_records(path):
    path_lower = path.name.lower()
    if any(path_lower.endswith(suffix) for suffix in FASTA_ARCHIVE_EXTENSIONS):
        with tarfile.open(path, "r:*") as archive:
            members = [member for member in archive.getmembers() if member.isfile()]
            fasta_members = [
                member for member in members
                if is_fasta_filename(Path(member.name).name)
            ]
            if len(fasta_members) == 0 and len(members) == 1:
                fasta_members = members
            for member in fasta_members:
                extracted = archive.extractfile(member)
                if extracted is None:
                    continue
                with io.TextIOWrapper(extracted, encoding="utf-8") as handle:
                    yield from _iter_fasta_records_from_handle(handle)
        return

    with open_text(path, "rt") as handle:
        yield from _iter_fasta_records_from_handle(handle)


def count_fasta_records(path):
    count = 0
    first_name = ""
    for header, _ in iter_fasta_records(path):
        count += 1
        if first_name == "":
            first_name = first_token(header)
    return count, first_name


def write_fasta_record(handle, record_id, sequence, width=80):
    handle.write(">{}\n".format(record_id))
    for i in range(0, len(sequence), width):
        handle.write(sequence[i:i + width] + "\n")


def extract_ensembl_id(header):
    match = re.search(r"(?:^|\s)gene:([^\s]+)", header)
    if match:
        return match.group(1)
    return first_token(header)


def extract_phycocosm_id(header):
    token = first_token(header)
    if token.startswith("jgi|"):
        parts = [part for part in token.split("|") if part != ""]
        if len(parts) >= 2:
            return parts[-1]
    if "|" in token:
        return token.split("|")[-1]
    return token


def extract_phytozome_id(header):
    return first_token(header)


def extract_gwh_id(header):
    for tag in ("Gene", "OriGeneID"):
        candidate = extract_header_tag_value(header, tag)
        if candidate != "":
            return candidate
    return first_token(header)


def extract_provider_id(provider, header):
    if provider in ("ensembl", "ensemblplants"):
        return extract_ensembl_id(header)
    if provider == "phycocosm":
        return extract_phycocosm_id(header)
    if provider == "gwh":
        return extract_gwh_id(header)
    if provider in ("ncbi", "refseq", "genbank"):
        return first_token(header)
    return extract_phytozome_id(header)


def extract_provider_transcript_id(provider, header):
    token = first_token(header)
    if provider in ("ensembl", "ensemblplants"):
        token = re.sub(r"^(?:transcript|mrna|rna):", "", token, flags=re.IGNORECASE)
        if token != "":
            return token
    extracted = extract_provider_id(provider, header)
    if extracted != "":
        return extracted
    return token


def collapse_transcript_suffix(provider, identifier):
    text = identifier
    if provider == "phytozome":
        text = re.sub(r"[.][0-9]+$", "", text)
    if provider in (
        "phycocosm",
        "phytozome",
        "ensembl",
        "ensemblplants",
        "ncbi",
        "refseq",
        "genbank",
        "coge",
        "cngb",
        "gwh",
        "flybase",
        "wormbase",
        "vectorbase",
        "fernbase",
        "insectbase",
        "direct",
        "local",
    ):
        text = re.sub(r"[._-]t[0-9]+$", "", text, flags=re.IGNORECASE)
        text = re.sub(r"[._-](?:transcript|mrna|rna|isoform)[._-]?[0-9]+$", "", text, flags=re.IGNORECASE)
        text = re.sub(r"[._-]amt[0-9]+$", "", text, flags=re.IGNORECASE)
    return text


def normalize_gene_grouping_mode(value):
    text = str(value or "").strip().lower()
    if text == "":
        return "strict"
    if text not in GENE_GROUPING_MODES:
        raise ValueError(
            "Unknown gene_grouping_mode: {} (allowed: {})".format(
                text,
                ",".join(GENE_GROUPING_MODES),
            )
        )
    return text


def gene_grouping_mode_for_task(task):
    return normalize_gene_grouping_mode(task.get("gene_grouping_mode", "strict"))


def task_missing_annotation_label(cds_path, gff_path, gbff_path, genome_path):
    if task_has_usable_source_bundle(cds_path, gff_path, gbff_path, genome_path):
        return ""
    if cds_path is None and gff_path is None and gbff_path is None and genome_path is None:
        return "CDS-or-GBFF-or-(GFF+genome)"
    if gff_path is not None and genome_path is None:
        return "CDS-or-GBFF-or-genome"
    if genome_path is not None and gff_path is None and gbff_path is None:
        return "CDS-or-GBFF-or-GFF"
    if gbff_path is None:
        return "CDS-or-GBFF-or-(GFF+genome)"
    return ""


def describe_task_cds_input(task):
    cds_path = task.get("cds_path")
    if cds_path is not None:
        return cds_path
    gff_path = task.get("gff_path")
    gbff_path = task.get("gbff_path")
    genome_path = task.get("genome_path")
    if gff_path is not None and genome_path is not None:
        return "{} + {} (derived CDS)".format(gff_path, genome_path)
    if gbff_path is not None and genome_path is not None:
        return "{} + {} (derived CDS)".format(gbff_path, genome_path)
    if gbff_path is not None:
        return "{} (derived CDS)".format(gbff_path)
    return ""


def describe_task_gff_input(task):
    gff_path = task.get("gff_path")
    if gff_path is not None:
        return gff_path
    gbff_path = task.get("gbff_path")
    if gbff_path is not None:
        return "{} (derived GFF)".format(gbff_path)
    return ""


def describe_task_genome_input(task):
    genome_path = task.get("genome_path")
    if genome_path is not None:
        return genome_path
    gbff_path = task.get("gbff_path")
    if gbff_path is not None:
        return "{} (derived genome)".format(gbff_path)
    return ""


def build_derived_cds_output_basename(task):
    gff_path = task.get("gff_path")
    species_prefix = task["species_prefix"]
    if gff_path is not None:
        normalized = normalize_output_basename(gff_path.name, species_prefix)
        stem = strip_suffix_case_insensitive(normalized, GFF_EXTENSIONS)
        return stem + ".derived.cds.fa.gz"
    gbff_path = task.get("gbff_path")
    if gbff_path is not None:
        normalized = normalize_output_basename(gbff_path.name, species_prefix)
        stem = strip_suffix_case_insensitive(normalized, GENBANK_EXTENSIONS)
        return stem + ".derived.cds.fa.gz"
    genome_path = task.get("genome_path")
    if genome_path is not None:
        normalized = normalize_output_basename(genome_path.name, species_prefix)
        stem = strip_suffix_case_insensitive(normalized, FASTA_EXTENSIONS)
        return stem + ".derived.cds.fa.gz"
    return species_prefix + ".derived.cds.fa.gz"


def build_derived_gff_output_basename(task):
    gbff_path = task.get("gbff_path")
    species_prefix = task["species_prefix"]
    if gbff_path is not None:
        normalized = normalize_output_basename(gbff_path.name, species_prefix)
        stem = strip_suffix_case_insensitive(normalized, GENBANK_EXTENSIONS)
        return stem + ".derived.gff.gz"
    return species_prefix + ".derived.gff.gz"


def build_derived_genome_output_basename(task):
    gbff_path = task.get("gbff_path")
    species_prefix = task["species_prefix"]
    if gbff_path is not None:
        normalized = normalize_output_basename(gbff_path.name, species_prefix)
        stem = strip_suffix_case_insensitive(normalized, GENBANK_EXTENSIONS)
        return stem + ".derived.genome.fa.gz"
    return species_prefix + ".derived.genome.fa.gz"


def iter_task_cds_records(task):
    cds_path = task.get("cds_path")
    if cds_path is not None:
        yield from iter_fasta_records(cds_path)
        return
    gff_path = task.get("gff_path")
    genome_path = task.get("genome_path")
    gbff_path = task.get("gbff_path")
    if gff_path is not None and genome_path is not None:
        yield from derive_cds_records_from_gff_and_genome(task)
        return
    if gbff_path is not None:
        yield from derive_cds_records_from_gbff(task)
        return
    raise ValueError("No CDS source is available for {}".format(task.get("species_key", "")))


def discover_generic_species_dir_tasks(provider, input_dir):
    warnings = []
    errors = []
    tasks = []

    for species_dir in sorted(input_dir.iterdir()):
        if not species_dir.is_dir():
            continue
        species_key = species_dir.name
        species_prefix = species_prefix_from_value(species_key)
        if species_prefix == "":
            warnings.append(
                "[{}] skipped '{}': unable to parse species prefix.".format(provider, species_key)
            )
            continue

        files = [path for path in species_dir.iterdir() if path.is_file()]
        cds_matches = [
            path
            for path in files
            if is_probable_cds_filename(provider, path.name)
        ]
        if len(cds_matches) == 0:
            cds_matches = [
                path
                for path in files
                if is_fasta_filename(path.name) and not is_probable_genome_filename(provider, path.name)
            ]
        gff_matches = [path for path in files if is_gff_filename(path.name)]
        gbff_matches = [path for path in files if is_gbff_filename(path.name)]
        genome_matches = [path for path in files if is_probable_genome_filename(provider, path.name)]

        cds_path = pick_single_file(cds_matches, provider, species_key, "CDS", warnings)
        gff_path = pick_single_file(gff_matches, provider, species_key, "GFF", warnings)
        gbff_path = pick_single_file(gbff_matches, provider, species_key, "GBFF", warnings)
        genome_path = pick_single_file(genome_matches, provider, species_key, "genome", warnings)

        missing_label = task_missing_annotation_label(cds_path, gff_path, gbff_path, genome_path)
        if missing_label != "":
            errors.append(
                "[{}] {}: missing {}".format(
                    provider, species_key, missing_label
                )
            )
            continue
        tasks.append(
            {
                "provider": provider,
                "species_key": species_key,
                "species_prefix": species_prefix,
                "cds_path": cds_path,
                "gff_path": gff_path,
                "gbff_path": gbff_path,
                "genome_path": genome_path,
            }
        )
    return tasks, warnings, errors


def build_gene_aggregate_id(task, header, transcript_id):
    provider = task["provider"]
    species_prefix = task["species_prefix"]
    if provider in ("ncbi", "refseq", "genbank"):
        ensembl_gene_id = extract_ncbi_ensembl_gene_id_from_header(header)
        gene_symbol = extract_header_tag_value(header, "gene")
        gene_id = extract_ncbi_gene_id_from_header(header)
        locus_tag = extract_header_tag_value(header, "locus_tag")
        protein_id = extract_header_tag_value(header, "protein_id")
        if ensembl_gene_id != "":
            gene_token = ensembl_gene_id
        elif gene_id != "":
            gene_token = "GeneID{}".format(gene_id)
        elif gene_symbol != "":
            gene_token = gene_symbol
        elif locus_tag != "":
            gene_token = locus_tag
        elif protein_id != "":
            gene_token = protein_id
        else:
            gene_token = extract_provider_id(provider, header)
    else:
        gene_symbol = extract_header_tag_value(header, "gene")
        extracted = extract_provider_id(provider, header)
        collapsed = collapse_transcript_suffix(provider, extracted)
        if gene_symbol != "":
            if (
                collapsed != ""
                and collapsed != gene_symbol
                and re.search(r"(?:^|[._-]){}$".format(re.escape(gene_symbol)), collapsed, re.IGNORECASE)
            ):
                gene_token = collapsed
            else:
                gene_token = gene_symbol
        else:
            gene_token = collapsed if collapsed != "" else extracted
    prefixed = "{}_{}".format(species_prefix, sanitize_identifier(gene_token))
    return sanitize_identifier(prefixed)


def pick_single_file(matches, provider, species_key, label, warnings):
    if len(matches) == 0:
        return None
    ordered = sorted(matches, key=lambda path: provider_candidate_sort_key(provider, label, path.name))
    if len(ordered) > 1:
        warnings.append(
            "[{}] {}: multiple {} files found. Using '{}'".format(
                provider, species_key, label, ordered[0].name
            )
        )
    return ordered[0]


def discover_ensembl_like_tasks(input_dir, provider):
    warnings = []
    errors = []
    tasks = []

    cds_by_species = defaultdict(list)
    gff_by_species = defaultdict(list)
    gbff_by_species = defaultdict(list)
    genome_by_species = defaultdict(list)

    for path in sorted(input_dir.iterdir()):
        if not path.is_file():
            continue
        name = path.name
        species_key = name.split(".", 1)[0]
        if species_key == "":
            continue
        if is_fasta_filename(name) and ENSEMBL_CDS_PATTERN.search(name):
            cds_by_species[species_key].append(path)
            continue
        if is_gff_filename(name):
            gff_by_species[species_key].append(path)
            continue
        if is_gbff_filename(name):
            gbff_by_species[species_key].append(path)
            continue
        if is_probable_genome_filename(provider, name):
            genome_by_species[species_key].append(path)

    for species_key in sorted(
        set(cds_by_species.keys())
        | set(gff_by_species.keys())
        | set(gbff_by_species.keys())
        | set(genome_by_species.keys())
    ):
        species_prefix = species_prefix_from_value(species_key)
        if species_prefix == "":
            warnings.append(
                "[{}] skipped '{}': unable to parse species prefix.".format(provider, species_key)
            )
            continue

        cds_path = pick_single_file(
            cds_by_species.get(species_key, []),
            provider,
            species_key,
            "CDS",
            warnings,
        )
        gff_path = pick_single_file(
            gff_by_species.get(species_key, []),
            provider,
            species_key,
            "GFF",
            warnings,
        )
        gbff_path = pick_single_file(
            gbff_by_species.get(species_key, []),
            provider,
            species_key,
            "GBFF",
            warnings,
        )
        genome_path = pick_single_file(
            genome_by_species.get(species_key, []),
            provider,
            species_key,
            "genome",
            warnings,
        )
        missing_label = task_missing_annotation_label(cds_path, gff_path, gbff_path, genome_path)
        if missing_label != "":
            errors.append(
                "[{}] {}: missing {}".format(
                    provider,
                    species_key, missing_label
                )
            )
            continue
        tasks.append(
            {
                "provider": provider,
                "species_key": species_key,
                "species_prefix": species_prefix,
                "cds_path": cds_path,
                "gff_path": gff_path,
                "gbff_path": gbff_path,
                "genome_path": genome_path,
            }
        )

    return tasks, warnings, errors


def discover_phycocosm_tasks(input_dir):
    warnings = []
    errors = []
    tasks = []

    for species_dir in sorted(input_dir.iterdir()):
        if not species_dir.is_dir():
            continue
        species_key = species_dir.name
        species_prefix = species_prefix_from_value(species_key)
        if species_prefix == "":
            warnings.append(
                "[phycocosm] skipped '{}': unable to parse species prefix.".format(species_key)
            )
            continue

        files = [path for path in species_dir.iterdir() if path.is_file()]
        cds_matches = [path for path in files if "fasta" in path.name.lower() and is_fasta_filename(path.name)]
        gff_matches = [path for path in files if "gff" in path.name.lower() and is_gff_filename(path.name)]
        gbff_matches = [path for path in files if is_gbff_filename(path.name)]
        genome_matches = [path for path in files if is_probable_genome_filename("phycocosm", path.name)]

        cds_path = pick_single_file(cds_matches, "phycocosm", species_key, "CDS", warnings)
        gff_path = pick_single_file(gff_matches, "phycocosm", species_key, "GFF", warnings)
        gbff_path = pick_single_file(gbff_matches, "phycocosm", species_key, "GBFF", warnings)
        genome_path = pick_single_file(genome_matches, "phycocosm", species_key, "genome", warnings)

        missing_label = task_missing_annotation_label(cds_path, gff_path, gbff_path, genome_path)
        if missing_label != "":
            errors.append(
                "[phycocosm] {}: missing {}".format(
                    species_key, missing_label
                )
            )
            continue
        tasks.append(
            {
                "provider": "phycocosm",
                "species_key": species_key,
                "species_prefix": species_prefix,
                "cds_path": cds_path,
                "gff_path": gff_path,
                "gbff_path": gbff_path,
                "genome_path": genome_path,
            }
        )

    return tasks, warnings, errors


def discover_phytozome_tasks(input_dir):
    warnings = []
    errors = []
    tasks = []

    for species_dir in sorted(input_dir.iterdir()):
        if not species_dir.is_dir():
            continue
        species_key = species_dir.name
        species_prefix = species_prefix_from_value(species_key)
        if species_prefix == "":
            warnings.append(
                "[phytozome] skipped '{}': unable to parse species prefix.".format(species_key)
            )
            continue

        files = [path for path in species_dir.iterdir() if path.is_file()]
        cds_matches = [
            path
            for path in files
            if ("cds_" in path.name.lower() or ".cds." in path.name.lower()) and is_fasta_filename(path.name)
        ]
        gff_matches = [path for path in files if "gene.gff3" in path.name.lower() and is_gff_filename(path.name)]
        gbff_matches = [path for path in files if is_gbff_filename(path.name)]
        genome_matches = [path for path in files if is_probable_genome_filename("phytozome", path.name)]

        cds_path = pick_single_file(cds_matches, "phytozome", species_key, "CDS", warnings)
        gff_path = pick_single_file(gff_matches, "phytozome", species_key, "GFF", warnings)
        gbff_path = pick_single_file(gbff_matches, "phytozome", species_key, "GBFF", warnings)
        genome_path = pick_single_file(genome_matches, "phytozome", species_key, "genome", warnings)

        missing_label = task_missing_annotation_label(cds_path, gff_path, gbff_path, genome_path)
        if missing_label != "":
            errors.append(
                "[phytozome] {}: missing {}".format(
                    species_key, missing_label
                )
            )
            continue
        tasks.append(
            {
                "provider": "phytozome",
                "species_key": species_key,
                "species_prefix": species_prefix,
                "cds_path": cds_path,
                "gff_path": gff_path,
                "gbff_path": gbff_path,
                "genome_path": genome_path,
            }
        )

    return tasks, warnings, errors


def discover_ncbi_like_tasks(input_dir, provider):
    warnings = []
    errors = []
    tasks = []

    for species_dir in sorted(input_dir.iterdir()):
        if not species_dir.is_dir():
            continue
        species_key = species_dir.name
        species_prefix = species_prefix_from_value(species_key)
        if species_prefix == "":
            warnings.append(
                "[{}] skipped '{}': unable to parse species prefix.".format(provider, species_key)
            )
            continue

        files = [path for path in species_dir.iterdir() if path.is_file()]
        cds_matches = [
            path
            for path in files
            if "cds_from_genomic" in path.name.lower() and is_fasta_filename(path.name)
        ]
        gff_matches = [
            path
            for path in files
            if "genomic.gff" in path.name.lower() and is_gff_filename(path.name)
        ]
        gbff_matches = [
            path
            for path in files
            if "genomic.gbff" in path.name.lower() and is_gbff_filename(path.name)
        ]
        genome_matches = [path for path in files if is_probable_genome_filename(provider, path.name)]

        cds_path = pick_single_file(cds_matches, provider, species_key, "CDS", warnings)
        gff_path = pick_single_file(gff_matches, provider, species_key, "GFF", warnings)
        gbff_path = pick_single_file(gbff_matches, provider, species_key, "GBFF", warnings)
        genome_path = pick_single_file(genome_matches, provider, species_key, "genome", warnings)

        missing_label = task_missing_annotation_label(cds_path, gff_path, gbff_path, genome_path)
        if missing_label != "":
            errors.append(
                "[{}] {}: missing {}".format(
                    provider,
                    species_key, missing_label
                )
            )
            continue
        tasks.append(
            {
                "provider": provider,
                "species_key": species_key,
                "species_prefix": species_prefix,
                "cds_path": cds_path,
                "gff_path": gff_path,
                "gbff_path": gbff_path,
                "genome_path": genome_path,
            }
        )

    return tasks, warnings, errors


def discover_tasks(provider, input_dir):
    if provider in ("ensembl", "ensemblplants"):
        return discover_ensembl_like_tasks(input_dir, provider)
    if provider == "phycocosm":
        return discover_phycocosm_tasks(input_dir)
    if provider == "phytozome":
        return discover_phytozome_tasks(input_dir)
    if provider in ("ncbi", "refseq", "genbank"):
        return discover_ncbi_like_tasks(input_dir, provider)
    if provider == "coge":
        return discover_generic_species_dir_tasks(provider, input_dir)
    if provider == "cngb":
        return discover_generic_species_dir_tasks(provider, input_dir)
    if provider in ("gwh", "flybase", "wormbase", "vectorbase", "fernbase", "veupathdb", "dictybase", "insectbase", "direct"):
        return discover_generic_species_dir_tasks(provider, input_dir)
    if provider == "local":
        return discover_generic_species_dir_tasks(provider, input_dir)
    raise ValueError("Unknown provider: {}".format(provider))


def build_formatted_cds_id(task, header):
    extracted = extract_provider_transcript_id(task["provider"], header)
    sanitized = sanitize_identifier(extracted)
    prefixed = "{}_{}".format(task["species_prefix"], sanitized)
    return sanitize_identifier(prefixed)


def format_cds(task, output_dir, overwrite, dry_run):
    cds_input = task.get("cds_path")
    if cds_input is not None:
        output_name = normalize_cds_output_basename(cds_input.name, task["species_prefix"])
    else:
        output_name = build_derived_cds_output_basename(task)
    output_path = output_dir / output_name

    if output_path.exists() and output_path.stat().st_size > 0 and not overwrite:
        before_count = 0
        for _header, _sequence in iter_task_cds_records(task):
            before_count += 1
        after_count, first_existing = count_fasta_records(output_path)
        return {
            "status": "skip",
            "output_path": output_path,
            "input_path": describe_task_cds_input(task),
            "written": after_count,
            "duplicates": max(0, before_count - after_count),
            "before_count": before_count,
            "after_count": after_count,
            "first_sequence_name": first_existing,
        }

    before_count = 0
    aggregated_away = 0
    first_sequence_name = ""
    records_by_gene = {}
    for header, sequence in iter_task_cds_records(task):
        before_count += 1
        transcript_id = build_formatted_cds_id(task, header)
        gene_id = build_gene_aggregate_id(task, header, transcript_id)
        seq = re.sub(r"\s+", "", sequence).upper()
        # Keep codon-frame-safe length (equivalent role of `cdskit pad` in shell pipelines).
        seq = pad_to_codon_length(seq)

        previous = records_by_gene.get(gene_id)
        if previous is None:
            records_by_gene[gene_id] = {
                "id": gene_id,
                "sequence": seq,
                "transcript_id": transcript_id,
            }
            continue

        previous_seq = previous["sequence"]
        previous_transcript_id = previous["transcript_id"]
        # Keep one representative CDS per gene; prefer longer CDS and then lexicographically stable tie-breaker.
        if len(seq) > len(previous_seq) or (
            len(seq) == len(previous_seq) and transcript_id < previous_transcript_id
        ):
            records_by_gene[gene_id] = {
                "id": gene_id,
                "sequence": seq,
                "transcript_id": transcript_id,
            }

    ordered_ids = sorted(records_by_gene.keys())
    after_count = len(ordered_ids)
    aggregated_away = max(0, before_count - after_count)
    if len(ordered_ids) > 0:
        first_sequence_name = ordered_ids[0]

    if not dry_run:
        write_fasta_records_gzip(
            output_path,
            (
                (records_by_gene[gene_id]["id"], records_by_gene[gene_id]["sequence"])
                for gene_id in ordered_ids
            ),
        )

    status = "dry-run" if dry_run else "write"
    return {
        "status": status,
        "output_path": output_path,
        "input_path": describe_task_cds_input(task),
        "written": after_count,
        "duplicates": aggregated_away,
        "before_count": before_count,
        "after_count": after_count,
        "first_sequence_name": first_sequence_name,
    }


def format_genome(task, output_dir, overwrite, dry_run):
    genome_path = task.get("genome_path")
    gbff_path = task.get("gbff_path")
    if genome_path is None and gbff_path is None:
        return {"status": "missing", "output_path": None, "written": 0}

    if genome_path is not None:
        output_name = normalize_genome_output_basename(genome_path.name, task["species_prefix"])
    else:
        output_name = build_derived_genome_output_basename(task)
    output_path = output_dir / output_name
    if output_path.exists() and output_path.stat().st_size > 0 and not overwrite:
        return {"status": "skip", "output_path": output_path, "written": 0}
    if dry_run:
        return {"status": "dry-run", "output_path": output_path, "written": 0}

    written = 0
    def iter_genome_output_records():
        nonlocal written
        if genome_path is not None:
            for header, sequence in iter_fasta_records(genome_path):
                record_id = first_token(apply_common_replacements(header))
                if record_id == "":
                    record_id = "unnamed"
                seq = re.sub(r"\s+", "", sequence).upper()
                written += 1
                yield record_id, seq
            return
        for record_id, sequence in iter_genome_records_from_gbff(gbff_path):
            written += 1
            yield record_id, sequence

    write_fasta_records_gzip(output_path, iter_genome_output_records())
    return {"status": "write", "output_path": output_path, "written": written}


def format_gff(task, output_dir, overwrite, dry_run):
    gff_path = task.get("gff_path")
    gbff_path = task.get("gbff_path")
    if gff_path is not None:
        output_name = normalize_gff_output_basename(gff_path.name, task["species_prefix"])
    elif gbff_path is not None:
        output_name = build_derived_gff_output_basename(task)
    else:
        return {"status": "missing", "output_path": None, "lines": 0}
    output_path = output_dir / output_name
    if output_path.exists() and output_path.stat().st_size > 0 and not overwrite:
        return {"status": "skip", "output_path": output_path, "lines": 0}
    if dry_run:
        return {"status": "dry-run", "output_path": output_path, "lines": 0}

    if gff_path is not None:
        line_count = write_gff_gzip(gff_path, output_path)
    else:
        line_count = write_gff_lines_gzip(output_path, iter_gff_lines_from_gbff(task))
    return {"status": "write", "output_path": output_path, "lines": line_count}


def resolve_provider_inputs(args):
    if args.provider == "all":
        if args.input_dir == "":
            raise ValueError("--input-dir is required when --provider all is used.")
        input_root = Path(args.input_dir).expanduser().resolve()
        return [
            (provider, (input_root / DEFAULT_INPUT_RELATIVE_DIRS[provider]).resolve())
            for provider in PROVIDERS
        ]

    provider = args.provider
    if args.input_dir != "":
        return [(provider, Path(args.input_dir).expanduser().resolve())]
    raise ValueError("Specify --input-dir.")


def utc_now_iso():
    return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())


def species_row_key(provider, species_key, species_prefix):
    stable_species = species_key if species_key != "" else species_prefix
    return "{}\t{}".format(provider, stable_species)


def read_species_summary_rows(path):
    rows = {}
    if not path.exists() or path.stat().st_size == 0:
        return rows
    with open(path, "rt", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for raw in reader:
            provider = (raw.get("provider") or "").strip()
            species_key = (raw.get("species_key") or "").strip()
            species_prefix = (raw.get("species_prefix") or "").strip()
            if provider == "" or (species_key == "" and species_prefix == ""):
                continue
            key = species_row_key(provider, species_key, species_prefix)
            rows[key] = {col: str(raw.get(col) or "") for col in SPECIES_SUMMARY_COLUMNS}
    return rows


def species_summary_row_has_existing_outputs(row):
    cds_output = (row.get("cds_output_path") or "").strip()
    gff_output = (row.get("gff_output_path") or "").strip()
    genome_output = (row.get("genome_output_path") or "").strip()
    gff_status = (row.get("gff_status") or "").strip()
    genome_status = (row.get("genome_status") or "").strip()

    if cds_output == "":
        return False
    if not Path(cds_output).expanduser().exists():
        return False
    if gff_status != "missing":
        if gff_output == "":
            return False
        if not Path(gff_output).expanduser().exists():
            return False
    if genome_output != "" and genome_status != "missing" and not Path(genome_output).expanduser().exists():
        return False
    return True


def retain_existing_species_summary_rows(rows):
    retained = {}
    for key in sorted(rows.keys()):
        row = rows[key]
        if species_summary_row_has_existing_outputs(row):
            retained[key] = row
    return retained


def write_species_summary_rows(path, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = Path(str(path) + ".tmp.{}".format(os.getpid()))
    try:
        with open(tmp_path, "wt", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=SPECIES_SUMMARY_COLUMNS, delimiter="\t")
            writer.writeheader()
            for key in sorted(rows.keys()):
                row = rows[key]
                payload = {col: str(row.get(col) or "") for col in SPECIES_SUMMARY_COLUMNS}
                writer.writerow(payload)
        tmp_path.replace(path)
    except Exception:
        try:
            tmp_path.unlink()
        except FileNotFoundError:
            pass
        except OSError:
            pass
        raise


def format_task_succeeded(cds_result, gff_result, genome_result, dry_run):
    if dry_run:
        return False
    if cds_result["status"] not in ("write", "skip"):
        return False
    if gff_result["status"] not in ("write", "skip", "missing"):
        return False
    if genome_result["status"] not in ("write", "skip", "missing"):
        return False
    return True


def build_species_summary_row(
    task,
    cds_result,
    gff_result,
    genome_result,
    run_started_utc,
    overwrite,
    dry_run,
    taxonomy_metadata=None,
):
    metadata = blank_species_taxonomy_metadata()
    if taxonomy_metadata is not None:
        metadata.update({key: str(value or "") for key, value in taxonomy_metadata.items()})
    return {
        "updated_utc": utc_now_iso(),
        "run_started_utc": run_started_utc,
        "provider": task["provider"],
        "species_key": task["species_key"],
        "species_prefix": task["species_prefix"],
        "taxid": metadata["taxid"],
        "nuclear_genetic_code_id": metadata["nuclear_genetic_code_id"],
        "nuclear_genetic_code_name": metadata["nuclear_genetic_code_name"],
        "mitochondrial_genetic_code_id": metadata["mitochondrial_genetic_code_id"],
        "mitochondrial_genetic_code_name": metadata["mitochondrial_genetic_code_name"],
        "plastid_genetic_code_id": metadata["plastid_genetic_code_id"],
        "plastid_genetic_code_name": metadata["plastid_genetic_code_name"],
        "cds_input_path": str(cds_result.get("input_path") or ""),
        "gff_input_path": str(describe_task_gff_input(task) or ""),
        "genome_input_path": str(describe_task_genome_input(task) or ""),
        "cds_output_path": str(cds_result["output_path"]) if cds_result.get("output_path") is not None else "",
        "gff_output_path": str(gff_result["output_path"]) if gff_result.get("output_path") is not None else "",
        "genome_output_path": str(genome_result["output_path"]) if genome_result.get("output_path") is not None else "",
        "cds_status": cds_result["status"],
        "gff_status": gff_result["status"],
        "genome_status": genome_result["status"],
        "cds_sequences_before": str(cds_result.get("before_count", "")),
        "cds_sequences_after": str(cds_result.get("after_count", "")),
        "cds_first_sequence_name": cds_result.get("first_sequence_name", "") or "NA",
        "aggregated_cds_removed": str(cds_result.get("duplicates", "")),
        "gene_grouping_mode": gene_grouping_mode_for_task(task),
        "overwrite": str(int(bool(overwrite))),
        "dry_run": str(int(bool(dry_run))),
    }


def main():
    parser = build_arg_parser()
    args = parser.parse_args()
    run_started_utc = utc_now_iso()

    if args.download_only and args.download_manifest == "":
        parser.error("--download-only requires --download-manifest.")
    if args.download_manifest != "" and args.provider in ("phycocosm", "phytozome"):
        parser.error(
            "--download-manifest does not support provider '{}'. Use --input-dir for local formatting.".format(
                args.provider
            )
        )

    try:
        http_headers = parse_http_headers(args.http_header, args.auth_bearer_token_env)
    except ValueError as exc:
        parser.error(str(exc))

    if args.download_manifest != "":
        parallel_jobs = resolve_parallel_jobs(args.jobs)
        download_root = Path(args.download_dir).expanduser().resolve()
        download_root.mkdir(parents=True, exist_ok=True)
        resolved_manifest_output_path = None
        if args.resolved_manifest_output != "":
            resolved_manifest_output_path = Path(args.resolved_manifest_output).expanduser().resolve()
        report = download_from_manifest(
            manifest_path=Path(args.download_manifest).expanduser().resolve(),
            download_root=download_root,
            provider_filter=args.provider,
            overwrite=args.overwrite,
            headers=http_headers,
            timeout=float(args.download_timeout),
            dry_run=args.dry_run,
            jobs=parallel_jobs,
            resolved_manifest_output_path=resolved_manifest_output_path,
        )
        for warning in report["warnings"]:
            sys.stderr.write("Warning: {}\n".format(warning))
        for error in report["errors"]:
            sys.stderr.write("Error: {}\n".format(error))
        print(
            "Download stage: rows processed={}, files downloaded={}, planned downloads={}, download root={}, resolved manifest={}, dry_run={}, jobs={}".format(
                report["processed"],
                report["downloaded"],
                report["planned"],
                download_root,
                report.get("resolved_manifest_output", ""),
                int(args.dry_run),
                parallel_jobs,
            )
        )
        if args.download_only:
            return 0 if len(report["errors"]) == 0 else 1
        if len(report["errors"]) > 0 and args.strict:
            return 1

    if args.input_dir == "" and args.download_manifest != "":
        download_input_root = Path(args.download_dir).expanduser().resolve()
        if args.provider == "all":
            args.input_dir = str(download_input_root)
        else:
            args.input_dir = str((download_input_root / DEFAULT_INPUT_RELATIVE_DIRS[args.provider]).resolve())

    try:
        provider_inputs = resolve_provider_inputs(args)
    except ValueError as exc:
        parser.error(str(exc))

    output_cds_dir = Path(args.species_cds_dir).expanduser().resolve()
    output_gff_dir = Path(args.species_gff_dir).expanduser().resolve()
    output_genome_dir = Path(args.species_genome_dir).expanduser().resolve()
    species_summary_path = Path(args.species_summary_output).expanduser().resolve()
    output_cds_dir.mkdir(parents=True, exist_ok=True)
    output_gff_dir.mkdir(parents=True, exist_ok=True)
    output_genome_dir.mkdir(parents=True, exist_ok=True)
    species_summary_rows = retain_existing_species_summary_rows(read_species_summary_rows(species_summary_path))
    write_species_summary_rows(species_summary_path, species_summary_rows)
    taxonomy_resolver = SpeciesTaxonomyMetadataResolver.from_environment()

    all_tasks = []
    all_warnings = []
    all_errors = []

    for provider, input_dir in provider_inputs:
        if not input_dir.exists() or not input_dir.is_dir():
            message = "[{}] input directory not found: {}".format(provider, input_dir)
            if args.provider == "all":
                all_warnings.append(message)
            else:
                all_errors.append(message)
            continue
        tasks, warnings, errors = discover_tasks(provider, input_dir)
        for task in tasks:
            task["gene_grouping_mode"] = args.gene_grouping_mode
        all_tasks.extend(tasks)
        all_warnings.extend(warnings)
        all_errors.extend(errors)

    for warning in all_warnings:
        sys.stderr.write("Warning: {}\n".format(warning))
    for error in all_errors:
        sys.stderr.write("Error: {}\n".format(error))

    if args.strict and len(all_errors) > 0:
        return 1
    if len(all_tasks) == 0:
        sys.stderr.write("No species inputs with CDS, GBFF, or GFF plus genome were discovered.\n")
        return 1

    processed = 0
    total_duplicates = 0
    total_cds_before = 0
    total_cds_after = 0
    first_cds_sequence_name = ""
    species_with_genome = 0
    for task in all_tasks:
        cds_result = format_cds(task, output_cds_dir, args.overwrite, args.dry_run)
        gff_result = format_gff(task, output_gff_dir, args.overwrite, args.dry_run)
        genome_result = format_genome(task, output_genome_dir, args.overwrite, args.dry_run)
        if format_task_succeeded(cds_result, gff_result, genome_result, args.dry_run):
            key = species_row_key(task["provider"], task["species_key"], task["species_prefix"])
            taxonomy_metadata = taxonomy_resolver.resolve(task["species_prefix"])
            species_summary_rows[key] = build_species_summary_row(
                task,
                cds_result,
                gff_result,
                genome_result,
                run_started_utc=run_started_utc,
                overwrite=args.overwrite,
                dry_run=args.dry_run,
                taxonomy_metadata=taxonomy_metadata,
            )
            # Persist each successful species incrementally.
            write_species_summary_rows(species_summary_path, species_summary_rows)
        processed += 1
        total_duplicates += cds_result["duplicates"]
        total_cds_before += cds_result["before_count"]
        total_cds_after += cds_result["after_count"]
        if first_cds_sequence_name == "":
            first_cds_sequence_name = cds_result["first_sequence_name"]
        if genome_result["status"] != "missing":
            species_with_genome += 1
        print(
            "[{}] {}: CDS={} ({}, {}, aggregated_away={}, before={}, after={}), GFF={} ({}, lines={}), GENOME={} ({})".format(
                task["provider"],
                task["species_prefix"],
                task["cds_path"].name if task.get("cds_path") is not None else Path(str(describe_task_cds_input(task) or "derived_cds")).name,
                cds_result["status"],
                cds_result["output_path"].name,
                cds_result["duplicates"],
                cds_result["before_count"],
                cds_result["after_count"],
                Path(str(describe_task_gff_input(task) or "derived_gff")).name,
                gff_result["status"],
                gff_result["lines"],
                Path(str(describe_task_genome_input(task) or "NA")).name,
                genome_result["status"],
            )
        )

    stats = {
        "species_processed": processed,
        "num_species_cds_files": processed,
        "num_species_gff_files": processed,
        "num_species_genome_files": species_with_genome,
        "cds_sequences_before": total_cds_before,
        "cds_sequences_after": total_cds_after,
        "cds_first_sequence_name": first_cds_sequence_name,
        "aggregated_cds_removed": total_duplicates,
        "duplicate_cds_ids_skipped": total_duplicates,
        "dry_run": int(args.dry_run),
    }
    if args.stats_output != "":
        stats_path = Path(args.stats_output).expanduser().resolve()
        stats_path.parent.mkdir(parents=True, exist_ok=True)
        with open(stats_path, "wt", encoding="utf-8") as handle:
            json.dump(stats, handle, ensure_ascii=True, indent=2, sort_keys=True)

    print(
        "Finished. species processed: {}, CDS aggregated away: {}, CDS before/after: {}/{}, first CDS sequence: {}, output CDS dir: {}, output GFF dir: {}, output genome dir: {}, species summary: {}, dry_run={}".format(
            processed,
            total_duplicates,
            total_cds_before,
            total_cds_after,
            first_cds_sequence_name if first_cds_sequence_name != "" else "NA",
            output_cds_dir,
            output_gff_dir,
            output_genome_dir,
            species_summary_path,
            int(args.dry_run),
        )
    )
    if len(all_errors) > 0:
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main())
