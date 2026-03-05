#!/usr/bin/env python3

import argparse
import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
import json
import math
from pathlib import Path
import re
import shutil
import sys
from typing import Dict, List, Optional, Sequence, Set
import zipfile
from urllib.parse import quote, urlparse
from urllib.request import Request, urlopen

import pandas

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - exercised when openpyxl is unavailable
    load_workbook = None


SUPPORTED_DATABASES = {
    "try": {
        "acquisition_mode": "bulk",
        "scope": "all_species_snapshot",
        "notes": "Use locally provided TRY export due to access restrictions.",
    },
    "gift": {
        "acquisition_mode": "gift_api",
        "scope": "trait_subset_api",
        "notes": "Resolve target work_IDs via species lookup, then fetch requested trait IDs.",
    },
    "bien": {
        "acquisition_mode": "species_api",
        "scope": "target_species_only",
        "notes": "Species-level retrieval via API/package layer is expected.",
    },
    "eol_traitbank": {
        "acquisition_mode": "species_api",
        "scope": "target_species_only",
        "notes": "Species-level structured API retrieval is expected.",
    },
    "austraits": {
        "acquisition_mode": "bulk",
        "scope": "all_species_snapshot",
        "notes": "Download full release snapshot then subset target species.",
    },
    "eltontraits": {
        "acquisition_mode": "bulk",
        "scope": "all_species_snapshot",
        "notes": "Download full release snapshot then subset target species.",
    },
    "combine": {
        "acquisition_mode": "bulk",
        "scope": "all_species_snapshot",
        "notes": "Download full release snapshot then subset target species.",
    },
    "birdbase": {
        "acquisition_mode": "bulk",
        "scope": "all_species_snapshot",
        "notes": "Download full release snapshot then subset target species.",
    },
    "amniote": {
        "acquisition_mode": "bulk",
        "scope": "all_species_snapshot",
        "notes": "Download full release snapshot then subset target species.",
    },
    "amphibio": {
        "acquisition_mode": "bulk",
        "scope": "all_species_snapshot",
        "notes": "Download full release snapshot then subset target species.",
    },
    "animaltraits": {
        "acquisition_mode": "bulk",
        "scope": "all_species_snapshot",
        "notes": "Download full release snapshot then subset target species.",
    },
    "fishbase": {
        "acquisition_mode": "bulk",
        "scope": "all_species_snapshot",
        "notes": "Cache bulk tables/snapshot then subset target species.",
    },
}

FASTA_EXTENSIONS = (
    ".fa",
    ".fas",
    ".fasta",
    ".fna",
    ".fa.gz",
    ".fas.gz",
    ".fasta.gz",
    ".fna.gz",
)
SPECIES_COLUMN_CANDIDATES = (
    "species",
    "species_name",
    "scientific_name",
    "scientificname",
    "binomial",
    "taxon_name",
    "taxon",
)
DEFAULT_MANIFEST_PATH = Path("workspace/input/input_generation/download_plan.xlsx")
DEFAULT_TRAIT_PLAN_PATH = Path("workspace/input/input_generation/trait_plan.tsv")
DEFAULT_DB_SOURCES_PATH = Path("workspace/input/input_generation/trait_database_sources.tsv")
DEFAULT_DOWNLOADS_DIR = Path("workspace/downloads/trait_datasets")
DEFAULT_OUTPUT_PATH = Path("workspace/input/species_trait/species_trait.tsv")
DEFAULT_GIFT_API = "https://gift.uni-goettingen.de/api/extended/"
DEFAULT_GIFT_PAGE_SIZE = 10000
GIFT_TRAIT_ID_PATTERN = re.compile(r"^\d+(?:\.\d+)+$")


@dataclass
class TraitPlanRow:
    database: str
    source_column: str
    output_trait: str
    value_type: str
    aggregation: str
    positive_values: Set[str]
    trait_key: str
    trait_key_column: str


def _log(message: str) -> None:
    print(message, flush=True)


def normalize_species_name(value: object) -> str:
    raw = str(value or "").strip()
    if raw == "":
        return ""
    raw = raw.replace("−", "-")
    raw = re.sub(r"\([^)]*\)", " ", raw)
    raw = raw.replace("-", "_").replace("/", "_")
    raw = re.sub(r"[^A-Za-z0-9_]+", "_", raw)
    tokens = [token for token in raw.split("_") if token]
    if len(tokens) < 2:
        return ""
    genus = tokens[0]
    species = tokens[1]
    if genus == "":
        return ""
    genus_norm = genus[:1].upper() + genus[1:].lower()
    species_norm = species.lower()
    return f"{genus_norm}_{species_norm}"


def parse_species_from_id_label(value: str) -> str:
    text = str(value or "").strip()
    if text == "":
        return ""
    match = re.search(r"\(([^()]+)\)", text)
    if match:
        normalized = normalize_species_name(match.group(1))
        if normalized != "":
            return normalized
    return normalize_species_name(text)


def parse_delimiter(path: Path, explicit: str) -> str:
    explicit_norm = str(explicit or "").strip().lower()
    if explicit_norm in ("tab", "\\t", "tsv"):
        return "\t"
    if explicit_norm in ("comma", ",", "csv"):
        return ","
    if explicit_norm == "pipe":
        return "|"
    suffixes = "".join(path.suffixes).lower()
    if suffixes.endswith(".tsv") or suffixes.endswith(".tsv.gz"):
        return "\t"
    if suffixes.endswith(".csv") or suffixes.endswith(".csv.gz"):
        return ","
    return "\t"


def strip_known_extension(path: Path) -> str:
    name = path.name
    lowered = name.lower()
    for ext in FASTA_EXTENSIONS:
        if lowered.endswith(ext):
            return name[: len(name) - len(ext)]
    return path.stem


def read_manifest_rows(path: Path) -> List[Dict[str, str]]:
    suffix = "".join(path.suffixes).lower()
    if suffix.endswith(".xlsx"):
        if load_workbook is None:
            raise RuntimeError("openpyxl is required to parse XLSX manifest files.")
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        row_iter = sheet.iter_rows(values_only=True)
        header_values = next(row_iter, None)
        if header_values is None:
            return []
        header = [str(cell or "").strip() for cell in header_values]
        rows: List[Dict[str, str]] = []
        for cells in row_iter:
            if cells is None:
                continue
            row = {header[i]: str(cells[i] or "").strip() for i in range(min(len(header), len(cells)))}
            if any(value.strip() != "" for value in row.values()):
                rows.append(row)
        return rows
    delimiter = parse_delimiter(path, "")
    with open(path, "rt", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        return [{key: str(value or "").strip() for key, value in row.items()} for row in reader]


def species_from_manifest(path: Path) -> Set[str]:
    rows = read_manifest_rows(path)
    species: Set[str] = set()
    for row in rows:
        provider = str(row.get("provider", "") or "").strip()
        if provider.startswith("#"):
            continue
        candidate = normalize_species_name(row.get("species_key", ""))
        if candidate == "":
            candidate = parse_species_from_id_label(row.get("id", ""))
        if candidate != "":
            species.add(candidate)
    return species


def species_from_cds_dir(path: Path) -> Set[str]:
    species: Set[str] = set()
    for file_path in sorted(path.iterdir()):
        if not file_path.is_file():
            continue
        if file_path.name.startswith("."):
            continue
        stem = strip_known_extension(file_path)
        candidate = normalize_species_name(stem)
        if candidate != "":
            species.add(candidate)
    return species


def read_trait_plan(path: Path) -> List[TraitPlanRow]:
    with open(path, "rt", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        rows = list(reader)
    required_columns = {"database", "source_column", "output_trait"}
    missing = [col for col in required_columns if col not in (reader.fieldnames or [])]
    if missing:
        raise ValueError("Trait plan is missing required columns: {}".format(", ".join(sorted(missing))))
    out: List[TraitPlanRow] = []
    for raw in rows:
        database = str(raw.get("database", "") or "").strip().lower()
        source_column = str(raw.get("source_column", "") or "").strip()
        output_trait = str(raw.get("output_trait", "") or "").strip()
        if database == "" or source_column == "" or output_trait == "":
            continue
        value_type = str(raw.get("value_type", "numeric") or "numeric").strip().lower()
        aggregation = str(raw.get("aggregation", "") or "").strip().lower()
        if aggregation == "":
            aggregation = "any" if value_type == "binary" else "median"
        positive_raw = str(raw.get("positive_values", "") or "").strip()
        positive_values = {
            token.strip().lower()
            for token in positive_raw.split(",")
            if token.strip() != ""
        }
        trait_key = str(raw.get("trait_key", "") or "").strip()
        trait_key_column = str(raw.get("trait_key_column", "") or "").strip()
        out.append(
            TraitPlanRow(
                database=database,
                source_column=source_column,
                output_trait=output_trait,
                value_type=value_type,
                aggregation=aggregation,
                positive_values=positive_values,
                trait_key=trait_key,
                trait_key_column=trait_key_column,
            )
        )
    return out


def read_database_sources(path: Path) -> Dict[str, Dict[str, str]]:
    with open(path, "rt", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        rows = list(reader)
    out: Dict[str, Dict[str, str]] = {}
    for raw in rows:
        database = str(raw.get("database", "") or "").strip().lower()
        if database == "":
            continue
        normalized = {str(k or "").strip(): str(v or "").strip() for k, v in raw.items()}
        uri = normalized.get("uri", "")
        if uri != "":
            parsed = urlparse(uri)
            if parsed.scheme == "" and not Path(uri).is_absolute():
                normalized["uri"] = str((path.parent / uri).resolve())
        out[database] = normalized
    return out


def resolve_requested_databases(
    databases_arg: str,
    plan_rows: Sequence[TraitPlanRow],
    source_rows: Dict[str, Dict[str, str]],
) -> List[str]:
    requested = str(databases_arg or "").strip().lower()
    if requested == "" or requested == "auto":
        return sorted({row.database for row in plan_rows})
    if requested == "all":
        return sorted(SUPPORTED_DATABASES.keys())
    dbs = [token.strip().lower() for token in requested.split(",") if token.strip() != ""]
    if len(dbs) == 0:
        return sorted({row.database for row in plan_rows})
    unknown = [db for db in dbs if db not in SUPPORTED_DATABASES and db not in source_rows]
    if unknown:
        raise ValueError("Unknown database(s): {}".format(", ".join(sorted(unknown))))
    return dbs


def detect_species_column(df: pandas.DataFrame, requested: str) -> str:
    requested_norm = str(requested or "").strip()
    if requested_norm != "" and requested_norm in df.columns:
        return requested_norm
    lowered = {col.lower(): col for col in df.columns}
    for candidate in SPECIES_COLUMN_CANDIDATES:
        if candidate in lowered:
            return lowered[candidate]
    return str(df.columns[0])


def read_table(path: Path, delimiter: str) -> pandas.DataFrame:
    suffix = "".join(path.suffixes).lower()
    if suffix.endswith(".xlsx"):
        return pandas.read_excel(path, dtype=str)
    sep = parse_delimiter(path, delimiter)
    return pandas.read_csv(path, sep=sep, dtype=str, encoding_errors="replace")


def read_table_from_text(text: str, delimiter: str) -> pandas.DataFrame:
    sep = parse_delimiter(Path("response.tsv"), delimiter)
    return pandas.read_csv(StringIO(text), sep=sep, dtype=str)


def split_uri_list(uri_raw: str) -> List[str]:
    return [token.strip() for token in str(uri_raw or "").split(",") if token.strip() != ""]


def choose_archive_member(members: Sequence[str], requested: str) -> Optional[str]:
    if len(members) == 0:
        return None
    requested_norm = str(requested or "").strip()
    if requested_norm != "":
        if requested_norm in members:
            return requested_norm
        return None
    preferred_suffixes = (".tsv", ".csv", ".txt", ".xlsx")
    for member in members:
        lowered = member.lower()
        for suffix in preferred_suffixes:
            if lowered.endswith(suffix):
                return member
    return members[0]


def read_table_from_zip(path: Path, delimiter: str, archive_member: str) -> pandas.DataFrame:
    with zipfile.ZipFile(path) as archive:
        members = [name for name in archive.namelist() if not name.endswith("/")]
        selected_member = choose_archive_member(members=members, requested=archive_member)
        if selected_member is None:
            raise ValueError("archive_member not found in {}: {}".format(path, archive_member))
        payload = archive.read(selected_member)
    if selected_member.lower().endswith(".xlsx"):
        return pandas.read_excel(BytesIO(payload), dtype=str)
    text: Optional[str] = None
    for encoding in ("utf-8", "latin1"):
        try:
            text = payload.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = payload.decode("utf-8", errors="replace")
    sep = parse_delimiter(Path(selected_member), delimiter)
    return pandas.read_csv(StringIO(text), sep=sep, dtype=str)


def read_bulk_table_from_path(path: Path, config: Dict[str, str]) -> pandas.DataFrame:
    delimiter = config.get("delimiter", "")
    archive_member = str(config.get("archive_member", "") or "").strip()
    suffix = "".join(path.suffixes).lower()
    if suffix.endswith(".zip"):
        return read_table_from_zip(path=path, delimiter=delimiter, archive_member=archive_member)
    return read_table(path=path, delimiter=delimiter)


def copy_or_download_file(
    uri: str,
    destination: Path,
    timeout: float,
    dry_run: bool,
) -> Optional[Path]:
    if uri == "":
        return None
    parsed = urlparse(uri)
    if parsed.scheme in ("",):
        src = Path(uri).expanduser().resolve()
        if not src.exists():
            return None
        if src == destination:
            return src
        if dry_run:
            _log("[dry-run] copy {} -> {}".format(src, destination))
            return destination
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(src, destination)
        return destination
    if parsed.scheme == "file":
        src = Path(parsed.path).resolve()
        if not src.exists():
            return None
        if src == destination:
            return src
        if dry_run:
            _log("[dry-run] copy {} -> {}".format(src, destination))
            return destination
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(src, destination)
        return destination
    if parsed.scheme in ("http", "https", "ftp"):
        if dry_run:
            _log("[dry-run] download {} -> {}".format(uri, destination))
            return destination
        destination.parent.mkdir(parents=True, exist_ok=True)
        request = Request(uri, headers={"User-Agent": "genegalleon-trait-generator"})
        with urlopen(request, timeout=timeout) as response, open(destination, "wb") as handle:
            shutil.copyfileobj(response, handle)
        return destination
    raise ValueError("Unsupported URI scheme in: {}".format(uri))


def detect_existing_bulk_file(database: str, downloads_dir: Path) -> Optional[Path]:
    db_dir = downloads_dir / database
    candidates = [
        db_dir / "{}.tsv".format(database),
        db_dir / "{}.csv".format(database),
        db_dir / "{}.txt".format(database),
        db_dir / "{}.zip".format(database),
        db_dir / "{}.tsv.gz".format(database),
        db_dir / "{}.csv.gz".format(database),
        db_dir / "source.tsv",
        db_dir / "source.csv",
        db_dir / "source.txt",
        db_dir / "source.zip",
        db_dir / "source.tsv.gz",
        db_dir / "source.csv.gz",
        db_dir / "source.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def load_bulk_database(
    database: str,
    config: Dict[str, str],
    downloads_dir: Path,
    timeout: float,
    dry_run: bool,
) -> Optional[pandas.DataFrame]:
    uri_values = split_uri_list(config.get("uri", ""))
    db_dir = downloads_dir / database
    table_paths: List[Path] = []

    if len(uri_values) > 0:
        for index, uri in enumerate(uri_values):
            parsed = urlparse(uri)
            basename = Path(parsed.path).name or "{}.{}.tsv".format(database, index + 1)
            target = db_dir / basename
            source_path = copy_or_download_file(uri=uri, destination=target, timeout=timeout, dry_run=dry_run)
            if source_path is None:
                continue
            if not dry_run:
                table_paths.append(source_path)
        if dry_run:
            return None
    else:
        existing = detect_existing_bulk_file(database=database, downloads_dir=downloads_dir)
        if existing is None:
            return None
        table_paths.append(existing)

    if len(table_paths) == 0:
        return None
    frames: List[pandas.DataFrame] = []
    for table_path in table_paths:
        frames.append(read_bulk_table_from_path(path=table_path, config=config))
    if len(frames) == 1:
        return frames[0]
    return pandas.concat(frames, ignore_index=True)


def fetch_species_api_table(
    database: str,
    config: Dict[str, str],
    species: Sequence[str],
    timeout: float,
    dry_run: bool,
) -> Optional[pandas.DataFrame]:
    uri_template = str(config.get("uri", "") or "").strip()
    if uri_template == "":
        return None
    response_format = str(config.get("response_format", "tsv") or "tsv").strip().lower()
    delimiter = config.get("delimiter", "\t")
    frames: List[pandas.DataFrame] = []
    for species_name in species:
        safe_species = quote(species_name.replace("_", " "))
        url = (
            uri_template
            .replace("{species}", species_name)
            .replace("{species_space}", species_name.replace("_", " "))
            .replace("{species_urlencoded}", safe_species)
        )
        if dry_run:
            _log("[dry-run] {} request: {}".format(database, url))
            continue
        request = Request(url, headers={"User-Agent": "genegalleon-trait-generator"})
        with urlopen(request, timeout=timeout) as response:
            payload = response.read().decode("utf-8", errors="replace")
        if response_format in ("tsv", "csv"):
            sep = "\t" if response_format == "tsv" else ","
            if str(delimiter or "").strip() != "":
                sep = parse_delimiter(Path("dummy.{}".format(response_format)), delimiter)
            frames.append(read_table_from_text(payload, sep))
        else:
            raise ValueError(
                "Unsupported species_api response_format '{}' for '{}'.".format(response_format, database)
            )
    if len(frames) == 0:
        return None
    return pandas.concat(frames, ignore_index=True)


def parse_bool_option(value: object, key_name: str, default: bool) -> bool:
    text = str(value or "").strip().lower()
    if text == "":
        return default
    if text in ("1", "true", "yes", "y", "on"):
        return True
    if text in ("0", "false", "no", "n", "off"):
        return False
    raise ValueError("Invalid boolean value for {}: {}".format(key_name, value))


def parse_positive_int_option(value: object, key_name: str, default: int) -> int:
    text = str(value or "").strip()
    if text == "":
        return default
    parsed = int(text)
    if parsed <= 0:
        raise ValueError("{} must be a positive integer: {}".format(key_name, value))
    return parsed


def normalize_base_uri(uri: str, fallback: str) -> str:
    base = str(uri or "").strip()
    if base == "":
        base = fallback
    if not base.endswith("/"):
        base = base + "/"
    return base


def fetch_json_payload(url: str, timeout: float) -> object:
    request = Request(
        url,
        headers={
            "User-Agent": "genegalleon-trait-generator",
            "Accept": "application/json",
        },
    )
    with urlopen(request, timeout=timeout) as response:
        payload = response.read().decode("utf-8", errors="replace")
    try:
        return json.loads(payload)
    except json.JSONDecodeError as exc:
        raise ValueError("Failed to parse JSON response from {}: {}".format(url, exc))


def json_payload_to_rows(payload: object) -> List[Dict[str, object]]:
    if payload is None:
        return []
    if isinstance(payload, list):
        return [row for row in payload if isinstance(row, dict)]
    if isinstance(payload, dict):
        for candidate_key in ("data", "results", "items"):
            value = payload.get(candidate_key)
            if isinstance(value, list):
                return [row for row in value if isinstance(row, dict)]
        return [payload]
    return []


def resolve_gift_version(
    data_api_base: str,
    requested_version: str,
    timeout: float,
    versions_api: str,
) -> str:
    version = str(requested_version or "").strip()
    if version == "":
        version = "latest"
    if version != "latest":
        return version
    if str(versions_api or "").strip() != "":
        versions_base = normalize_base_uri(versions_api, DEFAULT_GIFT_API)
    elif "/extended/" in data_api_base:
        versions_base = normalize_base_uri(data_api_base.replace("/extended/", "/", 1), DEFAULT_GIFT_API)
    else:
        versions_base = data_api_base
    versions_url = "{}index.php?query=versions".format(versions_base)
    rows = json_payload_to_rows(fetch_json_payload(url=versions_url, timeout=timeout))
    if len(rows) == 0:
        raise ValueError("GIFT versions endpoint returned no rows: {}".format(versions_url))
    resolved = str(rows[-1].get("version", "") or "").strip()
    if resolved == "":
        raise ValueError("GIFT versions endpoint did not include 'version': {}".format(versions_url))
    return resolved


def build_gift_index_url(data_api_base: str, version: str) -> str:
    if version == "beta":
        index_name = "index.php"
    else:
        index_name = "index{}.php".format(version)
    return "{}{}".format(data_api_base, index_name)


def split_genus_epithet(species_name: str) -> Optional[Sequence[str]]:
    tokens = [token for token in str(species_name or "").strip().split("_") if token]
    if len(tokens) < 2:
        return None
    return (tokens[0], tokens[1])


def score_gift_species_match(row: Dict[str, object], target_species: str) -> Sequence[float]:
    normalized_work_species = normalize_species_name(row.get("work_species", ""))
    exact_match = 1.0 if normalized_work_species == target_species else 0.0
    accepted = 1.0 if str(row.get("accepted", "")).strip() == "1" else 0.0
    resolved = 1.0 if str(row.get("resolved", "")).strip() == "1" else 0.0
    matched = 1.0 if str(row.get("matched", "")).strip() == "1" else 0.0
    overall_score_raw = str(row.get("overallscore", "")).strip()
    try:
        overall_score = float(overall_score_raw)
    except Exception:
        overall_score = 0.0
    return (exact_match, accepted, resolved, matched, overall_score)


def pick_best_gift_species_match(
    rows: Sequence[Dict[str, object]],
    target_species: str,
) -> Optional[Dict[str, object]]:
    if len(rows) == 0:
        return None
    ranked_rows = sorted(
        rows,
        key=lambda row: score_gift_species_match(row=row, target_species=target_species),
        reverse=True,
    )
    return ranked_rows[0]


def resolve_gift_species_map(
    index_url: str,
    species: Sequence[str],
    timeout: float,
) -> pandas.DataFrame:
    records: List[Dict[str, str]] = []
    work_id_to_species: Dict[str, str] = {}
    for species_name in species:
        genus_epithet = split_genus_epithet(species_name)
        if genus_epithet is None:
            continue
        genus, epithet = genus_epithet[0], genus_epithet[1]
        lookup_url = "{}?query=names_matched_unique&genus={}&epithet={}".format(
            index_url,
            quote(genus),
            quote(epithet),
        )
        rows = json_payload_to_rows(fetch_json_payload(url=lookup_url, timeout=timeout))
        best = pick_best_gift_species_match(rows=rows, target_species=species_name)
        if best is None:
            continue
        work_id = str(best.get("work_ID", "") or "").strip()
        if work_id == "":
            continue
        if work_id in work_id_to_species and work_id_to_species[work_id] != species_name:
            _log(
                "WARNING: [gift] work_ID {} matched multiple target species ('{}', '{}'); keeping first.".format(
                    work_id,
                    work_id_to_species[work_id],
                    species_name,
                )
            )
            continue
        work_id_to_species[work_id] = species_name
        work_species = str(best.get("work_species", "") or "").strip()
        if work_species == "":
            work_species = species_name.replace("_", " ")
        records.append(
            {
                "species": species_name,
                "work_ID": work_id,
                "work_species": work_species,
            }
        )
    return pandas.DataFrame(records)


def dedupe_keep_order(values: Sequence[str]) -> List[str]:
    deduped: List[str] = []
    seen: Set[str] = set()
    for value in values:
        key = str(value or "").strip()
        if key == "" or key in seen:
            continue
        seen.add(key)
        deduped.append(key)
    return deduped


def collect_gift_trait_tokens(
    database: str,
    config: Dict[str, str],
    plan_rows: Sequence[TraitPlanRow],
) -> List[str]:
    out: List[str] = []
    config_trait_ids = str(config.get("gift_trait_ids", "") or config.get("trait_ids", "") or "").strip()
    if config_trait_ids != "":
        out.extend([token.strip() for token in config_trait_ids.split(",") if token.strip() != ""])
    for row in plan_rows:
        if row.database != database:
            continue
        trait_id = str(row.trait_key or "").strip()
        if trait_id != "":
            out.append(trait_id)
    return dedupe_keep_order(out)


def is_gift_trait_id(value: str) -> bool:
    return bool(GIFT_TRAIT_ID_PATTERN.fullmatch(str(value or "").strip()))


def parse_float_or_default(value: object, default: float) -> float:
    try:
        return float(str(value or "").strip())
    except Exception:
        return default


def fetch_gift_traits_meta_rows(index_url: str, timeout: float) -> List[Dict[str, object]]:
    traits_meta_url = "{}?query=traits_meta".format(index_url)
    rows = json_payload_to_rows(fetch_json_payload(url=traits_meta_url, timeout=timeout))
    return [row for row in rows if isinstance(row, dict)]


def pick_best_gift_trait_match(rows: Sequence[Dict[str, object]]) -> Optional[Dict[str, object]]:
    if len(rows) == 0:
        return None
    ranked = sorted(
        rows,
        key=lambda row: (
            parse_float_or_default(row.get("count", 0), 0.0),
            str(row.get("Lvl3", "") or ""),
        ),
        reverse=True,
    )
    return ranked[0]


def resolve_gift_trait_token_map(
    index_url: str,
    trait_tokens: Sequence[str],
    timeout: float,
) -> Dict[str, str]:
    tokens = dedupe_keep_order([str(token or "").strip() for token in trait_tokens])
    if len(tokens) == 0:
        return {}
    resolved_map: Dict[str, str] = {}
    direct_ids = [token for token in tokens if is_gift_trait_id(token)]
    name_tokens = [token for token in tokens if not is_gift_trait_id(token)]
    for token in direct_ids:
        resolved_map[token] = token
    if len(name_tokens) == 0:
        return resolved_map

    traits_meta_rows = fetch_gift_traits_meta_rows(index_url=index_url, timeout=timeout)
    if len(traits_meta_rows) == 0:
        _log("WARNING: [gift] traits_meta returned no rows; trait name resolution skipped.")
        return resolved_map
    for token in name_tokens:
        token_norm = token.strip().lower()
        exact = [
            row
            for row in traits_meta_rows
            if str(row.get("Trait2", "") or "").strip().lower() == token_norm
            or str(row.get("Trait1", "") or "").strip().lower() == token_norm
        ]
        candidates = exact
        if len(candidates) == 0:
            candidates = [
                row
                for row in traits_meta_rows
                if token_norm in str(row.get("Trait2", "") or "").strip().lower()
                or token_norm in str(row.get("Trait1", "") or "").strip().lower()
            ]
        if len(candidates) == 0:
            _log("WARNING: [gift] trait token '{}' was not found in traits_meta.".format(token))
            continue
        selected = pick_best_gift_trait_match(candidates)
        if selected is None:
            continue
        selected_id = str(selected.get("Lvl3", "") or "").strip()
        if selected_id == "":
            _log("WARNING: [gift] trait token '{}' matched row without Lvl3.".format(token))
            continue
        if len(candidates) > 1:
            _log(
                "WARNING: [gift] trait token '{}' matched {} traits; selected '{}' ({}) by highest count.".format(
                    token,
                    len(candidates),
                    selected_id,
                    str(selected.get("Trait2", "") or "").strip(),
                )
            )
        resolved_map[token] = selected_id
    return resolved_map


def print_gift_traits(
    data_api_base: str,
    requested_version: str,
    versions_api: str,
    timeout: float,
    search: str,
    limit: int,
) -> None:
    resolved_version = resolve_gift_version(
        data_api_base=data_api_base,
        requested_version=requested_version,
        timeout=timeout,
        versions_api=versions_api,
    )
    index_url = build_gift_index_url(data_api_base=data_api_base, version=resolved_version)
    rows = fetch_gift_traits_meta_rows(index_url=index_url, timeout=timeout)
    if len(rows) == 0:
        print("trait_id\ttrait_name\ttrait_group\tvalue_type\tunits\tcount")
        return

    search_norm = str(search or "").strip().lower()
    filtered: List[Dict[str, object]] = []
    for row in rows:
        row_text = "\t".join(
            [
                str(row.get("Lvl3", "") or ""),
                str(row.get("Trait2", "") or ""),
                str(row.get("Trait1", "") or ""),
                str(row.get("Category", "") or ""),
            ]
        ).lower()
        if search_norm != "" and search_norm not in row_text:
            continue
        filtered.append(row)
    sorted_rows = sorted(
        filtered,
        key=lambda row: (
            parse_float_or_default(row.get("count", 0), 0.0),
            str(row.get("Lvl3", "") or ""),
        ),
        reverse=True,
    )
    if limit > 0:
        sorted_rows = sorted_rows[:limit]

    print("trait_id\ttrait_name\ttrait_group\tvalue_type\tunits\tcount")
    for row in sorted_rows:
        print(
            "{}\t{}\t{}\t{}\t{}\t{}".format(
                str(row.get("Lvl3", "") or "").strip(),
                str(row.get("Trait2", "") or "").strip(),
                str(row.get("Trait1", "") or "").strip(),
                str(row.get("type", "") or "").strip(),
                str(row.get("Units", "") or "").strip(),
                str(row.get("count", "") or "").strip(),
            )
        )


def fetch_gift_api_table(
    database: str,
    config: Dict[str, str],
    plan_rows: Sequence[TraitPlanRow],
    species: Sequence[str],
    timeout: float,
    dry_run: bool,
) -> Optional[pandas.DataFrame]:
    data_api_base = normalize_base_uri(config.get("uri", ""), DEFAULT_GIFT_API)
    requested_version = str(config.get("gift_version", "latest") or "latest").strip().lower()
    trait_tokens = collect_gift_trait_tokens(database=database, config=config, plan_rows=plan_rows)
    if len(trait_tokens) == 0:
        _log(
            "WARNING: [gift] no trait IDs resolved. Set trait_key in trait_plan or gift_trait_ids in database_sources."
        )
        return None
    page_size = parse_positive_int_option(
        value=config.get("gift_page_size", ""),
        key_name="gift_page_size",
        default=DEFAULT_GIFT_PAGE_SIZE,
    )
    max_pages = 0
    max_pages_raw = str(config.get("gift_max_pages_per_trait", "") or "").strip()
    if max_pages_raw != "":
        max_pages = parse_positive_int_option(
            value=max_pages_raw,
            key_name="gift_max_pages_per_trait",
            default=0,
        )
    bias_ref = parse_bool_option(config.get("gift_bias_ref", ""), key_name="gift_bias_ref", default=True)
    bias_deriv = parse_bool_option(config.get("gift_bias_deriv", ""), key_name="gift_bias_deriv", default=True)
    agreement_min: Optional[float] = None
    agreement_min_raw = str(config.get("gift_agreement_min", "") or "").strip()
    if agreement_min_raw != "":
        agreement_min = float(agreement_min_raw)

    if dry_run:
        display_version = requested_version if requested_version != "latest" else "<latest>"
        display_index_url = build_gift_index_url(data_api_base=data_api_base, version=display_version)
        for species_name in species:
            genus_epithet = split_genus_epithet(species_name)
            if genus_epithet is None:
                continue
            genus, epithet = genus_epithet[0], genus_epithet[1]
            _log(
                "[dry-run] {} request: {}?query=names_matched_unique&genus={}&epithet={}".format(
                    database,
                    display_index_url,
                    quote(genus),
                    quote(epithet),
                )
            )
        for trait_token in trait_tokens:
            _log(
                "[dry-run] {} request: {}?query=traits&traitid={}&biasref={}&biasderiv={}&startat=0".format(
                    database,
                    display_index_url,
                    quote(trait_token),
                    int(bias_ref),
                    int(bias_deriv),
                )
            )
        return None

    resolved_version = resolve_gift_version(
        data_api_base=data_api_base,
        requested_version=requested_version,
        timeout=timeout,
        versions_api=str(config.get("gift_versions_api", "") or "").strip(),
    )
    index_url = build_gift_index_url(data_api_base=data_api_base, version=resolved_version)
    trait_token_map = resolve_gift_trait_token_map(
        index_url=index_url,
        trait_tokens=trait_tokens,
        timeout=timeout,
    )
    trait_ids = dedupe_keep_order(list(trait_token_map.values()))
    if len(trait_ids) == 0:
        _log("WARNING: [gift] no trait IDs resolved after applying traits_meta lookup.")
        return None
    species_map = resolve_gift_species_map(index_url=index_url, species=species, timeout=timeout)
    if species_map.shape[0] == 0:
        return pandas.DataFrame(columns=["species", "work_ID", "work_species", "trait_ID", "trait_value"])
    work_ids = set(species_map["work_ID"].astype(str).tolist())

    frames: List[pandas.DataFrame] = []
    for trait_token in trait_tokens:
        trait_id = trait_token_map.get(trait_token, "")
        if trait_id == "":
            continue
        page_index = 0
        while True:
            start_at = page_index * page_size
            traits_url = "{}?query=traits&traitid={}&biasref={}&biasderiv={}&startat={}".format(
                index_url,
                quote(trait_id),
                int(bias_ref),
                int(bias_deriv),
                start_at,
            )
            rows = json_payload_to_rows(fetch_json_payload(url=traits_url, timeout=timeout))
            if len(rows) == 0:
                break
            page_df = pandas.DataFrame(rows)
            if "work_ID" not in page_df.columns:
                break
            page_df = page_df.copy()
            page_df["work_ID"] = page_df["work_ID"].astype(str).str.strip()
            matched_df = page_df.loc[page_df["work_ID"].isin(work_ids), :].copy()
            if matched_df.shape[0] > 0:
                matched_df["trait_ID"] = trait_id
                matched_df["trait_token"] = trait_token
                frames.append(matched_df)
            page_index += 1
            if max_pages > 0 and page_index >= max_pages:
                _log(
                    "WARNING: [gift] reached gift_max_pages_per_trait={} for trait_ID '{}'.".format(
                        max_pages,
                        trait_id,
                    )
                )
                break
            if len(rows) < page_size:
                break

    if len(frames) == 0:
        return pandas.DataFrame(columns=["species", "work_ID", "work_species", "trait_ID", "trait_value"])

    merged = pandas.concat(frames, ignore_index=True)
    species_map = species_map.copy()
    species_map["work_ID"] = species_map["work_ID"].astype(str)
    species_map = species_map.loc[:, ["work_ID", "species", "work_species"]].drop_duplicates(subset=["work_ID"])
    merged = merged.merge(species_map, how="left", on="work_ID")

    if agreement_min is not None and "agreement" in merged.columns:
        agreement_numeric = pandas.to_numeric(merged["agreement"], errors="coerce")
        merged = merged.loc[(agreement_numeric >= agreement_min) | agreement_numeric.isna(), :]
    return merged


def load_database_table(
    database: str,
    config: Dict[str, str],
    plan_rows: Sequence[TraitPlanRow],
    species: Sequence[str],
    downloads_dir: Path,
    timeout: float,
    dry_run: bool,
) -> Optional[pandas.DataFrame]:
    default_mode = SUPPORTED_DATABASES.get(database, {}).get("acquisition_mode", "bulk")
    acquisition_mode = str(config.get("acquisition_mode", default_mode) or default_mode).strip().lower()
    if acquisition_mode == "bulk":
        return load_bulk_database(
            database=database,
            config=config,
            downloads_dir=downloads_dir,
            timeout=timeout,
            dry_run=dry_run,
        )
    if acquisition_mode == "species_api":
        return fetch_species_api_table(
            database=database,
            config=config,
            species=species,
            timeout=timeout,
            dry_run=dry_run,
        )
    if acquisition_mode == "gift_api":
        return fetch_gift_api_table(
            database=database,
            config=config,
            plan_rows=plan_rows,
            species=species,
            timeout=timeout,
            dry_run=dry_run,
        )
    raise ValueError("Unsupported acquisition_mode '{}' for '{}'".format(acquisition_mode, database))


def as_numeric_or_nan(series: pandas.Series) -> pandas.Series:
    return pandas.to_numeric(series, errors="coerce")


def aggregate_numeric(values: pandas.Series, aggregation: str) -> object:
    numeric = as_numeric_or_nan(values).dropna()
    if numeric.shape[0] == 0:
        return pandas.NA
    if aggregation == "mean":
        return float(numeric.mean())
    if aggregation == "min":
        return float(numeric.min())
    if aggregation == "max":
        return float(numeric.max())
    return float(numeric.median())


def aggregate_binary(values: pandas.Series, aggregation: str, positive_values: Set[str]) -> object:
    if positive_values:
        lowered = values.astype(str).str.strip().str.lower()
        mapped = lowered.isin(positive_values).astype(int)
    else:
        numeric = as_numeric_or_nan(values).fillna(0)
        mapped = (numeric > 0).astype(int)
    if mapped.shape[0] == 0:
        return pandas.NA
    if aggregation in ("all", "min"):
        return int(mapped.min())
    if aggregation in ("sum",):
        return int(mapped.sum())
    if aggregation in ("mean",):
        return float(mapped.mean())
    return int(mapped.max())


def aggregate_categorical(values: pandas.Series, aggregation: str) -> object:
    valid = values.astype(str).str.strip()
    valid = valid[valid != ""]
    if valid.shape[0] == 0:
        return pandas.NA
    if aggregation == "first":
        return str(valid.iloc[0])
    mode_values = valid.mode(dropna=True)
    if mode_values.shape[0] == 0:
        return str(valid.iloc[0])
    return str(mode_values.iloc[0])


def format_output_value(value: object) -> str:
    if value is pandas.NA:
        return ""
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return "{:.6g}".format(value)
    if isinstance(value, (int,)):
        return str(value)
    text = str(value).strip()
    return text


def aggregate_trait_column(
    db_df: pandas.DataFrame,
    plan_row: TraitPlanRow,
) -> pandas.Series:
    grouped = db_df.groupby("__species_norm", observed=True)[plan_row.source_column]
    if plan_row.value_type == "binary":
        aggregated = grouped.apply(
            lambda s: aggregate_binary(
                values=s,
                aggregation=plan_row.aggregation,
                positive_values=plan_row.positive_values,
            )
        )
    elif plan_row.value_type == "categorical":
        aggregated = grouped.apply(
            lambda s: aggregate_categorical(
                values=s,
                aggregation=plan_row.aggregation,
            )
        )
    else:
        aggregated = grouped.apply(
            lambda s: aggregate_numeric(
                values=s,
                aggregation=plan_row.aggregation,
            )
        )
    return aggregated


def validate_species_source(
    species_source: str,
    manifest_path: Path,
    species_cds_dir: Path,
) -> Set[str]:
    if species_source == "download_manifest":
        if not manifest_path.exists():
            raise FileNotFoundError("Download manifest not found: {}".format(manifest_path))
        species = species_from_manifest(manifest_path)
        if len(species) == 0:
            raise ValueError("No species could be parsed from manifest: {}".format(manifest_path))
        return species
    if species_source == "species_cds":
        if not species_cds_dir.exists():
            raise FileNotFoundError("species_cds directory not found: {}".format(species_cds_dir))
        species = species_from_cds_dir(species_cds_dir)
        if len(species) == 0:
            raise ValueError("No species could be parsed from species_cds directory: {}".format(species_cds_dir))
        return species
    raise ValueError("Unsupported species-source: {}".format(species_source))


def print_supported_databases() -> None:
    print("database\tacquisition_mode\tretrieval_scope\tnotes")
    for database in sorted(SUPPORTED_DATABASES.keys()):
        info = SUPPORTED_DATABASES[database]
        print(
            "{}\t{}\t{}\t{}".format(
                database,
                info.get("acquisition_mode", ""),
                info.get("scope", ""),
                info.get("notes", ""),
            )
        )


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Generate workspace/input/species_trait/species_trait.tsv from selected species and trait sources. "
            "Species can be resolved from download_plan manifest (default) or species_cds filenames."
        )
    )
    parser.add_argument(
        "--species-source",
        choices=("download_manifest", "species_cds"),
        default="download_manifest",
        help="Source of target species set.",
    )
    parser.add_argument(
        "--download-manifest",
        default=str(DEFAULT_MANIFEST_PATH),
        help="Path to download_plan manifest (XLSX/TSV/CSV). Used when --species-source=download_manifest.",
    )
    parser.add_argument(
        "--species-cds-dir",
        default="workspace/output/input_generation/species_cds",
        help="species_cds directory used when --species-source=species_cds.",
    )
    parser.add_argument(
        "--trait-plan",
        default=str(DEFAULT_TRAIT_PLAN_PATH),
        help="Trait extraction plan TSV path.",
    )
    parser.add_argument(
        "--database-sources",
        default=str(DEFAULT_DB_SOURCES_PATH),
        help="Trait database source map TSV path.",
    )
    parser.add_argument(
        "--databases",
        default="auto",
        help="Comma-separated database IDs, or 'auto' (from trait_plan), or 'all'.",
    )
    parser.add_argument(
        "--downloads-dir",
        default=str(DEFAULT_DOWNLOADS_DIR),
        help="Directory for cached/downloaded trait database files.",
    )
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT_PATH),
        help="Output species trait TSV path.",
    )
    parser.add_argument(
        "--download-timeout",
        type=float,
        default=120.0,
        help="Network timeout in seconds for database retrieval.",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Fail when required inputs are missing or no trait columns are produced.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Resolve and report actions without writing files.",
    )
    parser.add_argument(
        "--stats-output",
        default="",
        help="Optional JSON stats output path.",
    )
    parser.add_argument(
        "--print-supported-databases",
        action="store_true",
        help="Print supported database IDs with retrieval mode and exit.",
    )
    parser.add_argument(
        "--print-gift-traits",
        action="store_true",
        help="Print GIFT traits_meta table and exit.",
    )
    parser.add_argument(
        "--gift-api",
        default=DEFAULT_GIFT_API,
        help="Base URI for GIFT API (used with --print-gift-traits).",
    )
    parser.add_argument(
        "--gift-version",
        default="latest",
        help="GIFT version (latest|beta|stable version; used with --print-gift-traits).",
    )
    parser.add_argument(
        "--gift-versions-api",
        default="",
        help="Optional base URI for GIFT versions endpoint (used with --print-gift-traits).",
    )
    parser.add_argument(
        "--gift-trait-search",
        default="",
        help="Case-insensitive keyword filter for --print-gift-traits.",
    )
    parser.add_argument(
        "--gift-trait-limit",
        type=int,
        default=0,
        help="Max rows for --print-gift-traits (0 means no limit).",
    )
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)

    if args.print_supported_databases:
        print_supported_databases()
        return 0
    if args.print_gift_traits:
        if int(args.gift_trait_limit) < 0:
            parser.error("--gift-trait-limit must be >= 0")
        gift_api_base = normalize_base_uri(args.gift_api, DEFAULT_GIFT_API)
        gift_versions_api = str(args.gift_versions_api or "").strip()
        if gift_versions_api != "":
            gift_versions_api = normalize_base_uri(gift_versions_api, DEFAULT_GIFT_API)
        print_gift_traits(
            data_api_base=gift_api_base,
            requested_version=str(args.gift_version or "latest").strip().lower(),
            versions_api=gift_versions_api,
            timeout=float(args.download_timeout),
            search=str(args.gift_trait_search or "").strip(),
            limit=int(args.gift_trait_limit),
        )
        return 0

    manifest_path = Path(args.download_manifest).expanduser().resolve()
    species_cds_dir = Path(args.species_cds_dir).expanduser().resolve()
    trait_plan_path = Path(args.trait_plan).expanduser().resolve()
    db_sources_path = Path(args.database_sources).expanduser().resolve()
    downloads_dir = Path(args.downloads_dir).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    stats_output_path = Path(args.stats_output).expanduser().resolve() if args.stats_output else None

    warnings: List[str] = []
    errors: List[str] = []

    try:
        target_species = validate_species_source(
            species_source=args.species_source,
            manifest_path=manifest_path,
            species_cds_dir=species_cds_dir,
        )
    except Exception as exc:
        parser.error(str(exc))

    _log("Target species source: {}".format(args.species_source))
    _log("Resolved target species: {}".format(len(target_species)))

    if not trait_plan_path.exists():
        message = "Trait plan not found: {}".format(trait_plan_path)
        if args.strict:
            parser.error(message)
        warnings.append(message)
        plan_rows: List[TraitPlanRow] = []
    else:
        plan_rows = read_trait_plan(trait_plan_path)

    if db_sources_path.exists():
        source_rows = read_database_sources(db_sources_path)
    else:
        source_rows = {}
        warnings.append("Database source map not found: {}".format(db_sources_path))

    requested_databases = resolve_requested_databases(
        databases_arg=args.databases,
        plan_rows=plan_rows,
        source_rows=source_rows,
    )
    if len(requested_databases) == 0 and len(plan_rows) > 0:
        requested_databases = sorted({row.database for row in plan_rows})

    species_sorted = sorted(target_species)
    result = pandas.DataFrame({"species": species_sorted}).set_index("species")
    db_frames: Dict[str, pandas.DataFrame] = {}
    db_configs: Dict[str, Dict[str, str]] = {}

    for database in requested_databases:
        config = source_rows.get(database, {})
        default_mode = SUPPORTED_DATABASES.get(database, {}).get("acquisition_mode", "bulk")
        if "acquisition_mode" not in config:
            config = dict(config)
            config["acquisition_mode"] = default_mode
        try:
            db_table = load_database_table(
                database=database,
                config=config,
                plan_rows=plan_rows,
                species=species_sorted,
                downloads_dir=downloads_dir,
                timeout=float(args.download_timeout),
                dry_run=bool(args.dry_run),
            )
        except Exception as exc:
            message = "[{}] failed to load source: {}".format(database, exc)
            if args.strict:
                errors.append(message)
            else:
                warnings.append(message)
            continue

        if db_table is None:
            warnings.append("[{}] source table is unavailable.".format(database))
            continue
        if db_table.shape[0] == 0:
            warnings.append("[{}] source table is empty.".format(database))
            continue

        db_configs[database] = config
        species_column = detect_species_column(db_table, config.get("species_column", ""))
        db_table = db_table.copy()
        db_table["__species_norm"] = db_table[species_column].map(normalize_species_name)
        db_table = db_table.loc[db_table["__species_norm"].isin(target_species), :]
        if db_table.shape[0] == 0:
            warnings.append("[{}] no rows matched target species.".format(database))
            continue
        db_frames[database] = db_table
        _log("[{}] matched rows: {}".format(database, db_table.shape[0]))

    for plan_row in plan_rows:
        if plan_row.database not in requested_databases:
            continue
        db_df = db_frames.get(plan_row.database)
        if db_df is None:
            message = "[{}] no data frame available for trait '{}'.".format(
                plan_row.database,
                plan_row.output_trait,
            )
            if args.strict:
                errors.append(message)
            else:
                warnings.append(message)
            continue
        db_config = db_configs.get(plan_row.database, {})
        db_filtered = db_df
        if plan_row.trait_key != "":
            trait_key_column = plan_row.trait_key_column or str(db_config.get("trait_key_column", "") or "").strip()
            if trait_key_column == "":
                trait_key_column = "trait_name"
            if trait_key_column not in db_filtered.columns:
                message = "[{}] missing trait_key_column '{}'.".format(plan_row.database, trait_key_column)
                if args.strict:
                    errors.append(message)
                else:
                    warnings.append(message)
                continue
            db_filtered = db_filtered.loc[
                db_filtered[trait_key_column].astype(str).str.strip() == plan_row.trait_key,
                :,
            ]
            if (
                db_filtered.shape[0] == 0
                and plan_row.database == "gift"
                and trait_key_column == "trait_ID"
                and "trait_token" in db_df.columns
            ):
                # Allow GIFT plan rows to use trait names in trait_key while source uses resolved trait_ID.
                db_filtered = db_df.loc[
                    db_df["trait_token"].astype(str).str.strip() == plan_row.trait_key,
                    :,
                ]
            if db_filtered.shape[0] == 0:
                message = "[{}] no rows matched trait_key '{}' in '{}'.".format(
                    plan_row.database,
                    plan_row.trait_key,
                    trait_key_column,
                )
                if args.strict:
                    errors.append(message)
                else:
                    warnings.append(message)
                continue

        if plan_row.source_column not in db_filtered.columns:
            message = "[{}] missing source column '{}'.".format(plan_row.database, plan_row.source_column)
            if args.strict:
                errors.append(message)
            else:
                warnings.append(message)
            continue

        aggregated = aggregate_trait_column(db_df=db_filtered, plan_row=plan_row)
        colname = plan_row.output_trait
        if colname not in result.columns:
            result[colname] = ""
        for species_name, value in aggregated.items():
            if species_name not in result.index:
                continue
            formatted = format_output_value(value)
            if formatted == "":
                continue
            existing = str(result.at[species_name, colname] or "").strip()
            if existing == "":
                result.at[species_name, colname] = formatted

    if len(errors) > 0:
        for message in errors:
            _log("ERROR: {}".format(message))
        return 1

    output_df = result.reset_index()
    trait_columns = [col for col in output_df.columns if col != "species"]
    num_species_with_any_trait = 0
    if len(trait_columns) > 0:
        has_any = output_df.loc[:, trait_columns].astype(str).apply(
            lambda row: any(value.strip() != "" for value in row.tolist()),
            axis=1,
        )
        num_species_with_any_trait = int(has_any.sum())

    if args.strict and len(trait_columns) == 0:
        _log("ERROR: No trait columns were generated.")
        return 1

    if args.dry_run:
        _log("[dry-run] species_trait output would be written to: {}".format(output_path))
    else:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_df.to_csv(output_path, sep="\t", index=False)
        _log("species_trait.tsv written: {}".format(output_path))

    if stats_output_path is not None:
        stats = {
            "species_source": args.species_source,
            "num_target_species": len(species_sorted),
            "num_trait_columns": len(trait_columns),
            "num_species_with_any_trait": num_species_with_any_trait,
            "num_requested_databases": len(requested_databases),
            "num_loaded_databases": len(db_frames),
            "output_path": str(output_path),
            "dry_run": int(bool(args.dry_run)),
        }
        if not args.dry_run:
            stats_output_path.parent.mkdir(parents=True, exist_ok=True)
            with open(stats_output_path, "wt", encoding="utf-8") as handle:
                json.dump(stats, handle, indent=2, ensure_ascii=False)
        else:
            _log("[dry-run] stats output would be written to: {}".format(stats_output_path))

    for message in warnings:
        _log("WARNING: {}".format(message))
    _log("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
