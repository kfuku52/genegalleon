#!/usr/bin/env python3
"""Download-manifest parsing and resolved-manifest helpers."""

import csv

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency guard
    load_workbook = None

RESOLVED_MANIFEST_PREFERRED_COLUMNS = (
    "line_number",
    "provider",
    "id",
    "species_key",
    "cds_url",
    "gff_url",
    "gbff_url",
    "genome_url",
    "cds_archive_member",
    "gff_archive_member",
    "gbff_archive_member",
    "genome_archive_member",
    "cds_filename",
    "gff_filename",
    "gbff_filename",
    "genome_filename",
    "fernbase_confidence_mode",
)
DIRECT_CATALOG_XLSX_SHEET = "_direct_catalog"
DIRECT_CATALOG_IGNORED_FIELDS = frozenset(("choice", "species", "provider"))


def manifest_has_annotation_source(gff_url, gbff_url):
    return str(gff_url or "").strip() != "" or str(gbff_url or "").strip() != ""


def manifest_has_cds_source(cds_url, genome_url, gbff_url):
    return (
        str(cds_url or "").strip() != ""
        or str(genome_url or "").strip() != ""
        or str(gbff_url or "").strip() != ""
    )


def manifest_has_usable_source_bundle(cds_url, gff_url, gbff_url, genome_url):
    if str(cds_url or "").strip() != "":
        return True
    if str(gbff_url or "").strip() != "":
        return True
    return str(gff_url or "").strip() != "" and str(genome_url or "").strip() != ""


def detect_manifest_delimiter(path):
    with open(path, "rt", encoding="utf-8") as handle:
        first_line = handle.readline()
    if "\t" in first_line:
        return "\t"
    return ","


def read_xlsx_table(sheet):
    row_iter = sheet.iter_rows(values_only=True)
    header_row = next(row_iter, None)
    if header_row is None:
        return [], []
    headers = []
    for value in header_row:
        text = str(value).strip() if value is not None else ""
        headers.append(text)
    while len(headers) > 0 and headers[-1] == "":
        headers.pop()
    if len(headers) == 0:
        return [], []

    rows = []
    for raw_values in row_iter:
        values = list(raw_values or ())
        if len(values) < len(headers):
            values.extend([None] * (len(headers) - len(values)))
        row = {}
        has_nonempty = False
        for idx, key in enumerate(headers):
            if key == "":
                continue
            cell = values[idx] if idx < len(values) else None
            text = str(cell).strip() if cell is not None else ""
            if text != "":
                has_nonempty = True
            row[key] = text
        if has_nonempty:
            rows.append(row)
    return headers, rows


def build_direct_catalog_lookup_from_xlsx(records):
    lookup = {}
    for record in records:
        for key in ("choice", "id"):
            value = str(record.get(key, "") or "").strip()
            if value != "" and value not in lookup:
                lookup[value] = record
    return lookup


def apply_direct_catalog_xlsx_defaults(rows, catalog_lookup):
    if len(catalog_lookup) == 0:
        return rows
    out = []
    for row in rows:
        updated = dict(row)
        provider = str(updated.get("provider", "") or "").strip().lower()
        source_id = str(updated.get("id", "") or "").strip()
        record = catalog_lookup.get(source_id)
        if (
            provider == "direct"
            and record is not None
            and not manifest_has_usable_source_bundle(
                updated.get("cds_url", ""),
                updated.get("gff_url", ""),
                updated.get("gbff_url", ""),
                updated.get("genome_url", ""),
            )
        ):
            catalog_id = str(record.get("id", "") or "").strip()
            if catalog_id != "":
                updated["id"] = catalog_id
            for key, value in record.items():
                if key in DIRECT_CATALOG_IGNORED_FIELDS:
                    continue
                text = str(value or "").strip()
                if text != "" and str(updated.get(key, "") or "").strip() == "":
                    updated[key] = text
        out.append(updated)
    return out


def read_download_manifest_xlsx(path):
    if load_workbook is None:
        raise ValueError("openpyxl is required to read .xlsx download manifests.")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        headers, rows = read_xlsx_table(sheet)
        if len(headers) == 0:
            return []
        if len(headers) < 2 or headers[0] != "provider" or headers[1] != "id":
            raise ValueError("XLSX download manifest must have 'provider' and 'id' as the first two columns.")

        if DIRECT_CATALOG_XLSX_SHEET in workbook.sheetnames:
            _catalog_headers, catalog_rows = read_xlsx_table(workbook[DIRECT_CATALOG_XLSX_SHEET])
            rows = apply_direct_catalog_xlsx_defaults(
                rows,
                build_direct_catalog_lookup_from_xlsx(catalog_rows),
            )
        return rows
    finally:
        workbook.close()


def read_download_manifest(path):
    if path.suffix.lower() == ".xlsx":
        return read_download_manifest_xlsx(path)
    delimiter = detect_manifest_delimiter(path)
    with open(path, "rt", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        rows = [row for row in reader]
    return rows


def resolved_manifest_fieldnames(rows):
    discovered = []
    for row in rows:
        for key in row.keys():
            if key not in discovered:
                discovered.append(key)
    ordered = []
    for key in RESOLVED_MANIFEST_PREFERRED_COLUMNS:
        ordered.append(key)
    for key in discovered:
        if key not in ordered:
            ordered.append(key)
    return ordered


def build_resolved_manifest_row(
    raw_row,
    fieldnames,
    provider,
    source_id,
    species_key,
    cds_url,
    gff_url,
    gbff_url,
    genome_url,
    cds_archive_member,
    gff_archive_member,
    gbff_archive_member,
    genome_archive_member,
    cds_filename,
    gff_filename,
    gbff_filename,
    genome_filename,
):
    row = {}
    for key in fieldnames:
        row[key] = str(raw_row.get(key) or "")
    row["provider"] = provider
    row["id"] = source_id
    row["species_key"] = species_key
    row["cds_url"] = cds_url
    row["gff_url"] = gff_url
    row["gbff_url"] = gbff_url
    row["genome_url"] = genome_url
    row["cds_archive_member"] = cds_archive_member
    row["gff_archive_member"] = gff_archive_member
    row["gbff_archive_member"] = gbff_archive_member
    row["genome_archive_member"] = genome_archive_member
    row["cds_filename"] = cds_filename
    row["gff_filename"] = gff_filename
    row["gbff_filename"] = gbff_filename
    row["genome_filename"] = genome_filename
    return row


def write_resolved_manifest_tsv(output_path, fieldnames, rows):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "wt", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: str(row.get(key) or "") for key in fieldnames})
