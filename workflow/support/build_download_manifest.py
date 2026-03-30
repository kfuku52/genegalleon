#!/usr/bin/env python3

import argparse
import csv
import json
from pathlib import Path
import re
import sys
from typing import Dict, Iterable, List, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.styles import Font
    from openpyxl.comments import Comment
    from openpyxl.worksheet.datavalidation import DataValidation
except Exception:  # pragma: no cover - exercised in runtime environments without openpyxl
    Workbook = None
    get_column_letter = None
    DefinedName = None
    Font = None
    Comment = None
    DataValidation = None


FASTA_EXTENSIONS = (
    ".fa",
    ".fas",
    ".fasta",
    ".fna",
    ".fa.bz2",
    ".fas.bz2",
    ".fasta.bz2",
    ".fna.bz2",
    ".fa.gz",
    ".fas.gz",
    ".fasta.gz",
    ".fna.gz",
    ".fa.tar.gz",
    ".fas.tar.gz",
    ".fasta.tar.gz",
    ".fna.tar.gz",
    ".fa.tar.bz2",
    ".fas.tar.bz2",
    ".fasta.tar.bz2",
    ".fna.tar.bz2",
)

GFF_EXTENSIONS = (
    ".gff",
    ".gff3",
    ".gtf",
    ".gff.gz",
    ".gff3.gz",
    ".gtf.gz",
)

GENBANK_EXTENSIONS = (
    ".gb",
    ".gbk",
    ".gbff",
    ".genbank",
    ".gb.gz",
    ".gbk.gz",
    ".gbff.gz",
    ".genbank.gz",
)

PROVIDERS = (
    "ensembl",
    "ensemblplants",
    "ncbi",
    "coge",
    "cngb",
    "gwh",
    "plantaedb",
    "flybase",
    "wormbase",
    "vectorbase",
    "fernbase",
    "veupathdb",
    "dictybase",
    "insectbase",
    "oryza_minuta",
    "direct",
    "local",
)
LEGACY_NCBI_PROVIDER_ALIASES = ("refseq", "genbank")
FERNBASE_CONFIDENCE_MODE_FIELD = "fernbase_confidence_mode"
FERNBASE_CONFIDENCE_MODE_HIGH_ONLY = "high-confidence only"
FERNBASE_CONFIDENCE_MODE_HIGH_LOW_COMBINED = "high-low combined"
FERNBASE_CONFIDENCE_MODE_CHOICES = (
    FERNBASE_CONFIDENCE_MODE_HIGH_ONLY,
    FERNBASE_CONFIDENCE_MODE_HIGH_LOW_COMBINED,
)
FERNBASE_COMBINED_FILENAME_MARKER = "highlowcombined"

DEFAULT_INPUT_RELATIVE_DIRS = {
    "ensembl": Path("Ensembl") / "original_files",
    "ensemblplants": Path("20230216_EnsemblPlants") / "original_files",
    "ncbi": Path("NCBI_Genome") / "species_wise_original",
    "coge": Path("CoGe") / "species_wise_original",
    "cngb": Path("CNGB") / "species_wise_original",
    "gwh": Path("GWH") / "species_wise_original",
    "plantaedb": Path("PlantaeDB") / "species_wise_original",
    "flybase": Path("FlyBase") / "species_wise_original",
    "wormbase": Path("WormBase") / "species_wise_original",
    "vectorbase": Path("VectorBase") / "species_wise_original",
    "fernbase": Path("FernBase") / "species_wise_original",
    "veupathdb": Path("VEuPathDB") / "species_wise_original",
    "dictybase": Path("dictyBase") / "species_wise_original",
    "insectbase": Path("InsectBase") / "species_wise_original",
    "oryza_minuta": Path("OryzaMinuta") / "species_wise_original",
    "direct": Path("Direct") / "species_wise_original",
    "local": Path("Local") / "species_wise_original",
}
NCBI_MERGED_INPUT_RELATIVE_DIRS = (
    Path("NCBI_Genome") / "species_wise_original",
    Path("NCBI_RefSeq") / "species_wise_original",
    Path("NCBI_GenBank") / "species_wise_original",
)

MANIFEST_FIELDNAMES = (
    "provider",
    "id",
    "species_key",
    "cds_url",
    "gff_url",
    "gbff_url",
    "genome_url",
    "cds_archive_member",
    "gff_archive_member",
    "gbff_archive_member",
    "genome_archive_member",
    "cds_filename",
    "gff_filename",
    "gbff_filename",
    "genome_filename",
    "cds_url_template",
    "gff_url_template",
    "genome_url_template",
    "local_cds_path",
    "local_gff_path",
    "local_gbff_path",
    "local_genome_path",
    FERNBASE_CONFIDENCE_MODE_FIELD,
)

DIRECT_CATALOG_FIELDS = tuple(field for field in MANIFEST_FIELDNAMES if field not in ("provider", "id"))
DIRECT_CATALOG_SNAPSHOT_FIELDS = ("id", "species") + DIRECT_CATALOG_FIELDS

HEADER_COMMENTS = {
    "provider": "\n".join(
        (
            "Required.",
            "",
            "Allowed values: {}.".format(", ".join(PROVIDERS)),
            "Use the drop-down when possible.",
            "In XLSX templates this must remain the first column.",
        )
    ),
    "id": "\n".join(
        (
            "Required.",
            "",
            "Provider-specific identifier.",
            "Examples:",
            "- ensembl: homo_sapiens",
            "- ncbi: GCF_000001405.40 or GCA_000001635.9",
            "- coge: numeric genome_id (gid), for example 24739",
            "- cngb: CNA... or GCA/GCF accession",
            "- gwh: GWH... accession, for example GWHIGRM00000000.1",
            "- plantaedb: PlantaeDB species page URL or /taxa/... path",
            "- veupathdb: EnuttalliP19",
            "- dictybase: Dictyostelium_discoideum",
            "- insectbase: IBG_00001",
            "- oryza_minuta: gramene_tetraploids",
            "- direct: stable label or index URL token for explicit URLs",
            "- local: local species directory ID or path-style ID",
            "The drop-down is provider-specific, but any valid value can still be typed manually.",
            "For non-local providers, labels like 'GCF_000001405.40 (Homo sapiens)' are accepted;",
            "the parser uses the token before the first space as the actual ID.",
            "In XLSX templates this must remain the second column.",
        )
    ),
    "species_key": "\n".join(
        (
            "Optional.",
            "",
            "Stable species label used for output directories and filename templating.",
            "Recommended format: letters, numbers, and underscores only, for example Homo_sapiens.",
            "If blank, it is inferred from provider metadata, local directory names, or id when possible.",
        )
    ),
    "cds_url": "\n".join(
        (
            "Conditionally required.",
            "",
            "URL to the CDS FASTA file.",
            "Accepted schemes: https://, http://, ftp://, file://.",
            "Leave blank when provider/id auto-resolution or cds_url_template will fill it.",
            "May also remain blank when both gff_url and genome_url are available; CDS can be derived later.",
            "Typical suffixes: .fa, .fasta, .fna with optional .gz.",
        )
    ),
    "gff_url": "\n".join(
        (
            "Conditionally required.",
            "",
            "URL to the annotation file.",
            "Accepted schemes: https://, http://, ftp://, file://.",
            "Leave blank when provider/id auto-resolution or gff_url_template will fill it.",
            "Typical suffixes: .gff, .gff3, .gtf with optional .gz.",
        )
    ),
    "gbff_url": "\n".join(
        (
            "Optional fallback annotation source.",
            "",
            "URL to a GenBank/GBFF annotation file.",
            "Accepted schemes: https://, http://, ftp://, file://.",
            "Use this when GFF is unavailable but GBFF plus genome (or GBFF alone) can drive CDS/GFF derivation.",
            "Typical suffixes: .gbff, .gbk, .gb with optional .gz.",
        )
    ),
    "genome_url": "\n".join(
        (
            "Conditionally required when cds_url is blank.",
            "",
            "URL to the genome FASTA file.",
            "Accepted schemes: https://, http://, ftp://, file://.",
            "Leave blank to skip genome download when CDS is already available; provider/id auto-resolution or genome_url_template can fill it.",
            "Typical suffixes: .fa, .fasta, .fna with optional .gz.",
        )
    ),
    "cds_archive_member": "\n".join(
        (
            "Optional.",
            "",
            "Path to the CDS file inside an archive referenced by cds_url.",
            "Use this when cds_url points to a .zip or .tar.* bundle instead of a plain FASTA.",
            "Example: SpeciesA/SpeciesA.cds.fasta",
        )
    ),
    "gff_archive_member": "\n".join(
        (
            "Optional.",
            "",
            "Path to the GFF/GTF file inside an archive referenced by gff_url.",
            "Use this when gff_url points to a .zip or .tar.* bundle instead of a plain annotation file.",
            "Example: SpeciesA/SpeciesA.genes.gff3",
        )
    ),
    "gbff_archive_member": "\n".join(
        (
            "Optional.",
            "",
            "Path to the GBFF/GenBank file inside an archive referenced by gbff_url.",
            "Use this when gbff_url points to a .zip or .tar.* bundle instead of a plain annotation file.",
            "Example: SpeciesA/SpeciesA.genomic.gbff",
        )
    ),
    "genome_archive_member": "\n".join(
        (
            "Optional.",
            "",
            "Path to the genome FASTA inside an archive referenced by genome_url.",
            "Use this when genome_url points to a .zip or .tar.* bundle instead of a plain FASTA.",
            "Example: SpeciesA/SpeciesA.fasta",
        )
    ),
    "cds_filename": "\n".join(
        (
            "Optional.",
            "",
            "Output filename used when the CDS file is written to the raw download directory.",
            "Use a basename only, not a directory path.",
            "If blank, it is inferred from cds_archive_member, cds_url, or provider metadata.",
        )
    ),
    "gff_filename": "\n".join(
        (
            "Optional.",
            "",
            "Output filename used when the GFF/GTF file is written to the raw download directory.",
            "Use a basename only, not a directory path.",
            "If blank, it is inferred from gff_archive_member, gff_url, or provider metadata.",
        )
    ),
    "gbff_filename": "\n".join(
        (
            "Optional.",
            "",
            "Output filename used when the GBFF/GenBank file is written to the raw download directory.",
            "Use a basename only, not a directory path.",
            "If blank, it is inferred from gbff_archive_member, gbff_url, or provider metadata.",
        )
    ),
    "genome_filename": "\n".join(
        (
            "Optional.",
            "",
            "Output filename used when the genome FASTA is written to the raw download directory.",
            "Use a basename only, not a directory path.",
            "If blank, it is inferred from genome_archive_member, genome_url, or provider metadata.",
        )
    ),
    "cds_url_template": "\n".join(
        (
            "Optional.",
            "",
            "Template used to build cds_url when cds_url is blank.",
            "Supported placeholders: {id}, {species_key}, {provider}.",
            "Example: https://example.org/{species_key}/{id}.cds.fa.gz",
        )
    ),
    "gff_url_template": "\n".join(
        (
            "Optional.",
            "",
            "Template used to build gff_url when gff_url is blank.",
            "Supported placeholders: {id}, {species_key}, {provider}.",
            "Example: https://example.org/{species_key}/{id}.gff3.gz",
        )
    ),
    "genome_url_template": "\n".join(
        (
            "Optional.",
            "",
            "Template used to build genome_url when genome_url is blank.",
            "Supported placeholders: {id}, {species_key}, {provider}.",
            "Example: https://example.org/{species_key}/{id}.genome.fa.gz",
        )
    ),
    "local_cds_path": "\n".join(
        (
            "Optional; mainly for provider=local.",
            "",
            "Absolute path or path relative to this manifest file.",
            "Converted to file:// at runtime.",
            "Use this when the CDS already exists locally and you do not want to write cds_url manually.",
        )
    ),
    "local_gff_path": "\n".join(
        (
            "Optional; mainly for provider=local.",
            "",
            "Absolute path or path relative to this manifest file.",
            "Converted to file:// at runtime.",
            "Use this when the annotation file already exists locally and you do not want to write gff_url manually.",
        )
    ),
    "local_gbff_path": "\n".join(
        (
            "Optional; mainly for provider=local.",
            "",
            "Absolute path or path relative to this manifest file.",
            "Converted to file:// at runtime.",
            "Use this when only a GBFF/GenBank annotation file exists locally and you do not want to write gbff_url manually.",
        )
    ),
    "local_genome_path": "\n".join(
        (
            "Optional; mainly for provider=local.",
            "",
            "Absolute path or path relative to this manifest file.",
            "Converted to file:// at runtime.",
            "Use this when the genome FASTA already exists locally and you do not want to write genome_url manually.",
        )
    ),
    FERNBASE_CONFIDENCE_MODE_FIELD: "\n".join(
        (
            "Optional; only used for legacy FernBase releases with separate highconfidence/lowconfidence files.",
            "",
            "Allowed values: {} ; {}.".format(
                FERNBASE_CONFIDENCE_MODE_HIGH_ONLY,
                FERNBASE_CONFIDENCE_MODE_HIGH_LOW_COMBINED,
            ),
            "Blank defaults to high-confidence only.",
            "Ignored for non-FernBase rows and for newer single-annotation FernBase releases.",
        )
    ),
}


def missing_annotation_label(cds_path, gff_path, gbff_path, genome_path):
    if cds_path is not None or gbff_path is not None or (gff_path is not None and genome_path is not None):
        return ""
    if cds_path is None and gff_path is None and gbff_path is None and genome_path is None:
        return "CDS-or-GBFF-or-(GFF+genome)"
    if gff_path is not None and genome_path is None:
        return "CDS-or-GBFF-or-genome"
    if genome_path is not None and gff_path is None and gbff_path is None:
        return "CDS-or-GBFF-or-GFF"
    if gbff_path is None:
        return "CDS-or-GBFF-or-(GFF+genome)"
    return ""

HEADER_COMMENT_MIN_WIDTH_PX = 240
HEADER_COMMENT_MAX_WIDTH_PX = 520
HEADER_COMMENT_MIN_HEIGHT_PX = 100
HEADER_COMMENT_CHAR_WIDTH_PX = 7
HEADER_COMMENT_LINE_HEIGHT_PX = 18
HEADER_COMMENT_PADDING_X_PX = 24
HEADER_COMMENT_PADDING_Y_PX = 18


def build_header_comment(fieldname):
    if Comment is None:
        return None
    text = str(HEADER_COMMENTS.get(fieldname, "") or "").strip()
    if text == "":
        return None

    lines = text.splitlines() or [text]
    longest_line = max(len(line) for line in lines) if lines else 0
    width = HEADER_COMMENT_PADDING_X_PX + (longest_line * HEADER_COMMENT_CHAR_WIDTH_PX)
    width = max(HEADER_COMMENT_MIN_WIDTH_PX, min(HEADER_COMMENT_MAX_WIDTH_PX, width))
    wrap_chars = max(1, (width - HEADER_COMMENT_PADDING_X_PX) // HEADER_COMMENT_CHAR_WIDTH_PX)

    display_lines = 0
    for line in lines:
        if line == "":
            display_lines += 1
            continue
        display_lines += 1 + max(0, (len(line) - 1) // wrap_chars)

    height = HEADER_COMMENT_PADDING_Y_PX + (display_lines * HEADER_COMMENT_LINE_HEIGHT_PX)
    height = max(HEADER_COMMENT_MIN_HEIGHT_PX, height)
    return Comment(text, "genegalleon", width=width, height=height)

LARGE_ID_PROVIDERS = ("ncbi",)
SNAPSHOT_FULL_ID_PROVIDERS = (
    "ensembl",
    "ensemblplants",
    "gwh",
    "plantaedb",
    "flybase",
    "wormbase",
    "vectorbase",
    "fernbase",
    "veupathdb",
    "dictybase",
    "insectbase",
    "oryza_minuta",
    "direct",
    "local",
)
EXAMPLE_ONLY_PROVIDERS = LARGE_ID_PROVIDERS + ("coge", "cngb")

ID_EXAMPLES_BY_PROVIDER = {
    "ensembl": (("homo_sapiens", "Homo sapiens"), ("mus_musculus", "Mus musculus")),
    "ensemblplants": (("Ostreococcus_lucimarinus", "Ostreococcus lucimarinus"),),
    "ncbi": (
        ("GCF_000001405.40", "Homo sapiens"),
        ("GCA_000001635.9", "Mus musculus"),
        ("GCF_049306965.1", "Danio rerio"),
        ("GCA_000001215.4", "Drosophila melanogaster"),
        ("GCF_000002985.6", "Caenorhabditis elegans"),
    ),
    "coge": (
        ("24739", "Arabidopsis thaliana"),
        ("29177", "Oryza sativa"),
        ("42091", "Zea mays"),
        ("78085", "Drosophila melanogaster"),
        ("72417", "Caenorhabditis elegans"),
    ),
    "cngb": (
        ("CNA0012345", "Homo sapiens"),
        ("GCF_000001405.40", "Homo sapiens"),
        ("GCA_000001635.9", "Mus musculus"),
        ("GCF_049306965.1", "Danio rerio"),
        ("GCA_000001215.4", "Drosophila melanogaster"),
    ),
    "gwh": (
        ("GWHIGRM00000000.1", "Medicago sativa"),
        ("GWHCBHY00000000", "Allium sativum"),
    ),
    "plantaedb": (
        (
            "https://plantaedb.com/taxa/phylum/angiosperms/order/asterales/family/asteraceae/subfamily/asteroideae/tribe/astereae/subtribe/conyzinae/genus/erigeron/species/erigeron-breviscapus",
            "Erigeron breviscapus (PlantaeDB page)",
        ),
        (
            "https://plantaedb.com/taxa/phylum/angiosperms/order/ranunculales/family/berberidaceae/genus/berberis/species/berberis-thunbergii",
            "Berberis thunbergii (PlantaeDB page)",
        ),
    ),
    "flybase": (("dmel_r6.61", "Drosophila melanogaster"),),
    "wormbase": (("celegans_prjna13758_ws290", "Caenorhabditis elegans"),),
    "vectorbase": (("anopheles_gambiae_pest", "Anopheles gambiae"),),
    "fernbase": (("Azolla_filiculoides", "Azolla filiculoides"), ("Salvinia_cucullata_v2", "Salvinia cucullata v2")),
    "veupathdb": (("EnuttalliP19", "Entamoeba nuttalli"),),
    "dictybase": (("Dictyostelium_discoideum", "Dictyostelium discoideum"),),
    "insectbase": (("IBG_00001", "Abrostola tripartita"),),
    "oryza_minuta": (("gramene_tetraploids", "Oryza minuta Gramene tetraploids"),),
    "direct": (("direct_example_species", "Direct URL manifest row"),),
    "local": (("/absolute/path/to/local/species_dir", "Local species directory"),),
}


def species_label_from_species_key(species_key: str) -> str:
    text = str(species_key or "").strip()
    if text == "":
        return ""
    parts = text.replace("_", " ").split()
    normalized = []
    for idx, part in enumerate(parts):
        if re.fullmatch(r"[a-z]+", part):
            if idx == 0:
                normalized.append(part.capitalize())
            else:
                normalized.append(part.lower())
        else:
            normalized.append(part)
    return " ".join(normalized)


def format_id_choice(source_id: str, species_label: str) -> str:
    sid = str(source_id or "").strip()
    label = str(species_label or "").strip()
    if sid == "":
        return ""
    if label == "":
        return sid
    return "{} ({})".format(sid, label)


def dedupe_preserve_order(items):
    out = []
    seen = set()
    for item in items:
        if item in seen:
            continue
        seen.add(item)
        out.append(item)
    return out


def parse_snapshot_entries(raw_entries):
    out = []
    entries = raw_entries
    if isinstance(entries, dict):
        entries = [entries]
    if not isinstance(entries, list):
        return out
    for entry in entries:
        source_id = ""
        species_label = ""
        if isinstance(entry, str):
            source_id = entry
        elif isinstance(entry, dict):
            source_id = str(entry.get("id") or entry.get("source_id") or "").strip()
            species_label = str(entry.get("species") or entry.get("species_label") or entry.get("label") or "").strip()
        elif isinstance(entry, (tuple, list)) and len(entry) > 0:
            source_id = str(entry[0] or "").strip()
            if len(entry) > 1:
                species_label = str(entry[1] or "").strip()
        source_id = str(source_id or "").strip()
        if source_id == "":
            continue
        out.append((source_id, species_label))
    return dedupe_preserve_order(out)


def parse_direct_catalog_entries(raw_entries):
    out = []
    entries = raw_entries
    if isinstance(entries, dict):
        entries = [entries]
    if not isinstance(entries, list):
        return out
    seen = set()
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        source_id = str(entry.get("id") or entry.get("source_id") or "").strip()
        species_label = str(entry.get("species") or entry.get("species_label") or entry.get("label") or "").strip()
        if source_id == "":
            continue
        key = (source_id, species_label)
        if key in seen:
            continue
        seen.add(key)
        record = {"id": source_id, "species": species_label}
        for field in DIRECT_CATALOG_FIELDS:
            record[field] = str(entry.get(field) or "").strip()
        out.append(record)
    return out


def load_snapshot_payload(path: Path):
    with open(path, "rt", encoding="utf-8") as handle:
        payload = json.load(handle)
    providers_blob = payload.get("providers", payload)
    if not isinstance(providers_blob, dict):
        raise ValueError("id options snapshot must contain object key 'providers'")
    return providers_blob


def load_id_options_snapshot(path: Path):
    providers_blob = load_snapshot_payload(path)
    options = {}
    for provider in PROVIDERS:
        options[provider] = parse_snapshot_entries(providers_blob.get(provider, []))
    return options


def load_direct_catalog_snapshot(path: Path):
    providers_blob = load_snapshot_payload(path)
    return parse_direct_catalog_entries(providers_blob.get("direct", []))


def load_manifest_rows(path: Path):
    suffix = str(path.suffix or "").lower()
    if suffix == ".xlsx":
        if Workbook is None:
            raise RuntimeError("openpyxl is required to read .xlsx direct catalog manifests")
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            sheet = workbook.active
            values = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()
        if len(values) == 0:
            return []
        headers = [str(value or "").strip() for value in values[0]]
        rows = []
        for row_values in values[1:]:
            if row_values is None:
                continue
            record = {}
            nonempty = False
            for idx, field in enumerate(headers):
                if field == "":
                    continue
                value = row_values[idx] if idx < len(row_values) else ""
                text = str(value or "").strip()
                if text != "":
                    nonempty = True
                record[field] = text
            if nonempty:
                rows.append(record)
        return rows
    delimiter = "\t" if suffix == ".tsv" else ","
    with open(path, "rt", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter=delimiter))


def build_direct_catalog_from_manifest_rows(rows):
    out = []
    seen = set()
    for row in rows:
        provider = str(row.get("provider", "") or "").strip().lower()
        if provider != "direct":
            continue
        source_id = str(row.get("id", "") or "").strip()
        if source_id == "":
            continue
        species_key = str(row.get("species_key", "") or "").strip()
        species_label = species_label_from_species_key(species_key)
        key = (source_id, species_key, species_label)
        if key in seen:
            continue
        seen.add(key)
        record = {"id": source_id, "species": species_label}
        for field in DIRECT_CATALOG_FIELDS:
            record[field] = str(row.get(field, "") or "").strip()
        out.append(record)
    return sorted(out, key=lambda record: (str(record.get("id", "")).lower(), str(record.get("species_key", "")).lower()))


def infer_coge_genome_id_from_files(species_key: str, files: List[Path], warnings: List[str]) -> str:
    matches = set()
    key_match = re.search(r"(?:^|[._-])gid([0-9]+)(?:[._-]|$)", species_key, flags=re.IGNORECASE)
    if key_match is not None:
        matches.add(key_match.group(1))
    for path in files:
        name = path.name
        for match in re.finditer(r"(?:^|[._-])gid([0-9]+)(?:[._-]|$)", name, flags=re.IGNORECASE):
            matches.add(match.group(1))
        for match in re.finditer(r"dsgid[=_-]?([0-9]+)", name, flags=re.IGNORECASE):
            matches.add(match.group(1))
    if len(matches) == 0:
        return ""
    ordered = sorted(matches)
    if len(ordered) > 1:
        warnings.append(
            "[coge] {}: multiple genome_id candidates detected {}. Using '{}'".format(
                species_key, ",".join(ordered), ordered[0]
            )
        )
    return ordered[0]


def infer_gwh_accession_from_files(species_key: str, files: List[Path], warnings: List[str]) -> str:
    matches = set()
    for text in [species_key] + [path.name for path in files]:
        for match in re.finditer(r"(GWH[A-Z0-9]+(?:\.[0-9]+)?)", text, flags=re.IGNORECASE):
            matches.add(match.group(1).upper())
    if len(matches) == 0:
        return ""
    ordered = sorted(matches)
    if len(ordered) > 1:
        warnings.append(
            "[gwh] {}: multiple GWH accession candidates detected {}. Using '{}'".format(
                species_key, ",".join(ordered), ordered[0]
            )
        )
    return ordered[0]


def build_provider_id_options(rows: List[Dict[str, str]], snapshot_options=None):
    if snapshot_options is None:
        snapshot_options = {}
    discovered: Dict[str, Dict[str, str]] = {provider: {} for provider in PROVIDERS}
    for row in rows:
        provider = str(row.get("provider", "") or "").strip().lower()
        source_id = str(row.get("id", "") or "").strip()
        species_label = species_label_from_species_key(str(row.get("species_key", "") or ""))
        if provider not in discovered or source_id == "":
            continue
        if source_id not in discovered[provider]:
            discovered[provider][source_id] = species_label

    options = {}
    for provider in PROVIDERS:
        values = []
        if provider in EXAMPLE_ONLY_PROVIDERS:
            values = [
                format_id_choice(source_id, species_label)
                for source_id, species_label in ID_EXAMPLES_BY_PROVIDER.get(provider, ())
            ]
        else:
            snapshot_entries = list(snapshot_options.get(provider, []))
            if provider in SNAPSHOT_FULL_ID_PROVIDERS and len(snapshot_entries) > 0:
                if provider == "local":
                    values = [str(source_id).strip() for source_id, _ in snapshot_entries if str(source_id).strip()]
                else:
                    values = [
                        format_id_choice(source_id, species_label if species_label != "" else species_label_from_species_key(source_id))
                        for source_id, species_label in snapshot_entries
                    ]
            if len(values) == 0:
                if provider == "local":
                    values = [str(source_id).strip() for source_id in sorted(discovered[provider].keys()) if str(source_id).strip()]
                    if len(values) == 0:
                        values = [str(source_id).strip() for source_id, _species_label in ID_EXAMPLES_BY_PROVIDER.get(provider, ())]
                else:
                    values = [
                        format_id_choice(source_id, discovered[provider].get(source_id, ""))
                        for source_id in sorted(discovered[provider].keys())
                    ]
                    if len(values) == 0:
                        values = [
                            format_id_choice(source_id, species_label)
                            for source_id, species_label in ID_EXAMPLES_BY_PROVIDER.get(provider, ())
                        ]
        if len(values) == 0:
            values = [""]
        options[provider] = dedupe_preserve_order(values)
    return options


def build_direct_catalog_choice(entry):
    source_id = str(entry.get("id", "") or "").strip()
    species_label = str(entry.get("species", "") or "").strip()
    if species_label == "":
        species_label = species_label_from_species_key(str(entry.get("species_key", "") or ""))
    if species_label == "":
        species_label = species_label_from_species_key(source_id)
    return format_id_choice(source_id, species_label)


def direct_catalog_formula(catalog_column_letter, catalog_row_end, sheet_row):
    choice_range = "_direct_catalog!$A$2:$A${}".format(catalog_row_end)
    id_range = "_direct_catalog!$B$2:$B${}".format(catalog_row_end)
    value_range = "_direct_catalog!${}$2:${}${}".format(catalog_column_letter, catalog_column_letter, catalog_row_end)
    return (
        '=IF($A{row}<>"direct","",'
        'IFERROR(INDEX({value_range},MATCH($B{row},{choice_range},0)),'
        'IFERROR(INDEX({value_range},MATCH($B{row},{id_range},0)),"")))'
    ).format(
        row=sheet_row,
        value_range=value_range,
        choice_range=choice_range,
        id_range=id_range,
    )


def build_arg_parser():
    parser = argparse.ArgumentParser(
        description=(
            "Build a download manifest from an existing local provider-root directory. "
            "Generated URLs are file:// URLs for reuse in format_species_inputs.py --download-manifest."
        )
    )
    parser.add_argument(
        "--provider",
        choices=("all",) + PROVIDERS + LEGACY_NCBI_PROVIDER_ALIASES,
        default="all",
        help="Provider to scan (default: all).",
    )
    parser.add_argument(
        "--input-dir",
        required=True,
        help="Root directory containing provider subdirectories used by format_species_inputs.py.",
    )
    parser.add_argument(
        "--output",
        default="workspace/input/input_generation/download_plan.xlsx",
        help="Output path (.xlsx recommended; .tsv/.csv also supported).",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Exit with error when a species lacks a usable source bundle (CDS, GBFF, or GFF plus genome).",
    )
    parser.add_argument(
        "--id-options-snapshot",
        default="",
        help=(
            "Optional JSON snapshot for provider-specific id dropdown values. "
            "When set, non-large providers (ensembl/ensemblplants/gwh/flybase/wormbase/vectorbase/fernbase/veupathdb/dictybase/insectbase/direct/local) "
            "prefer snapshot IDs over locally discovered IDs."
        ),
    )
    parser.add_argument(
        "--direct-catalog-manifest",
        default="",
        help=(
            "Optional TSV/CSV/XLSX manifest containing provider=direct rows. "
            "When set, direct dropdown candidates and direct auto-fill catalog are read from this manifest."
        ),
    )
    return parser


def is_fasta_filename(name):
    lower = name.lower()
    return any(lower.endswith(ext) for ext in FASTA_EXTENSIONS)


def is_gff_filename(name):
    lower = name.lower()
    return any(lower.endswith(ext) for ext in GFF_EXTENSIONS)


def is_gbff_filename(name):
    lower = name.lower()
    return any(lower.endswith(ext) for ext in GENBANK_EXTENSIONS)


def is_probable_genome_filename(provider, name):
    if not is_fasta_filename(name):
        return False
    lower = name.lower()
    non_genome_markers = (
        "cds",
        "mrna",
        "trna",
        "rrna",
        "ncrna",
        "rna_from_genomic",
        "transcript",
        "cdna",
        "pep",
        "protein",
    )
    if any(marker in lower for marker in non_genome_markers):
        return False
    if provider == "ncbi":
        return "genomic" in lower
    if provider in ("ensembl", "ensemblplants"):
        return any(marker in lower for marker in ("dna", "genome", "toplevel", "primary_assembly", "chromosome"))
    return any(marker in lower for marker in ("genome", "assembly", "genomic", "dna", "chromosome", "scaffold"))


FERNBASE_GFF_EXCLUDE_PATTERN = re.compile(
    r"(?:^|[._-])(te|teanno|repeat|repeats|transpos(?:on)?)(?:[._-]|$)",
    re.IGNORECASE,
)


def is_probable_cds_filename(provider, name):
    lower = name.lower()
    if provider == "fernbase" and (lower.endswith(".cds") or lower.endswith(".cds.gz")):
        return True
    if not is_fasta_filename(name):
        return False
    return any(marker in lower for marker in ("cds", "transcript", "mrna", "cdna"))


def infer_fernbase_confidence_mode_from_files(files, cds_path, gff_path):
    if cds_path is None or gff_path is None:
        return ""
    cds_lower = cds_path.name.lower()
    gff_lower = gff_path.name.lower()
    if FERNBASE_COMBINED_FILENAME_MARKER in cds_lower or FERNBASE_COMBINED_FILENAME_MARKER in gff_lower:
        return FERNBASE_CONFIDENCE_MODE_HIGH_LOW_COMBINED
    has_high_cds = any("highconfidence" in path.name.lower() and is_probable_cds_filename("fernbase", path.name) for path in files)
    has_low_cds = any("lowconfidence" in path.name.lower() and is_probable_cds_filename("fernbase", path.name) for path in files)
    has_high_gff = any("highconfidence" in path.name.lower() and is_gff_filename(path.name) for path in files)
    has_low_gff = any("lowconfidence" in path.name.lower() and is_gff_filename(path.name) for path in files)
    if has_high_cds and has_low_cds and has_high_gff and has_low_gff:
        return FERNBASE_CONFIDENCE_MODE_HIGH_ONLY
    return ""


def provider_candidate_sort_key(provider, label, name):
    lower = str(name or "").lower()
    label_upper = str(label or "").upper()
    if provider != "fernbase":
        return (lower,)
    if label_upper == "CDS":
        return (
            0 if FERNBASE_COMBINED_FILENAME_MARKER in lower else 1,
            0 if "highconfidence" in lower else 1,
            1 if "lowconfidence" in lower else 0,
            0 if "cds" in lower else 1,
            1 if any(marker in lower for marker in ("transcript", "mrna", "cdna")) else 0,
            lower,
        )
    if label_upper == "GFF":
        return (
            1 if FERNBASE_GFF_EXCLUDE_PATTERN.search(lower) else 0,
            0 if FERNBASE_COMBINED_FILENAME_MARKER in lower else 1,
            0 if "highconfidence" in lower else 1,
            1 if "lowconfidence" in lower else 0,
            lower,
        )
    if label_upper == "GENOME":
        return (
            1 if "chloroplast" in lower else 0,
            0 if any(marker in lower for marker in ("genome", "assembly", "chr", "chromosome")) else 1,
            lower,
        )
    return (lower,)


def pick_single_file(matches, provider, species_key, label, warnings):
    if len(matches) == 0:
        return None
    ordered = sorted(matches, key=lambda path: provider_candidate_sort_key(provider, label, path.name))
    if len(ordered) > 1:
        warnings.append(
            "[{}] {}: multiple {} files found. Using '{}'".format(
                provider, species_key, label, ordered[0].name
            )
        )
    return ordered[0]


def discover_ensembl_like(input_dir, provider):
    warnings = []
    errors = []
    rows = []

    species_to_cds = {}
    species_to_gff = {}
    species_to_gbff = {}
    species_to_genome = {}
    for path in sorted(input_dir.iterdir()):
        if not path.is_file():
            continue
        species_key = path.name.split(".", 1)[0]
        if species_key == "":
            continue
        if ".cds." in path.name.lower() and is_fasta_filename(path.name):
            species_to_cds.setdefault(species_key, []).append(path)
        elif is_gff_filename(path.name):
            species_to_gff.setdefault(species_key, []).append(path)
        elif is_gbff_filename(path.name):
            species_to_gbff.setdefault(species_key, []).append(path)
        elif is_probable_genome_filename(provider, path.name):
            species_to_genome.setdefault(species_key, []).append(path)

    for species_key in sorted(
        set(species_to_cds.keys()) | set(species_to_gff.keys()) | set(species_to_gbff.keys()) | set(species_to_genome.keys())
    ):
        cds_path = pick_single_file(species_to_cds.get(species_key, []), provider, species_key, "CDS", warnings)
        gff_path = pick_single_file(species_to_gff.get(species_key, []), provider, species_key, "GFF", warnings)
        gbff_path = pick_single_file(species_to_gbff.get(species_key, []), provider, species_key, "GBFF", warnings)
        genome_path = pick_single_file(
            species_to_genome.get(species_key, []), provider, species_key, "genome", warnings
        )
        missing_label = missing_annotation_label(cds_path, gff_path, gbff_path, genome_path)
        if missing_label != "":
            errors.append(
                "[{}] {}: missing {}".format(
                    provider,
                    species_key, missing_label
                )
            )
            continue
        rows.append(
            {
                "provider": provider,
                "id": species_key,
                "species_key": species_key,
                "cds_url": cds_path.resolve().as_uri() if cds_path is not None else "",
                "gff_url": gff_path.resolve().as_uri() if gff_path is not None else "",
                "gbff_url": gbff_path.resolve().as_uri() if gbff_path is not None else "",
                "genome_url": genome_path.resolve().as_uri() if genome_path is not None else "",
                "cds_filename": cds_path.name if cds_path is not None else "",
                "gff_filename": gff_path.name if gff_path is not None else "",
                "gbff_filename": gbff_path.name if gbff_path is not None else "",
                "genome_filename": genome_path.name if genome_path is not None else "",
                FERNBASE_CONFIDENCE_MODE_FIELD: "",
            }
        )

    return rows, warnings, errors


def discover_species_dir_based(provider, input_dir):
    warnings = []
    errors = []
    rows = []
    for species_dir in sorted(input_dir.iterdir()):
        if not species_dir.is_dir():
            continue
        species_key = species_dir.name
        files = [path for path in species_dir.iterdir() if path.is_file()]
        if provider == "phycocosm":
            cds_matches = [path for path in files if "fasta" in path.name.lower() and is_fasta_filename(path.name)]
            gff_matches = [path for path in files if "gff" in path.name.lower() and is_gff_filename(path.name)]
            gbff_matches = [path for path in files if is_gbff_filename(path.name)]
            genome_matches = [path for path in files if is_probable_genome_filename(provider, path.name)]
        elif provider == "phytozome":
            cds_matches = [
                path
                for path in files
                if ("cds_" in path.name.lower() or ".cds." in path.name.lower()) and is_fasta_filename(path.name)
            ]
            gff_matches = [path for path in files if "gene.gff3" in path.name.lower() and is_gff_filename(path.name)]
            gbff_matches = [path for path in files if is_gbff_filename(path.name)]
            genome_matches = [path for path in files if is_probable_genome_filename(provider, path.name)]
        elif provider == "ncbi":
            cds_matches = [
                path
                for path in files
                if "cds_from_genomic" in path.name.lower() and is_fasta_filename(path.name)
            ]
            gff_matches = [
                path
                for path in files
                if "genomic.gff" in path.name.lower() and is_gff_filename(path.name)
            ]
            gbff_matches = [
                path
                for path in files
                if "genomic.gbff" in path.name.lower() and is_gbff_filename(path.name)
            ]
            genome_matches = [path for path in files if is_probable_genome_filename(provider, path.name)]
        else:
            cds_matches = [
                path
                for path in files
                if is_probable_cds_filename(provider, path.name)
            ]
            if len(cds_matches) == 0:
                cds_matches = [
                    path
                    for path in files
                    if is_fasta_filename(path.name) and not is_probable_genome_filename(provider, path.name)
                ]
            gff_matches = [path for path in files if is_gff_filename(path.name)]
            gbff_matches = [path for path in files if is_gbff_filename(path.name)]
            genome_matches = [path for path in files if is_probable_genome_filename(provider, path.name)]

        cds_path = pick_single_file(cds_matches, provider, species_key, "CDS", warnings)
        gff_path = pick_single_file(gff_matches, provider, species_key, "GFF", warnings)
        gbff_path = pick_single_file(gbff_matches, provider, species_key, "GBFF", warnings)
        genome_path = pick_single_file(genome_matches, provider, species_key, "genome", warnings)
        missing_label = missing_annotation_label(cds_path, gff_path, gbff_path, genome_path)
        if missing_label != "":
            errors.append(
                "[{}] {}: missing {}".format(
                    provider, species_key, missing_label
                )
            )
            continue
        source_id = species_key
        if provider == "coge":
            source_id = infer_coge_genome_id_from_files(species_key, files, warnings)
            if source_id == "":
                errors.append(
                    "[coge] {}: genome_id (gid) was not detected from file names. "
                    "Include 'gid<NUM>' in at least one input filename.".format(species_key)
                )
                continue
        if provider == "gwh":
            source_id = infer_gwh_accession_from_files(species_key, files, warnings)
            if source_id == "":
                errors.append(
                    "[gwh] {}: GWH accession was not detected from file names. "
                    "Include 'GWH...' in at least one input filename.".format(species_key)
                )
                continue
        rows.append(
            {
                "provider": provider,
                "id": source_id,
                "species_key": species_key,
                "cds_url": cds_path.resolve().as_uri() if cds_path is not None else "",
                "gff_url": gff_path.resolve().as_uri() if gff_path is not None else "",
                "gbff_url": gbff_path.resolve().as_uri() if gbff_path is not None else "",
                "genome_url": genome_path.resolve().as_uri() if genome_path is not None else "",
                "cds_filename": cds_path.name if cds_path is not None else "",
                "gff_filename": gff_path.name if gff_path is not None else "",
                "gbff_filename": gbff_path.name if gbff_path is not None else "",
                "genome_filename": genome_path.name if genome_path is not None else "",
                FERNBASE_CONFIDENCE_MODE_FIELD: infer_fernbase_confidence_mode_from_files(files, cds_path, gff_path)
                if provider == "fernbase"
                else "",
            }
        )
    return rows, warnings, errors


def discover(provider, input_root, allow_missing=False):
    if provider in LEGACY_NCBI_PROVIDER_ALIASES:
        provider = "ncbi"
    if provider == "ncbi":
        warnings = []
        errors = []
        rows = []
        seen_ids = set()
        existing_dirs = []
        for rel_dir in NCBI_MERGED_INPUT_RELATIVE_DIRS:
            input_dir = input_root / rel_dir
            if not input_dir.exists() or not input_dir.is_dir():
                continue
            existing_dirs.append(input_dir)
            sub_rows, sub_warnings, sub_errors = discover_species_dir_based("ncbi", input_dir)
            warnings.extend(sub_warnings)
            errors.extend(sub_errors)
            for row in sub_rows:
                source_id = str(row.get("id", "") or "").strip()
                if source_id != "" and source_id in seen_ids:
                    warnings.append(
                        "[ncbi] duplicate accession '{}' across NCBI_Genome/NCBI_RefSeq/NCBI_GenBank. Keeping first occurrence.".format(
                            source_id
                        )
                    )
                    continue
                if source_id != "":
                    seen_ids.add(source_id)
                rows.append(row)
        if len(existing_dirs) == 0:
            message = "[ncbi] input directories not found: {}".format(
                ", ".join(str((input_root / rel_dir)) for rel_dir in NCBI_MERGED_INPUT_RELATIVE_DIRS)
            )
            if allow_missing:
                return [], [message], []
            return [], [], [message]
        return rows, warnings, errors

    input_dir = input_root / DEFAULT_INPUT_RELATIVE_DIRS[provider]
    if not input_dir.exists() or not input_dir.is_dir():
        message = "[{}] input directory not found: {}".format(provider, input_dir)
        if allow_missing:
            return [], [message], []
        return [], [], [message]
    if provider in ("ensembl", "ensemblplants"):
        return discover_ensembl_like(input_dir, provider)
    return discover_species_dir_based(provider, input_dir)


def write_manifest_tsv(rows: Iterable[Dict[str, str]], output_path: Path):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "wt", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=list(MANIFEST_FIELDNAMES),
            delimiter="\t",
            lineterminator="\n",
        )
        writer.writeheader()
        for row in rows:
            writer.writerow({key: str(row.get(key, "") or "") for key in MANIFEST_FIELDNAMES})


def write_manifest_xlsx(rows: List[Dict[str, str]], output_path: Path, id_options_snapshot=None, direct_catalog_entries=None):
    if (
        Workbook is None
        or get_column_letter is None
        or DefinedName is None
        or Font is None
        or DataValidation is None
    ):
        raise RuntimeError(
            "openpyxl is required to write .xlsx manifests. Install openpyxl or use --output *.tsv."
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "download_plan"
    sheet.append(list(MANIFEST_FIELDNAMES))
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    if Comment is not None:
        for cell in sheet[1]:
            fieldname = str(cell.value or "").strip()
            comment = build_header_comment(fieldname)
            if comment is not None:
                cell.comment = comment
    sheet.freeze_panes = "A2"

    for row in rows:
        sheet.append([str(row.get(key, "") or "") for key in MANIFEST_FIELDNAMES])

    list_sheet = workbook.create_sheet("_lists")
    for idx, provider in enumerate(PROVIDERS, start=1):
        list_sheet.cell(row=idx, column=1, value=provider)
    id_options_by_provider = build_provider_id_options(rows, snapshot_options=id_options_snapshot)
    for col_idx, provider in enumerate(PROVIDERS, start=2):
        values = list(id_options_by_provider.get(provider, [""]))
        for row_idx, value in enumerate(values, start=1):
            list_sheet.cell(row=row_idx, column=col_idx, value=value)
        col_letter = get_column_letter(col_idx)
        workbook.defined_names.add(
            DefinedName(
                name="id_opts_{}".format(provider),
                attr_text="'_lists'!${}$1:${}${}".format(col_letter, col_letter, len(values)),
            )
        )
    list_sheet.sheet_state = "hidden"

    direct_catalog_entries = list(direct_catalog_entries or [])
    if len(direct_catalog_entries) > 0:
        direct_sheet = workbook.create_sheet("_direct_catalog")
        direct_headers = ["choice", "id", "species"] + list(DIRECT_CATALOG_FIELDS)
        direct_sheet.append(direct_headers)
        catalog_column_by_field = {}
        for idx, field in enumerate(direct_headers, start=1):
            catalog_column_by_field[field] = get_column_letter(idx)
        for entry in direct_catalog_entries:
            direct_sheet.append(
                [
                    build_direct_catalog_choice(entry),
                    str(entry.get("id", "") or "").strip(),
                    str(entry.get("species", "") or "").strip(),
                ]
                + [str(entry.get(field, "") or "").strip() for field in DIRECT_CATALOG_FIELDS]
            )
        direct_sheet.sheet_state = "hidden"

        sheet_column_by_field = {}
        for idx, field in enumerate(MANIFEST_FIELDNAMES, start=1):
            sheet_column_by_field[field] = idx
        catalog_row_end = len(direct_catalog_entries) + 1
        for row_idx in range(2, 5001):
            for field in DIRECT_CATALOG_FIELDS:
                cell = sheet.cell(row=row_idx, column=sheet_column_by_field[field])
                if str(cell.value or "").strip() != "":
                    continue
                cell.value = direct_catalog_formula(
                    catalog_column_by_field[field],
                    catalog_row_end,
                    row_idx,
                )

    provider_validation = DataValidation(
        type="list",
        formula1="=_lists!$A$1:$A${}".format(len(PROVIDERS)),
        allow_blank=False,
    )
    sheet.add_data_validation(provider_validation)
    provider_validation.add("A2:A5000")

    id_validation = DataValidation(
        type="list",
        formula1='=INDIRECT("id_opts_"&$A2)',
        allow_blank=False,
        errorStyle="information",
        errorTitle="ID candidates",
        error=(
            "This drop-down changes by provider. "
            "For provider=ncbi, five model-organism IDs are shown as examples. "
            "You can still type any other ID value."
        ),
    )
    id_validation.showErrorMessage = False
    sheet.add_data_validation(id_validation)
    id_validation.add("B2:B5000")

    fernbase_mode_col = get_column_letter(MANIFEST_FIELDNAMES.index(FERNBASE_CONFIDENCE_MODE_FIELD) + 1)
    fernbase_mode_validation = DataValidation(
        type="list",
        formula1='"{},{}"'.format(
            FERNBASE_CONFIDENCE_MODE_HIGH_ONLY,
            FERNBASE_CONFIDENCE_MODE_HIGH_LOW_COMBINED,
        ),
        allow_blank=True,
    )
    sheet.add_data_validation(fernbase_mode_validation)
    fernbase_mode_validation.add("{}2:{}5000".format(fernbase_mode_col, fernbase_mode_col))

    workbook.save(output_path)


def write_manifest(rows: List[Dict[str, str]], output_path: Path, id_options_snapshot=None, direct_catalog_entries=None):
    suffix = output_path.suffix.lower()
    if suffix == ".xlsx":
        write_manifest_xlsx(
            rows,
            output_path,
            id_options_snapshot=id_options_snapshot,
            direct_catalog_entries=direct_catalog_entries,
        )
        return
    write_manifest_tsv(rows, output_path)


def main():
    parser = build_arg_parser()
    args = parser.parse_args()
    input_root = Path(args.input_dir).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    id_options_snapshot = {}
    direct_catalog_entries = []
    if str(args.id_options_snapshot or "").strip() != "":
        snapshot_path = Path(args.id_options_snapshot).expanduser().resolve()
        if not snapshot_path.exists():
            sys.stderr.write("Warning: id options snapshot was not found and will be ignored: {}\n".format(snapshot_path))
        else:
            try:
                id_options_snapshot = load_id_options_snapshot(snapshot_path)
                direct_catalog_entries = load_direct_catalog_snapshot(snapshot_path)
            except Exception as exc:
                sys.stderr.write("Error: failed to read id options snapshot '{}': {}\n".format(snapshot_path, exc))
                return 1
    direct_catalog_text = str(args.direct_catalog_manifest or "").strip()
    if direct_catalog_text != "":
        direct_catalog_path = Path(direct_catalog_text).expanduser().resolve()
        if not direct_catalog_path.exists():
            sys.stderr.write("Error: direct catalog manifest was not found: {}\n".format(direct_catalog_path))
            return 1
        try:
            manifest_rows = load_manifest_rows(direct_catalog_path)
            direct_catalog_entries = build_direct_catalog_from_manifest_rows(manifest_rows)
        except Exception as exc:
            sys.stderr.write("Error: failed to read direct catalog manifest '{}': {}\n".format(direct_catalog_path, exc))
            return 1
        if len(direct_catalog_entries) > 0:
            id_options_snapshot = dict(id_options_snapshot)
            id_options_snapshot["direct"] = [
                (str(entry.get("id", "") or "").strip(), str(entry.get("species", "") or "").strip())
                for entry in direct_catalog_entries
            ]

    requested_provider = str(args.provider or "").strip().lower()
    if requested_provider in LEGACY_NCBI_PROVIDER_ALIASES:
        requested_provider = "ncbi"
    providers = PROVIDERS if requested_provider == "all" else (requested_provider,)
    all_rows = []
    all_warnings = []
    all_errors = []
    allow_missing = requested_provider == "all"
    for provider in providers:
        rows, warnings, errors = discover(provider, input_root, allow_missing=allow_missing)
        all_rows.extend(rows)
        all_warnings.extend(warnings)
        all_errors.extend(errors)

    for warning in all_warnings:
        sys.stderr.write("Warning: {}\n".format(warning))
    for error in all_errors:
        sys.stderr.write("Error: {}\n".format(error))

    provider_order = {provider: idx for idx, provider in enumerate(PROVIDERS)}
    all_rows = sorted(
        all_rows,
        key=lambda x: (
            provider_order.get(str(x.get("provider", "")), len(PROVIDERS)),
            str(x.get("species_key", "")),
        ),
    )
    write_manifest(
        all_rows,
        output_path,
        id_options_snapshot=id_options_snapshot,
        direct_catalog_entries=direct_catalog_entries,
    )
    print("Manifest written: {} (rows={})".format(output_path, len(all_rows)))

    if args.strict and len(all_errors) > 0:
        return 1
    if len(all_rows) == 0:
        if output_path.suffix.lower() == ".xlsx" and (len(direct_catalog_entries) > 0 or len(id_options_snapshot) > 0):
            return 0
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
