#!/usr/bin/env python3

import argparse
import csv
from datetime import datetime, timezone
import json
from pathlib import Path
import re
import subprocess
import sys
from urllib.parse import quote, urljoin, urlparse
from urllib.request import Request, urlopen

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


PROVIDERS = (
    "ensembl",
    "ensemblplants",
    "ncbi",
    "coge",
    "cngb",
    "gwh",
    "plantaedb",
    "flybase",
    "wormbase",
    "vectorbase",
    "fernbase",
    "veupathdb",
    "dictybase",
    "insectbase",
    "direct",
    "local",
)

FETCH_PROVIDERS = ("ensembl", "ensemblplants", "flybase", "wormbase", "vectorbase", "fernbase", "veupathdb", "insectbase", "local")

DEFAULT_INPUT_RELATIVE_DIRS = {
    "local": Path("Local") / "species_wise_original",
}

MANIFEST_FIELDNAMES = (
    "provider",
    "id",
    "species_key",
    "cds_url",
    "gff_url",
    "gbff_url",
    "genome_url",
    "cds_archive_member",
    "gff_archive_member",
    "gbff_archive_member",
    "genome_archive_member",
    "cds_filename",
    "gff_filename",
    "gbff_filename",
    "genome_filename",
    "cds_url_template",
    "gff_url_template",
    "genome_url_template",
    "local_cds_path",
    "local_gff_path",
    "local_gbff_path",
    "local_genome_path",
)
DIRECT_CATALOG_FIELDS = tuple(field for field in MANIFEST_FIELDNAMES if field not in ("provider", "id"))

ID_EXAMPLES_BY_PROVIDER = {
    "ensembl": (("homo_sapiens", "Homo sapiens"), ("mus_musculus", "Mus musculus")),
    "ensemblplants": (("Ostreococcus_lucimarinus", "Ostreococcus lucimarinus"),),
    "ncbi": (
        ("GCF_000001405.40", "Homo sapiens"),
        ("GCA_000001635.9", "Mus musculus"),
        ("GCF_049306965.1", "Danio rerio"),
        ("GCA_000001215.4", "Drosophila melanogaster"),
        ("GCF_000002985.6", "Caenorhabditis elegans"),
    ),
    "coge": (
        ("24739", "Arabidopsis thaliana"),
        ("29177", "Oryza sativa"),
        ("42091", "Zea mays"),
        ("78085", "Drosophila melanogaster"),
        ("72417", "Caenorhabditis elegans"),
    ),
    "cngb": (
        ("CNA0012345", "Homo sapiens"),
        ("GCF_000001405.40", "Homo sapiens"),
        ("GCA_000001635.9", "Mus musculus"),
        ("GCF_049306965.1", "Danio rerio"),
        ("GCA_000001215.4", "Drosophila melanogaster"),
    ),
    "gwh": (
        ("GWHIGRM00000000.1", "Medicago sativa"),
        ("GWHCBHY00000000", "Allium sativum"),
    ),
    "plantaedb": (
        (
            "https://plantaedb.com/taxa/phylum/angiosperms/order/asterales/family/asteraceae/subfamily/asteroideae/tribe/astereae/subtribe/conyzinae/genus/erigeron/species/erigeron-breviscapus",
            "Erigeron breviscapus (PlantaeDB page)",
        ),
        (
            "https://plantaedb.com/taxa/phylum/angiosperms/order/ranunculales/family/berberidaceae/genus/berberis/species/berberis-thunbergii",
            "Berberis thunbergii (PlantaeDB page)",
        ),
    ),
    "flybase": (("dmel_r6.61", "Drosophila melanogaster"),),
    "wormbase": (("celegans_prjna13758_ws290", "Caenorhabditis elegans"),),
    "vectorbase": (("anopheles_gambiae_pest", "Anopheles gambiae"),),
    "fernbase": (("Azolla_filiculoides", "Azolla filiculoides"), ("Salvinia_cucullata_v2", "Salvinia cucullata v2")),
    "veupathdb": (("EnuttalliP19", "Entamoeba nuttalli"),),
    "dictybase": (("Dictyostelium_discoideum", "Dictyostelium discoideum"),),
    "insectbase": (("IBG_00001", "Abrostola tripartita"),),
    "direct": (("direct_example_species", "Direct URL manifest row"),),
    "local": (("/absolute/path/to/local/species_dir", "Local species directory"),),
}

DEFAULT_TIMEOUT_SECONDS = 45.0


def species_label_from_species_key(species_key):
    text = str(species_key or "").strip()
    if text == "":
        return ""
    parts = text.replace("_", " ").split()
    normalized = []
    for idx, part in enumerate(parts):
        if re.fullmatch(r"[a-z]+", part):
            if idx == 0:
                normalized.append(part.capitalize())
            else:
                normalized.append(part.lower())
        else:
            normalized.append(part)
    return " ".join(normalized)


def dedupe_options(options):
    out = []
    seen = set()
    for source_id, species_label in options:
        sid = str(source_id or "").strip()
        label = str(species_label or "").strip()
        if sid == "":
            continue
        key = (sid, label)
        if key in seen:
            continue
        seen.add(key)
        out.append((sid, label))
    return out


def parse_snapshot_entries(raw_entries):
    entries = raw_entries
    if isinstance(entries, dict):
        entries = [entries]
    if not isinstance(entries, list):
        return []
    out = []
    for entry in entries:
        source_id = ""
        species_label = ""
        if isinstance(entry, str):
            source_id = entry
        elif isinstance(entry, dict):
            source_id = str(entry.get("id") or entry.get("source_id") or "").strip()
            species_label = str(entry.get("species") or entry.get("species_label") or entry.get("label") or "").strip()
        elif isinstance(entry, (tuple, list)) and len(entry) > 0:
            source_id = str(entry[0] or "").strip()
            if len(entry) > 1:
                species_label = str(entry[1] or "").strip()
        if source_id == "":
            continue
        out.append((source_id, species_label))
    return dedupe_options(out)


def parse_direct_catalog_entries(raw_entries):
    entries = raw_entries
    if isinstance(entries, dict):
        entries = [entries]
    if not isinstance(entries, list):
        return []
    out = []
    seen = set()
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        source_id = str(entry.get("id") or entry.get("source_id") or "").strip()
        if source_id == "":
            continue
        species_label = str(entry.get("species") or entry.get("species_label") or entry.get("label") or "").strip()
        species_key = str(entry.get("species_key") or "").strip()
        key = (source_id, species_key, species_label)
        if key in seen:
            continue
        seen.add(key)
        record = {"id": source_id, "species": species_label}
        for field in DIRECT_CATALOG_FIELDS:
            record[field] = str(entry.get(field) or "").strip()
        out.append(record)
    return out


def load_snapshot_payload(path):
    with open(path, "rt", encoding="utf-8") as handle:
        payload = json.load(handle)
    providers_blob = payload.get("providers", payload)
    if not isinstance(providers_blob, dict):
        raise ValueError("snapshot must contain object key 'providers'")
    return providers_blob


def load_snapshot_provider_options(path):
    providers_blob = load_snapshot_payload(path)
    out = {}
    for provider in PROVIDERS:
        out[provider] = parse_snapshot_entries(providers_blob.get(provider, []))
    return out


def load_snapshot_direct_catalog(path):
    providers_blob = load_snapshot_payload(path)
    return parse_direct_catalog_entries(providers_blob.get("direct", []))


def load_manifest_rows(path):
    suffix = str(path.suffix or "").lower()
    if suffix == ".xlsx":
        if load_workbook is None:
            raise RuntimeError("openpyxl is required to read .xlsx direct catalog manifests")
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            sheet = workbook.active
            values = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()
        if len(values) == 0:
            return []
        headers = [str(value or "").strip() for value in values[0]]
        rows = []
        for row_values in values[1:]:
            if row_values is None:
                continue
            record = {}
            nonempty = False
            for idx, field in enumerate(headers):
                if field == "":
                    continue
                value = row_values[idx] if idx < len(row_values) else ""
                text = str(value or "").strip()
                if text != "":
                    nonempty = True
                record[field] = text
            if nonempty:
                rows.append(record)
        return rows
    delimiter = "\t" if suffix == ".tsv" else ","
    with open(path, "rt", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter=delimiter))


def build_direct_catalog_from_manifest_rows(rows):
    out = []
    seen = set()
    for row in rows:
        provider = str(row.get("provider", "") or "").strip().lower()
        if provider != "direct":
            continue
        source_id = str(row.get("id", "") or "").strip()
        if source_id == "":
            continue
        species_key = str(row.get("species_key", "") or "").strip()
        species_label = species_label_from_species_key(species_key)
        key = (source_id, species_key, species_label)
        if key in seen:
            continue
        seen.add(key)
        record = {"id": source_id, "species": species_label}
        for field in DIRECT_CATALOG_FIELDS:
            record[field] = str(row.get(field, "") or "").strip()
        out.append(record)
    return sorted(out, key=lambda record: (str(record.get("id", "")).lower(), str(record.get("species_key", "")).lower()))


def fetch_text(url, timeout):
    urllib_exc = None
    try:
        req = Request(url, headers={"User-Agent": "genegalleon-id-options/1"})
        with urlopen(req, timeout=timeout) as response:
            payload = response.read()
        if len(payload) > 0:
            try:
                return payload.decode("utf-8")
            except UnicodeDecodeError:
                return payload.decode("utf-8", errors="replace")
    except Exception as exc:
        urllib_exc = exc

    curl_cmd = [
        "curl",
        "-fsSL",
        "--max-time",
        str(max(1, int(round(float(timeout))))),
        str(url),
    ]
    curl_proc = subprocess.run(curl_cmd, capture_output=True, text=True, check=False)
    if curl_proc.returncode == 0 and curl_proc.stdout != "":
        return curl_proc.stdout

    if urllib_exc is not None:
        raise urllib_exc
    raise ValueError("empty response from {}".format(url))


def fetch_json(url, timeout):
    return json.loads(fetch_text(url, timeout))


def parse_links(html_text, base_url):
    links = []
    for href in re.findall(r"""href\s*=\s*["']([^"']+)["']""", html_text, flags=re.IGNORECASE):
        candidate = href.strip()
        if candidate == "" or candidate.startswith("?") or candidate.startswith("#"):
            continue
        links.append(urljoin(base_url, candidate))
    return links


def parse_dirname_from_index_link(url):
    path = urlparse(url).path
    if path.endswith("/"):
        name = path.rstrip("/").split("/")[-1]
        return name
    if path.endswith("/index.html"):
        parent = path[: -len("/index.html")]
        return parent.rstrip("/").split("/")[-1]
    return ""


def fetch_ensembl_options(timeout):
    data = fetch_json("https://rest.ensembl.org/info/species?content-type=application/json", timeout)
    out = []
    for item in data.get("species", []):
        source_id = str(item.get("name") or "").strip()
        if source_id == "":
            continue
        species_label = str(item.get("display_name") or "").strip()
        if species_label == "":
            species_label = species_label_from_species_key(source_id)
        out.append((source_id, species_label))
    out = dedupe_options(out)
    return sorted(out, key=lambda x: x[0].lower())


def fetch_ensemblplants_options(timeout):
    index_url = "https://ftp.ensemblgenomes.ebi.ac.uk/pub/plants/current/fasta/"
    html_text = fetch_text(index_url, timeout)
    out = []
    for link in parse_links(html_text, index_url):
        source_id = parse_dirname_from_index_link(link)
        if source_id in ("", "current", "fasta", "pub", "plants"):
            continue
        if not re.fullmatch(r"[A-Za-z0-9_.-]+", source_id):
            continue
        species_label = species_label_from_species_key(source_id)
        out.append((source_id, species_label))
    out = dedupe_options(out)
    return sorted(out, key=lambda x: x[0].lower())


def fetch_flybase_options(timeout):
    base_url = "https://s3ftp.flybase.org/genomes/"
    html_text = fetch_text(base_url, timeout)
    species_dirs = set()
    for link in parse_links(html_text, base_url):
        path = urlparse(link).path
        match = re.search(r"/genomes/([^/]+)/index\.html$", path)
        if match is None:
            continue
        species_dir = match.group(1).strip()
        if species_dir == "" or not re.fullmatch(r"[A-Za-z0-9_.-]+", species_dir):
            continue
        species_dirs.add(species_dir)

    out = []
    for species_dir in sorted(species_dirs):
        fasta_index = "{}{}/current/fasta/index.html".format(base_url, species_dir)
        try:
            fasta_html = fetch_text(fasta_index, timeout)
        except Exception:
            continue
        candidate_ids = []
        for link in parse_links(fasta_html, fasta_index):
            filename = urlparse(link).path.split("/")[-1]
            match = re.search(r"([a-z0-9]+)-all-([a-z0-9_]+)-r([0-9][0-9.]*)[.]fasta(?:[.]gz)?$", filename, flags=re.IGNORECASE)
            if match is None:
                continue
            abbrev = match.group(1).lower()
            label_token = match.group(2).lower()
            release = match.group(3)
            priority = 2
            if label_token == "cds":
                priority = 0
            elif "transcript" in label_token:
                priority = 1
            candidate_ids.append((priority, "{}_r{}".format(abbrev, release)))

        resolved_id = ""
        if len(candidate_ids) > 0:
            candidate_ids = sorted(candidate_ids, key=lambda x: (x[0], x[1]))
            resolved_id = candidate_ids[0][1]
        if resolved_id == "":
            continue
        species_label = species_dir.replace("_", " ")
        out.append((resolved_id, species_label))
    out = dedupe_options(out)
    return sorted(out, key=lambda x: x[0].lower())


def fetch_wormbase_options(timeout):
    base_url = "https://ftp.ebi.ac.uk/pub/databases/wormbase/parasite/releases/current/species/"
    html_text = fetch_text(base_url, timeout)
    species_dirs = set()
    for link in parse_links(html_text, base_url):
        path = urlparse(link).path
        match = re.search(r"/species/([^/]+)/?$", path)
        if match is None:
            continue
        species_dir = match.group(1).strip()
        if species_dir == "" or not re.fullmatch(r"[a-z0-9_]+", species_dir):
            continue
        species_dirs.add(species_dir)

    out = []
    for species_dir in sorted(species_dirs):
        species_url = "{}{}/".format(base_url, species_dir)
        try:
            species_html = fetch_text(species_url, timeout)
        except Exception:
            continue
        projects = set()
        for link in parse_links(species_html, species_url):
            path = urlparse(link).path
            match = re.search(r"/(PRJ[A-Z0-9]+)/?$", path, flags=re.IGNORECASE)
            if match is None:
                continue
            projects.add(match.group(1).upper())
        if len(projects) == 0:
            continue
        species_label = species_dir.replace("_", " ")
        for project in sorted(projects):
            source_id = "{}_{}".format(species_dir, project.lower())
            out.append((source_id, species_label))
    out = dedupe_options(out)
    return sorted(out, key=lambda x: x[0].lower())


def vectorbase_species_label_from_id(source_id):
    match = re.match(r"^([A-Z])([a-z]+)", str(source_id or ""))
    if match is not None:
        return "{}. {}".format(match.group(1), match.group(2).lower())
    spaced = re.sub(r"([a-z0-9])([A-Z])", r"\1 \2", str(source_id or ""))
    return species_label_from_species_key(spaced.replace(" ", "_"))


def fetch_vectorbase_options(timeout):
    root_url = "https://vectorbase.org/common/downloads/"
    root_html = fetch_text(root_url, timeout)
    release_numbers = []
    for link in parse_links(root_html, root_url):
        path = urlparse(link).path
        match = re.search(r"/release-([0-9]+)/?$", path)
        if match is None:
            continue
        release_numbers.append(int(match.group(1)))
    if len(release_numbers) == 0:
        raise ValueError("vectorbase release directories were not found")
    latest_release = max(release_numbers)
    release_url = "{}release-{}/".format(root_url, latest_release)
    release_html = fetch_text(release_url, timeout)

    out = []
    for link in parse_links(release_html, release_url):
        path = urlparse(link).path
        match = re.search(r"/release-{}/([^/]+)/?$".format(latest_release), path)
        if match is None:
            continue
        source_id = match.group(1).strip()
        if source_id == "" or not re.fullmatch(r"[A-Za-z0-9_.-]+", source_id):
            continue
        if source_id.lower().startswith("release-"):
            continue
        species_label = vectorbase_species_label_from_id(source_id)
        out.append((source_id, species_label))
    out = dedupe_options(out)
    return sorted(out, key=lambda x: x[0].lower())


def fetch_fernbase_options(timeout):
    root_url = "https://fernbase.org/ftp/"
    html_text = fetch_text(root_url, timeout)
    out = []
    for link in parse_links(html_text, root_url):
        path = urlparse(link).path
        match = re.search(r"/ftp/([^/]+)/?$", path)
        if match is None:
            continue
        source_id = match.group(1).strip()
        if source_id == "":
            continue
        if not re.fullmatch(r"[A-Z][A-Za-z0-9.-]*_[A-Za-z0-9_.-]+", source_id):
            continue
        species_label = species_label_from_species_key(source_id)
        out.append((source_id, species_label))
    out = dedupe_options(out)
    return sorted(out, key=lambda x: x[0].lower())


def fetch_veupathdb_options(timeout):
    config = {
        "attributes": ["primary_key", "species", "URLGenomeFasta"],
        "tables": [],
        "attributeFormat": "text",
    }
    base_url = "https://veupathdb.org/veupathdb/service"
    query = json.dumps(config, separators=(",", ":"))
    service_url = (
        "{}/record-types/organism/searches/GenomeDataTypes/reports/standard?reportConfig={}".format(
            base_url,
            quote(query, safe=""),
        )
    )
    payload = fetch_json(service_url, timeout)
    out = []
    for record in payload.get("records", []):
        attributes = record.get("attributes", {})
        genome_url = str(attributes.get("URLGenomeFasta", "") or "").strip()
        if genome_url == "":
            continue
        parts = [part for part in urlparse(genome_url).path.split("/") if part]
        source_id = ""
        for idx, token in enumerate(parts[:-1]):
            if token == "Current_Release" and idx + 1 < len(parts):
                source_id = parts[idx + 1].strip()
                break
        if source_id == "":
            continue
        species_label = str(attributes.get("species", "") or "").strip()
        out.append((source_id, species_label))
    out = dedupe_options(out)
    return sorted(out, key=lambda x: x[0].lower())


def fetch_insectbase_options(timeout):
    next_url = "https://www.insect-genome.com/api/genome/genomes/?page=1&page_size=5000"
    out = []
    seen = set()
    while next_url:
        payload = fetch_json(next_url, timeout)
        for record in payload.get("results", []):
            source_id = str(record.get("ibg_id", "") or "").strip()
            species_label = str(record.get("species", "") or "").strip()
            if source_id == "" or source_id in seen:
                continue
            seen.add(source_id)
            out.append((source_id, species_label))
        next_url = str(payload.get("next", "") or "").strip()
    out = dedupe_options(out)
    return sorted(out, key=lambda x: x[0].lower())


def fetch_local_options(input_root):
    if input_root is None:
        return []
    local_dir = input_root / DEFAULT_INPUT_RELATIVE_DIRS["local"]
    if not local_dir.exists() or not local_dir.is_dir():
        return []
    out = []
    for species_dir in sorted(local_dir.iterdir()):
        if not species_dir.is_dir():
            continue
        source_id = species_dir.name
        species_label = species_label_from_species_key(source_id)
        out.append((source_id, species_label))
    out = dedupe_options(out)
    return sorted(out, key=lambda x: x[0].lower())


def fetch_provider_options(provider, timeout, input_root):
    if provider == "ensembl":
        return fetch_ensembl_options(timeout)
    if provider == "ensemblplants":
        return fetch_ensemblplants_options(timeout)
    if provider == "flybase":
        return fetch_flybase_options(timeout)
    if provider == "wormbase":
        return fetch_wormbase_options(timeout)
    if provider == "vectorbase":
        return fetch_vectorbase_options(timeout)
    if provider == "fernbase":
        return fetch_fernbase_options(timeout)
    if provider == "veupathdb":
        return fetch_veupathdb_options(timeout)
    if provider == "insectbase":
        return fetch_insectbase_options(timeout)
    if provider == "local":
        return fetch_local_options(input_root)
    raise ValueError("unexpected fetch provider: {}".format(provider))


def example_options(provider):
    return [(source_id, species_label) for source_id, species_label in ID_EXAMPLES_BY_PROVIDER.get(provider, ())]


def build_snapshot(timeout, input_root, fallback_options, fallback_direct_catalog, direct_catalog_entries):
    warnings = []
    providers = {}

    for provider in PROVIDERS:
        if provider == "direct":
            resolved_catalog = list(direct_catalog_entries)
            if len(resolved_catalog) == 0:
                resolved_catalog = list(fallback_direct_catalog)
            if len(resolved_catalog) == 0:
                resolved_catalog = [
                    {"id": source_id, "species": species_label}
                    for source_id, species_label in example_options(provider)
                ]
            providers[provider] = resolved_catalog
            continue

        resolved = []
        if provider in FETCH_PROVIDERS:
            try:
                resolved = fetch_provider_options(provider, timeout=timeout, input_root=input_root)
            except Exception as exc:
                warnings.append("[{}] fetch failed: {}".format(provider, exc))
                resolved = []
            if len(resolved) == 0:
                fallback_values = list(fallback_options.get(provider, []))
                if len(fallback_values) > 0:
                    warnings.append("[{}] using fallback snapshot entries ({})".format(provider, len(fallback_values)))
                    resolved = fallback_values
                else:
                    resolved = example_options(provider)
                    warnings.append("[{}] using built-in example entries ({})".format(provider, len(resolved)))
        else:
            fallback_values = list(fallback_options.get(provider, []))
            if len(fallback_values) > 0:
                resolved = fallback_values
            else:
                resolved = example_options(provider)

        resolved = dedupe_options(resolved)
        resolved = sorted(resolved, key=lambda x: x[0].lower())
        providers[provider] = [{"id": source_id, "species": species_label} for source_id, species_label in resolved]

    snapshot = {
        "schema_version": 1,
        "generated_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "providers": providers,
    }
    return snapshot, warnings


def build_arg_parser():
    parser = argparse.ArgumentParser(
        description=(
            "Build provider-specific id_options_snapshot.json used by build_download_manifest.py "
            "for XLSX drop-down ID candidates."
        )
    )
    parser.add_argument(
        "--output",
        required=True,
        help="Output JSON path.",
    )
    parser.add_argument(
        "--input-dir",
        default="",
        help=(
            "Optional provider-root input directory. "
            "Used to discover provider=local IDs from Local/species_wise_original."
        ),
    )
    parser.add_argument(
        "--fallback-snapshot",
        default="",
        help="Optional previous id_options_snapshot.json used as per-provider fallback.",
    )
    parser.add_argument(
        "--direct-catalog-manifest",
        default="",
        help=(
            "Optional TSV/CSV/XLSX manifest containing provider=direct rows. "
            "When set, direct entries in the snapshot preserve direct URL and filename fields."
        ),
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=DEFAULT_TIMEOUT_SECONDS,
        help="Timeout seconds per HTTP request (default: 45).",
    )
    return parser


def main():
    parser = build_arg_parser()
    args = parser.parse_args()

    output_path = Path(args.output).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    input_root = None
    if str(args.input_dir or "").strip() != "":
        input_root = Path(args.input_dir).expanduser().resolve()

    fallback_options = {}
    fallback_direct_catalog = []
    fallback_text = str(args.fallback_snapshot or "").strip()
    if fallback_text != "":
        fallback_path = Path(fallback_text).expanduser().resolve()
        if fallback_path.exists():
            try:
                fallback_options = load_snapshot_provider_options(fallback_path)
                fallback_direct_catalog = load_snapshot_direct_catalog(fallback_path)
            except Exception as exc:
                sys.stderr.write("Warning: failed to read fallback snapshot '{}': {}\n".format(fallback_path, exc))
        else:
            sys.stderr.write("Warning: fallback snapshot not found: {}\n".format(fallback_path))

    direct_catalog_entries = []
    direct_catalog_text = str(args.direct_catalog_manifest or "").strip()
    if direct_catalog_text != "":
        direct_catalog_path = Path(direct_catalog_text).expanduser().resolve()
        if not direct_catalog_path.exists():
            sys.stderr.write("Warning: direct catalog manifest not found: {}\n".format(direct_catalog_path))
        else:
            try:
                manifest_rows = load_manifest_rows(direct_catalog_path)
                direct_catalog_entries = build_direct_catalog_from_manifest_rows(manifest_rows)
            except Exception as exc:
                sys.stderr.write("Warning: failed to read direct catalog manifest '{}': {}\n".format(direct_catalog_path, exc))

    snapshot, warnings = build_snapshot(
        timeout=float(args.timeout),
        input_root=input_root,
        fallback_options=fallback_options,
        fallback_direct_catalog=fallback_direct_catalog,
        direct_catalog_entries=direct_catalog_entries,
    )
    for warning in warnings:
        sys.stderr.write("Warning: {}\n".format(warning))

    with open(output_path, "wt", encoding="utf-8") as handle:
        json.dump(snapshot, handle, indent=2, ensure_ascii=False)
        handle.write("\n")

    counts = []
    for provider in PROVIDERS:
        counts.append("{}={}".format(provider, len(snapshot.get("providers", {}).get(provider, []))))
    print("Snapshot written: {} ({})".format(output_path, ", ".join(counts)))
    return 0


if __name__ == "__main__":
    sys.exit(main())
