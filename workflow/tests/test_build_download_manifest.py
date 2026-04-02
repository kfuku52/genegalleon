from pathlib import Path
import csv
import json
import shutil
import subprocess
import sys
import zipfile

from openpyxl import load_workbook


SCRIPT_PATH = Path(__file__).resolve().parents[1] / "support" / "build_download_manifest.py"
SMALL_DATASET_ROOT = Path(__file__).resolve().parent / "data" / "small_gfe_dataset"


def run_script(*args):
    return subprocess.run(
        [sys.executable, str(SCRIPT_PATH), *args],
        capture_output=True,
        text=True,
        check=False,
    )


def read_manifest(path):
    if path.suffix.lower() == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()
        if len(rows) == 0:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        out = []
        for values in rows[1:]:
            if values is None:
                continue
            record = {}
            nonempty = False
            for idx, key in enumerate(headers):
                if key == "":
                    continue
                value = values[idx] if idx < len(values) else ""
                text = str(value or "").strip()
                if text != "":
                    nonempty = True
                record[key] = text
            if nonempty:
                out.append(record)
        return out
    with open(path, "rt", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def read_list_column_values(sheet, column):
    values = []
    row_idx = 1
    while True:
        value = sheet.cell(row=row_idx, column=column).value
        if value is None or str(value).strip() == "":
            break
        values.append(str(value).strip())
        row_idx += 1
    return values


def read_comment_vml(path):
    with zipfile.ZipFile(path) as archive:
        for name in archive.namelist():
            if name.endswith(".vml"):
                return archive.read(name).decode("utf-8")
    return ""


def test_build_download_manifest_from_small_fixture_all(tmp_path):
    out = tmp_path / "download_manifest.xlsx"
    completed = run_script(
        "--provider",
        "all",
        "--input-dir",
        str(SMALL_DATASET_ROOT),
        "--output",
        str(out),
    )
    assert completed.returncode == 0, completed.stderr + "\n" + completed.stdout
    assert out.exists()

    rows = read_manifest(out)
    assert len(rows) == 1
    providers = sorted([row["provider"] for row in rows])
    assert providers == ["ensemblplants"]

    for row in rows:
        assert row["cds_url"].startswith("file://")
        assert row["gff_url"].startswith("file://")
        assert "genome_url" in row
        assert row["cds_filename"] != ""
        assert row["gff_filename"] != ""
        assert "genome_filename" in row


def test_build_download_manifest_strict_allows_cds_only_inputs(tmp_path):
    dataset_copy = tmp_path / "small_dataset_copy"
    shutil.copytree(SMALL_DATASET_ROOT, dataset_copy)
    missing_gff = dataset_copy / "20230216_EnsemblPlants" / "original_files" / "Ostreococcus_lucimarinus.ASM9206v1.56.gff3"
    missing_gff.unlink()

    out = tmp_path / "download_manifest.xlsx"
    completed = run_script(
        "--provider",
        "ensemblplants",
        "--input-dir",
        str(dataset_copy),
        "--output",
        str(out),
        "--strict",
    )
    assert completed.returncode == 0, completed.stderr + "\n" + completed.stdout
    rows = read_manifest(out)
    assert len(rows) == 1
    assert rows[0]["cds_url"].startswith("file://")
    assert rows[0]["gff_url"] == ""
    assert rows[0]["genome_url"] == ""


def test_build_download_manifest_ncbi_like_provider_from_species_dir_fixture(tmp_path):
    input_root = tmp_path / "dataset"
    fixtures = (
        ("NCBI_Genome", "GCF_000001405.40", "GCF_000001405.40_GRCh38.p14"),
        ("NCBI_RefSeq", "GCF_000001405.40", "GCF_000001405.40_GRCh38.p14"),
        ("NCBI_GenBank", "GCA_000001405.29", "GCA_000001405.29_GRCh38.p14"),
    )
    for provider_dir, accession, stem in fixtures:
        species_dir = input_root / provider_dir / "species_wise_original" / accession
        species_dir.mkdir(parents=True, exist_ok=True)
        (species_dir / (stem + "_cds_from_genomic.fna.gz")).write_bytes(b"dummy")
        (species_dir / (stem + "_genomic.gff.gz")).write_bytes(b"dummy")
        (species_dir / (stem + "_genomic.fna.gz")).write_bytes(b"dummy")

    out = tmp_path / "download_manifest_ncbi.xlsx"
    completed = run_script(
        "--provider",
        "ncbi",
        "--input-dir",
        str(input_root),
        "--output",
        str(out),
    )
    assert completed.returncode == 0, completed.stderr + "\n" + completed.stdout
    rows = read_manifest(out)
    assert len(rows) == 2
    assert [row["provider"] for row in rows] == ["ncbi", "ncbi"]
    assert sorted([row["id"] for row in rows]) == ["GCA_000001405.29", "GCF_000001405.40"]
    for row in rows:
        assert row["species_key"] in ("GCA_000001405.29", "GCF_000001405.40")
        assert row["genome_url"].startswith("file://")
        assert row["genome_filename"].endswith("_genomic.fna.gz")


def test_build_download_manifest_ensembl_provider_from_flat_fixture(tmp_path):
    input_root = tmp_path / "dataset"
    ensembl_dir = input_root / "Ensembl" / "original_files"
    ensembl_dir.mkdir(parents=True)
    (ensembl_dir / "homo_sapiens.GRCh38.cds.all.fa.gz").write_bytes(b"dummy")
    (ensembl_dir / "homo_sapiens.GRCh38.112.gff3.gz").write_bytes(b"dummy")
    (ensembl_dir / "homo_sapiens.GRCh38.dna.primary_assembly.fa.gz").write_bytes(b"dummy")

    out = tmp_path / "download_manifest_ensembl.xlsx"
    completed = run_script(
        "--provider",
        "ensembl",
        "--input-dir",
        str(input_root),
        "--output",
        str(out),
    )
    assert completed.returncode == 0, completed.stderr + "\n" + completed.stdout
    rows = read_manifest(out)
    assert len(rows) == 1
    assert rows[0]["provider"] == "ensembl"
    assert rows[0]["id"] == "homo_sapiens"
    assert rows[0]["species_key"] == "homo_sapiens"
    assert rows[0]["cds_filename"] == "homo_sapiens.GRCh38.cds.all.fa.gz"
    assert rows[0]["gff_filename"] == "homo_sapiens.GRCh38.112.gff3.gz"
    assert rows[0]["genome_filename"] == "homo_sapiens.GRCh38.dna.primary_assembly.fa.gz"


def test_build_download_manifest_coge_cngb_and_gwh_provider_from_species_dir_fixture(tmp_path):
    input_root = tmp_path / "dataset"
    species_key = "Arabidopsis_thaliana"
    provider_dir_name = {"coge": "CoGe", "cngb": "CNGB", "gwh": "GWH"}

    for provider in ("coge", "cngb", "gwh"):
        species_dir = input_root / provider_dir_name[provider] / "species_wise_original" / species_key
        species_dir.mkdir(parents=True, exist_ok=True)
        if provider == "coge":
            cds_name = "Arabidopsis_thaliana.coge.gid24739.cds.fa"
            gff_name = "Arabidopsis_thaliana.gid24739.gff3"
            genome_name = "Arabidopsis_thaliana.coge.gid24739.genome.fa"
        elif provider == "gwh":
            cds_name = "Arabidopsis_thaliana.GWHABCD00000000.1.CDS.fasta.gz"
            gff_name = "Arabidopsis_thaliana.GWHABCD00000000.1.gff.gz"
            genome_name = "Arabidopsis_thaliana.GWHABCD00000000.1.genome.fasta.gz"
        else:
            cds_name = species_key + ".cds.fa"
            gff_name = species_key + ".gene.gff3"
            genome_name = species_key + ".genome.fa"
        (species_dir / cds_name).write_text(">AT1G01010_t1\nATGAA\n", encoding="utf-8")
        (species_dir / gff_name).write_text(
            "chr1\tsrc\tgene\t1\t9\t.\t+\t.\tID=AT1G01010\n",
            encoding="utf-8",
        )
        (species_dir / genome_name).write_text(">chr1\nATGCATGC\n", encoding="utf-8")

        out = tmp_path / ("download_manifest_{}.xlsx".format(provider))
        completed = run_script(
            "--provider",
            provider,
            "--input-dir",
            str(input_root),
            "--output",
            str(out),
        )
        assert completed.returncode == 0, completed.stderr + "\n" + completed.stdout
        rows = read_manifest(out)
        assert len(rows) == 1
        assert rows[0]["provider"] == provider
        expected_id = "24739" if provider == "coge" else ("GWHABCD00000000.1" if provider == "gwh" else species_key)
        assert rows[0]["id"] == expected_id
        assert rows[0]["species_key"] == species_key
        assert rows[0]["cds_filename"] == cds_name
        assert rows[0]["gff_filename"] == gff_name
        assert rows[0]["genome_filename"] == genome_name


def test_build_download_manifest_fernbase_provider_prefers_primary_annotation_files(tmp_path):
    input_root = tmp_path / "dataset"
    species_dir = input_root / "FernBase" / "species_wise_original" / "Azolla_filiculoides"
    species_dir.mkdir(parents=True, exist_ok=True)
    (species_dir / "Azolla_filiculoides.CDS.lowconfidence_v1.1.fasta").write_text(">g1.t1\nATG\n", encoding="utf-8")
    (species_dir / "Azolla_filiculoides.CDS.highconfidence_v1.1.fasta").write_text(">g1.t1\nATGAA\n", encoding="utf-8")
    (species_dir / "Azolla_filiculoides.transcript.highconfidence_v1.1.fasta").write_text(">g1.t1\nATGAAA\n", encoding="utf-8")
    (species_dir / "Azolla_filiculoides.gene_models.lowconfidence_v1.1.gff").write_text(
        "chr1\tsrc\tgene\t1\t3\t.\t+\t.\tID=g1\n",
        encoding="utf-8",
    )
    (species_dir / "Azolla_filiculoides.gene_models.highconfidence_v1.1.gff").write_text(
        "chr1\tsrc\tgene\t1\t5\t.\t+\t.\tID=g1\n",
        encoding="utf-8",
    )
    (species_dir / "Azolla_filiculoides.genome_v1.2.fasta").write_text(">chr1\nATGCATGC\n", encoding="utf-8")

    out = tmp_path / "download_manifest_fernbase.xlsx"
    completed = run_script(
        "--provider",
        "fernbase",
        "--input-dir",
        str(input_root),
        "--output",
        str(out),
    )
    assert completed.returncode == 0, completed.stderr + "\n" + completed.stdout
    rows = read_manifest(out)
    assert len(rows) == 1
    assert rows[0]["provider"] == "fernbase"
    assert rows[0]["id"] == "Azolla_filiculoides"
    assert rows[0]["species_key"] == "Azolla_filiculoides"
    assert rows[0]["cds_filename"] == "Azolla_filiculoides.CDS.highconfidence_v1.1.fasta"
    assert rows[0]["gff_filename"] == "Azolla_filiculoides.gene_models.highconfidence_v1.1.gff"
    assert rows[0]["genome_filename"] == "Azolla_filiculoides.genome_v1.2.fasta"
    assert rows[0]["fernbase_confidence_mode"] == "high-confidence only"


def test_build_download_manifest_allows_gff_and_genome_without_cds(tmp_path):
    input_root = tmp_path / "dataset"
    species_dir = input_root / "Direct" / "species_wise_original" / "Arabidopsis_thaliana"
    species_dir.mkdir(parents=True, exist_ok=True)
    (species_dir / "Arabidopsis_thaliana.gene.gff3").write_text(
        "chr1\tsrc\tgene\t1\t9\t.\t+\t.\tID=gene1\n",
        encoding="utf-8",
    )
    (species_dir / "Arabidopsis_thaliana.genome.fa").write_text(">chr1\nATGAAATTT\n", encoding="utf-8")

    out = tmp_path / "download_manifest_direct.xlsx"
    completed = run_script(
        "--provider",
        "direct",
        "--input-dir",
        str(input_root),
        "--output",
        str(out),
    )
    assert completed.returncode == 0, completed.stderr + "\n" + completed.stdout
    rows = read_manifest(out)
    assert len(rows) == 1
    assert rows[0]["provider"] == "direct"
    assert rows[0]["id"] == "Arabidopsis_thaliana"
    assert rows[0]["species_key"] == "Arabidopsis_thaliana"
    assert rows[0]["cds_url"] == ""
    assert rows[0]["cds_filename"] == ""
    assert rows[0]["gff_filename"] == "Arabidopsis_thaliana.gene.gff3"
    assert rows[0]["genome_filename"] == "Arabidopsis_thaliana.genome.fa"


def test_build_download_manifest_allows_cds_only(tmp_path):
    input_root = tmp_path / "dataset"
    species_dir = input_root / "Direct" / "species_wise_original" / "Croton_tiglium"
    species_dir.mkdir(parents=True, exist_ok=True)
    (species_dir / "Croton_tiglium.cds.fa").write_text(">ctg1.t1\nATGAAATTT\n", encoding="utf-8")

    out = tmp_path / "download_manifest_direct_cds_only.xlsx"
    completed = run_script(
        "--provider",
        "direct",
        "--input-dir",
        str(input_root),
        "--output",
        str(out),
    )
    assert completed.returncode == 0, completed.stderr + "\n" + completed.stdout
    rows = read_manifest(out)
    assert len(rows) == 1
    assert rows[0]["provider"] == "direct"
    assert rows[0]["id"] == "Croton_tiglium"
    assert rows[0]["species_key"] == "Croton_tiglium"
    assert rows[0]["cds_filename"] == "Croton_tiglium.cds.fa"
    assert rows[0]["gff_url"] == ""
    assert rows[0]["gbff_url"] == ""
    assert rows[0]["genome_url"] == ""


def test_build_download_manifest_allows_gbff_and_genome_without_gff_or_cds(tmp_path):
    input_root = tmp_path / "dataset"
    species_dir = input_root / "Direct" / "species_wise_original" / "Moringa_oleifera"
    species_dir.mkdir(parents=True, exist_ok=True)
    (species_dir / "Moringa_oleifera.genomic.gbff").write_text(
        "LOCUS       chr1               9 bp    DNA     linear   PLN 01-JAN-2000\n",
        encoding="utf-8",
    )
    (species_dir / "Moringa_oleifera.genome.fa").write_text(">chr1\nATGAAATTT\n", encoding="utf-8")

    out = tmp_path / "download_manifest_direct_gbff.xlsx"
    completed = run_script(
        "--provider",
        "direct",
        "--input-dir",
        str(input_root),
        "--output",
        str(out),
    )
    assert completed.returncode == 0, completed.stderr + "\n" + completed.stdout
    rows = read_manifest(out)
    assert len(rows) == 1
    assert rows[0]["provider"] == "direct"
    assert rows[0]["id"] == "Moringa_oleifera"
    assert rows[0]["species_key"] == "Moringa_oleifera"
    assert rows[0]["cds_url"] == ""
    assert rows[0]["gff_url"] == ""
    assert rows[0]["gbff_filename"] == "Moringa_oleifera.genomic.gbff"
    assert rows[0]["genome_filename"] == "Moringa_oleifera.genome.fa"


def test_build_download_manifest_xlsx_has_provider_and_id_dropdowns(tmp_path):
    out = tmp_path / "download_manifest.xlsx"
    completed = run_script(
        "--provider",
        "all",
        "--input-dir",
        str(SMALL_DATASET_ROOT),
        "--output",
        str(out),
    )
    assert completed.returncode == 0, completed.stderr + "\n" + completed.stdout

    vml = read_comment_vml(out)
    assert "width:144px;height:79px" not in vml
    assert "visibility:hidden" in vml

    workbook = load_workbook(out)
    try:
        sheet = workbook["download_plan"]
        headers = [cell.value for cell in sheet[1]]
        assert headers[0] == "provider"
        assert headers[1] == "id"
        header_comments = {str(cell.value): (cell.comment.text if cell.comment else "") for cell in sheet[1]}
        assert set(header_comments.keys()) == set(headers)
        assert all(header_comments[name] != "" for name in headers)
        assert "Allowed values:" in header_comments["provider"]
        assert "Provider-specific identifier." in header_comments["id"]
        assert "Supported placeholders: {id}, {species_key}, {provider}." in header_comments["cds_url_template"]
        assert "provider=local" in header_comments["local_cds_path"]

        validations = list(sheet.data_validations.dataValidation)
        provider_validation = next(v for v in validations if str(v.sqref) == "A2:A5000")
        id_validation = next(v for v in validations if str(v.sqref) == "B2:B5000")
        assert "_lists!$A$1:$A$" in str(provider_validation.formula1)
        assert 'INDIRECT("id_opts_"&$A2)' in str(id_validation.formula1)

        list_sheet = workbook["_lists"]
        provider_values = [list_sheet.cell(row=i, column=1).value for i in range(1, 21)]
        assert provider_values == [
            "ensembl",
            "ensemblplants",
            "ncbi",
            "coge",
            "cngb",
            "gwh",
            "citrusgenomedb",
            "figshare",
            "plantgarden",
            "plantaedb",
            "flybase",
            "wormbase",
            "vectorbase",
            "fernbase",
            "veupathdb",
            "dictybase",
            "insectbase",
            "oryza_minuta",
            "direct",
            "local",
        ]
        assert "id_opts_ensembl" in workbook.defined_names
        assert "id_opts_ncbi" in workbook.defined_names
        assert "id_opts_coge" in workbook.defined_names
        assert "id_opts_gwh" in workbook.defined_names
        assert "id_opts_citrusgenomedb" in workbook.defined_names
        assert "id_opts_figshare" in workbook.defined_names
        assert "id_opts_plantgarden" in workbook.defined_names
        assert "id_opts_plantaedb" in workbook.defined_names
        assert "id_opts_fernbase" in workbook.defined_names
        assert "id_opts_veupathdb" in workbook.defined_names
        assert "id_opts_dictybase" in workbook.defined_names
        assert "id_opts_insectbase" in workbook.defined_names
        assert "id_opts_oryza_minuta" in workbook.defined_names
        assert "id_opts_direct" in workbook.defined_names
        assert "id_opts_local" in workbook.defined_names
    finally:
        workbook.close()


def test_build_download_manifest_xlsx_id_lists_are_provider_specific(tmp_path):
    input_root = tmp_path / "dataset"

    coge_species_gid = (
        ("Arabidopsis_thaliana", "24739"),
        ("Oryza_sativa", "24740"),
    )
    for species, gid in coge_species_gid:
        species_dir = input_root / "CoGe" / "species_wise_original" / species
        species_dir.mkdir(parents=True, exist_ok=True)
        (species_dir / (species + ".coge.gid{}.cds.fa".format(gid))).write_text(">g1.t1\nATG\n", encoding="utf-8")
        (species_dir / (species + ".gid{}.gff3".format(gid))).write_text(
            "chr1\tsrc\tgene\t1\t3\t.\t+\t.\tID=g1\n",
            encoding="utf-8",
        )

    local_species = ("Hydrocotyle_leucocephala_HAP1v2.1", "Marchantia_polymorpha_v6")
    for species in local_species:
        species_dir = input_root / "Local" / "species_wise_original" / species
        species_dir.mkdir(parents=True, exist_ok=True)
        (species_dir / (species + ".cds.fa")).write_text(">g1.t1\nATG\n", encoding="utf-8")
        (species_dir / (species + ".gene.gff3")).write_text("chr1\tsrc\tgene\t1\t3\t.\t+\t.\tID=g1\n", encoding="utf-8")

    out = tmp_path / "download_manifest_provider_lists.xlsx"
    completed = run_script(
        "--provider",
        "all",
        "--input-dir",
        str(input_root),
        "--output",
        str(out),
    )
    assert completed.returncode == 0, completed.stderr + "\n" + completed.stdout

    workbook = load_workbook(out)
    try:
        list_sheet = workbook["_lists"]
        provider_order = [
            "ensembl",
            "ensemblplants",
            "ncbi",
            "coge",
            "cngb",
            "gwh",
            "citrusgenomedb",
            "figshare",
            "plantgarden",
            "plantaedb",
            "flybase",
            "wormbase",
            "vectorbase",
            "fernbase",
            "veupathdb",
            "dictybase",
            "insectbase",
            "oryza_minuta",
            "direct",
            "local",
        ]
        provider_col = {provider: idx + 2 for idx, provider in enumerate(provider_order)}
        coge_values = read_list_column_values(list_sheet, provider_col["coge"])
        cngb_values = read_list_column_values(list_sheet, provider_col["cngb"])
        gwh_values = read_list_column_values(list_sheet, provider_col["gwh"])
        citrusgenomedb_values = read_list_column_values(list_sheet, provider_col["citrusgenomedb"])
        figshare_values = read_list_column_values(list_sheet, provider_col["figshare"])
        plantgarden_values = read_list_column_values(list_sheet, provider_col["plantgarden"])
        plantaedb_values = read_list_column_values(list_sheet, provider_col["plantaedb"])
        fernbase_values = read_list_column_values(list_sheet, provider_col["fernbase"])
        veupathdb_values = read_list_column_values(list_sheet, provider_col["veupathdb"])
        dictybase_values = read_list_column_values(list_sheet, provider_col["dictybase"])
        insectbase_values = read_list_column_values(list_sheet, provider_col["insectbase"])
        oryza_minuta_values = read_list_column_values(list_sheet, provider_col["oryza_minuta"])
        direct_values = read_list_column_values(list_sheet, provider_col["direct"])
        local_values = read_list_column_values(list_sheet, provider_col["local"])
        ncbi_values = read_list_column_values(list_sheet, provider_col["ncbi"])

        assert coge_values == [
            "24739 (Arabidopsis thaliana)",
            "29177 (Oryza sativa)",
            "42091 (Zea mays)",
            "78085 (Drosophila melanogaster)",
            "72417 (Caenorhabditis elegans)",
        ]
        assert cngb_values == [
            "CNA0012345 (Homo sapiens)",
            "GCF_000001405.40 (Homo sapiens)",
            "GCA_000001635.9 (Mus musculus)",
            "GCF_049306965.1 (Danio rerio)",
            "GCA_000001215.4 (Drosophila melanogaster)",
        ]
        assert gwh_values == [
            "GWHIGRM00000000.1 (Medicago sativa)",
            "GWHCBHY00000000 (Allium sativum)",
        ]
        assert citrusgenomedb_values == [
            "https://www.citrusgenomedb.org/organism/5799 (Citrus australasica (CGD organism page))",
            "https://www.citrusgenomedb.org/Analysis/2530647 (Citrus australasica cv. AZM genome v1.0 (CGD analysis))",
        ]
        assert figshare_values == [
            "https://figshare.com/articles/dataset/Construction_of_the_super_pan-genome_for_the_genus_Actinidia_reveals_structural_variations_linked_to_phenotypic_diversity/28759280 (Actinidia super pan-genome (figshare article))",
        ]
        assert plantgarden_values == [
            "https://plantgarden.jp/en/list/t64480/genome/t64480.G001 (Actinidia polygama (PlantGARDEN assembly page))",
            "https://plantgarden.jp/en/list/t385388/genome/t385388.G001 (Camellia oleifera (PlantGARDEN assembly page))",
        ]
        assert plantaedb_values == [
            "https://plantaedb.com/taxa/phylum/angiosperms/order/asterales/family/asteraceae/subfamily/asteroideae/tribe/astereae/subtribe/conyzinae/genus/erigeron/species/erigeron-breviscapus (Erigeron breviscapus (PlantaeDB page))",
            "https://plantaedb.com/taxa/phylum/angiosperms/order/ranunculales/family/berberidaceae/genus/berberis/species/berberis-thunbergii (Berberis thunbergii (PlantaeDB page))",
        ]
        assert fernbase_values == [
            "Azolla_filiculoides (Azolla filiculoides)",
            "Salvinia_cucullata_v2 (Salvinia cucullata v2)",
        ]
        assert veupathdb_values == [
            "EnuttalliP19 (Entamoeba nuttalli)",
        ]
        assert dictybase_values == [
            "Dictyostelium_discoideum (Dictyostelium discoideum)",
        ]
        assert insectbase_values == [
            "IBG_00001 (Abrostola tripartita)",
        ]
        assert oryza_minuta_values == [
            "gramene_tetraploids (Oryza minuta Gramene tetraploids)",
        ]
        assert direct_values == [
            "direct_example_species (Direct URL manifest row)",
        ]
        assert local_values == [
            "Hydrocotyle_leucocephala_HAP1v2.1",
            "Marchantia_polymorpha_v6",
        ]
        assert ncbi_values == [
            "GCF_000001405.40 (Homo sapiens)",
            "GCA_000001635.9 (Mus musculus)",
            "GCF_049306965.1 (Danio rerio)",
            "GCA_000001215.4 (Drosophila melanogaster)",
            "GCF_000002985.6 (Caenorhabditis elegans)",
        ]
    finally:
        workbook.close()


def test_build_download_manifest_xlsx_prefers_snapshot_for_full_providers(tmp_path):
    out = tmp_path / "download_manifest.xlsx"
    snapshot = tmp_path / "id_options_snapshot.json"
    snapshot.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "providers": {
                    "ensembl": [
                        {"id": "homo_sapiens", "species": "Homo sapiens"},
                        {"id": "pan_troglodytes", "species": "Pan troglodytes"},
                    ],
                    "ensemblplants": [
                        {"id": "arabidopsis_thaliana", "species": "Arabidopsis thaliana"},
                    ],
                    "flybase": [
                        {"id": "dmel_r6.66", "species": "Drosophila melanogaster"},
                    ],
                    "wormbase": [
                        {"id": "caenorhabditis_elegans_prjna13758", "species": "Caenorhabditis elegans"},
                    ],
                    "vectorbase": [
                        {"id": "AgambiaePEST", "species": "Anopheles gambiae"},
                    ],
                    "fernbase": [
                        {"id": "Ceratopteris_richardii", "species": "Ceratopteris richardii"},
                    ],
                    "veupathdb": [
                        {"id": "EhistolyticaHM1IMSS", "species": "Entamoeba histolytica"},
                    ],
                    "dictybase": [
                        {"id": "Dictyostelium_discoideum", "species": "Dictyostelium discoideum"},
                    ],
                    "insectbase": [
                        {"id": "IBG_99999", "species": "Bombyx mori"},
                    ],
                    "gwh": [
                        {"id": "GWHZZZZ00000000.1", "species": "Fake gwh species"},
                    ],
                    "figshare": [
                        {
                            "id": "https://figshare.example/articles/dataset/fake_bundle/999",
                            "species": "Fake figshare species",
                        },
                    ],
                    "plantgarden": [
                        {
                            "id": "https://plantgarden.example/en/list/t999/genome/t999.G001",
                            "species": "Fake PlantGARDEN species",
                        },
                    ],
                    "plantaedb": [
                        {
                            "id": "https://plantaedb.example/species/example",
                            "species": "Fake PlantaeDB species",
                        },
                    ],
                    "direct": [
                        {
                            "id": "snapshot_direct_species",
                            "species": "Snapshot direct",
                            "species_key": "Snapshot_direct_species",
                            "cds_url": "https://example.org/snapshot_direct_species.cds.fa.gz",
                            "gff_url": "https://example.org/snapshot_direct_species.gff3.gz",
                            "genome_url": "https://example.org/snapshot_direct_species.genome.fa.gz",
                            "cds_filename": "snapshot_direct_species.cds.fa.gz",
                            "gff_filename": "snapshot_direct_species.gff3.gz",
                            "genome_filename": "snapshot_direct_species.genome.fa.gz",
                        },
                    ],
                    "local": [
                        {"id": "/data/local_species_1", "species": "Local species 1"},
                        {"id": "/data/local_species_2", "species": "Local species 2"},
                    ],
                    "ncbi": [
                        {"id": "FAKE_NCBI_ID", "species": "Fake species"},
                    ],
                    "coge": [
                        {"id": "99999", "species": "Fake coge species"},
                    ],
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    completed = run_script(
        "--provider",
        "all",
        "--input-dir",
        str(SMALL_DATASET_ROOT),
        "--output",
        str(out),
        "--id-options-snapshot",
        str(snapshot),
    )
    assert completed.returncode == 0, completed.stderr + "\n" + completed.stdout

    workbook = load_workbook(out)
    try:
        list_sheet = workbook["_lists"]
        provider_order = [
            "ensembl",
            "ensemblplants",
            "ncbi",
            "coge",
            "cngb",
            "gwh",
            "citrusgenomedb",
            "figshare",
            "plantgarden",
            "plantaedb",
            "flybase",
            "wormbase",
            "vectorbase",
            "fernbase",
            "veupathdb",
            "dictybase",
            "insectbase",
            "oryza_minuta",
            "direct",
            "local",
        ]
        provider_col = {provider: idx + 2 for idx, provider in enumerate(provider_order)}
        ensembl_values = read_list_column_values(list_sheet, provider_col["ensembl"])
        ensemblplants_values = read_list_column_values(list_sheet, provider_col["ensemblplants"])
        ncbi_values = read_list_column_values(list_sheet, provider_col["ncbi"])
        coge_values = read_list_column_values(list_sheet, provider_col["coge"])
        cngb_values = read_list_column_values(list_sheet, provider_col["cngb"])
        gwh_values = read_list_column_values(list_sheet, provider_col["gwh"])
        citrusgenomedb_values = read_list_column_values(list_sheet, provider_col["citrusgenomedb"])
        figshare_values = read_list_column_values(list_sheet, provider_col["figshare"])
        plantgarden_values = read_list_column_values(list_sheet, provider_col["plantgarden"])
        plantaedb_values = read_list_column_values(list_sheet, provider_col["plantaedb"])
        flybase_values = read_list_column_values(list_sheet, provider_col["flybase"])
        wormbase_values = read_list_column_values(list_sheet, provider_col["wormbase"])
        vectorbase_values = read_list_column_values(list_sheet, provider_col["vectorbase"])
        fernbase_values = read_list_column_values(list_sheet, provider_col["fernbase"])
        veupathdb_values = read_list_column_values(list_sheet, provider_col["veupathdb"])
        dictybase_values = read_list_column_values(list_sheet, provider_col["dictybase"])
        insectbase_values = read_list_column_values(list_sheet, provider_col["insectbase"])
        oryza_minuta_values = read_list_column_values(list_sheet, provider_col["oryza_minuta"])
        direct_values = read_list_column_values(list_sheet, provider_col["direct"])
        local_values = read_list_column_values(list_sheet, provider_col["local"])

        assert ensembl_values == [
            "homo_sapiens (Homo sapiens)",
            "pan_troglodytes (Pan troglodytes)",
        ]
        assert ensemblplants_values == [
            "arabidopsis_thaliana (Arabidopsis thaliana)",
        ]
        assert "FAKE_NCBI_ID (Fake species)" not in ncbi_values
        assert ncbi_values == [
            "GCF_000001405.40 (Homo sapiens)",
            "GCA_000001635.9 (Mus musculus)",
            "GCF_049306965.1 (Danio rerio)",
            "GCA_000001215.4 (Drosophila melanogaster)",
            "GCF_000002985.6 (Caenorhabditis elegans)",
        ]
        assert coge_values == [
            "24739 (Arabidopsis thaliana)",
            "29177 (Oryza sativa)",
            "42091 (Zea mays)",
            "78085 (Drosophila melanogaster)",
            "72417 (Caenorhabditis elegans)",
        ]
        assert cngb_values == [
            "CNA0012345 (Homo sapiens)",
            "GCF_000001405.40 (Homo sapiens)",
            "GCA_000001635.9 (Mus musculus)",
            "GCF_049306965.1 (Danio rerio)",
            "GCA_000001215.4 (Drosophila melanogaster)",
        ]
        assert gwh_values == ["GWHZZZZ00000000.1 (Fake gwh species)"]
        assert citrusgenomedb_values == [
            "https://www.citrusgenomedb.org/organism/5799 (Citrus australasica (CGD organism page))",
            "https://www.citrusgenomedb.org/Analysis/2530647 (Citrus australasica cv. AZM genome v1.0 (CGD analysis))",
        ]
        assert figshare_values == ["https://figshare.example/articles/dataset/fake_bundle/999 (Fake figshare species)"]
        assert plantgarden_values == ["https://plantgarden.example/en/list/t999/genome/t999.G001 (Fake PlantGARDEN species)"]
        assert plantaedb_values == ["https://plantaedb.example/species/example (Fake PlantaeDB species)"]
        assert flybase_values == ["dmel_r6.66 (Drosophila melanogaster)"]
        assert wormbase_values == ["caenorhabditis_elegans_prjna13758 (Caenorhabditis elegans)"]
        assert vectorbase_values == ["AgambiaePEST (Anopheles gambiae)"]
        assert fernbase_values == ["Ceratopteris_richardii (Ceratopteris richardii)"]
        assert veupathdb_values == ["EhistolyticaHM1IMSS (Entamoeba histolytica)"]
        assert dictybase_values == ["Dictyostelium_discoideum (Dictyostelium discoideum)"]
        assert insectbase_values == ["IBG_99999 (Bombyx mori)"]
        assert oryza_minuta_values == ["gramene_tetraploids (Oryza minuta Gramene tetraploids)"]
        assert direct_values == ["snapshot_direct_species (Snapshot direct)"]
        assert local_values == ["/data/local_species_1", "/data/local_species_2"]
    finally:
        workbook.close()


def test_build_download_manifest_xlsx_adds_direct_catalog_sheet_and_formulas(tmp_path):
    out = tmp_path / "download_manifest.xlsx"
    snapshot = tmp_path / "id_options_snapshot.json"
    snapshot.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "providers": {
                    "direct": [
                        {
                            "id": "snapshot_direct_species",
                            "species": "Snapshot direct",
                            "species_key": "Snapshot_direct_species",
                            "cds_url": "https://example.org/snapshot_direct_species.cds.fa.gz",
                            "gff_url": "https://example.org/snapshot_direct_species.gff3.gz",
                            "genome_url": "https://example.org/snapshot_direct_species.genome.fa.gz",
                            "cds_filename": "snapshot_direct_species.cds.fa.gz",
                            "gff_filename": "snapshot_direct_species.gff3.gz",
                            "genome_filename": "snapshot_direct_species.genome.fa.gz",
                        },
                    ],
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    completed = run_script(
        "--provider",
        "all",
        "--input-dir",
        str(SMALL_DATASET_ROOT),
        "--output",
        str(out),
        "--id-options-snapshot",
        str(snapshot),
    )
    assert completed.returncode == 0, completed.stderr + "\n" + completed.stdout

    workbook = load_workbook(out, data_only=False)
    try:
        sheet = workbook["download_plan"]
        direct_sheet = workbook["_direct_catalog"]

        assert direct_sheet.sheet_state == "hidden"
        assert direct_sheet["A2"].value == "snapshot_direct_species (Snapshot direct)"
        assert direct_sheet["B2"].value == "snapshot_direct_species"
        assert direct_sheet["D2"].value == "Snapshot_direct_species"
        assert direct_sheet["E2"].value == "https://example.org/snapshot_direct_species.cds.fa.gz"
        assert direct_sheet["F2"].value == "https://example.org/snapshot_direct_species.gff3.gz"
        assert direct_sheet["H2"].value == "https://example.org/snapshot_direct_species.genome.fa.gz"

        assert sheet["C2"].value == "Ostreococcus_lucimarinus"
        assert isinstance(sheet["C3"].value, str)
        assert sheet["C3"].value.startswith('=IF($A3<>"direct"')
        assert "_direct_catalog" in sheet["C3"].value
        assert "_direct_catalog" in str(sheet["D3"].value)
        assert "_direct_catalog" in str(sheet["F3"].value)
        assert "_direct_catalog" in str(sheet["M3"].value)
    finally:
        workbook.close()


def test_build_download_manifest_xlsx_reads_direct_catalog_manifest(tmp_path):
    direct_manifest = tmp_path / "direct_manifest.xlsx"
    from openpyxl import Workbook

    manifest_wb = Workbook()
    manifest_ws = manifest_wb.active
    manifest_ws.append(
        [
            "provider",
            "id",
            "species_key",
            "cds_url",
            "gff_url",
            "gbff_url",
            "genome_url",
            "cds_archive_member",
            "gff_archive_member",
            "gbff_archive_member",
            "genome_archive_member",
            "cds_filename",
            "gff_filename",
            "gbff_filename",
            "genome_filename",
            "cds_url_template",
            "gff_url_template",
            "genome_url_template",
            "local_cds_path",
            "local_gff_path",
            "local_gbff_path",
            "local_genome_path",
        ]
    )
    manifest_ws.append(
        [
            "direct",
            "Fragaria_ananassa_FAN_r2.3",
            "Fragaria_ananassa",
            "https://example.org/fragaria.cds.fa.gz",
            "https://example.org/fragaria.gff3.gz",
            "",
            "https://example.org/fragaria.genome.fa.gz",
            "",
            "",
            "",
            "",
            "fragaria.cds.fa.gz",
            "fragaria.gff3.gz",
            "",
            "fragaria.genome.fa.gz",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    manifest_wb.save(direct_manifest)

    out = tmp_path / "download_manifest.xlsx"
    completed = run_script(
        "--provider",
        "all",
        "--input-dir",
        str(SMALL_DATASET_ROOT),
        "--output",
        str(out),
        "--direct-catalog-manifest",
        str(direct_manifest),
    )
    assert completed.returncode == 0, completed.stderr + "\n" + completed.stdout

    workbook = load_workbook(out, data_only=False)
    try:
        sheet = workbook["download_plan"]
        list_sheet = workbook["_lists"]
        direct_sheet = workbook["_direct_catalog"]
        provider_values = [list_sheet.cell(row=i, column=1).value for i in range(1, 32)]
        provider_values = [value for value in provider_values if value]
        provider_col = {provider: idx + 2 for idx, provider in enumerate(provider_values)}
        assert direct_sheet["A2"].value == "Fragaria_ananassa_FAN_r2.3 (Fragaria ananassa)"
        assert direct_sheet["D2"].value == "Fragaria_ananassa"
        assert direct_sheet["E2"].value == "https://example.org/fragaria.cds.fa.gz"
        assert list_sheet.cell(row=1, column=provider_col["direct"]).value == "Fragaria_ananassa_FAN_r2.3 (Fragaria ananassa)"
        assert "_direct_catalog" in str(sheet["C3"].value)
        assert "_direct_catalog" in str(sheet["D3"].value)
    finally:
        workbook.close()


def test_build_download_manifest_xlsx_direct_catalog_manifest_allows_template_only_output(tmp_path):
    direct_manifest = tmp_path / "direct_manifest.tsv"
    direct_manifest.write_text(
        "\t".join(
            [
                "provider",
                "id",
                "species_key",
                "cds_url",
                "gff_url",
                "gbff_url",
                "genome_url",
                "cds_archive_member",
                "gff_archive_member",
                "gbff_archive_member",
                "genome_archive_member",
                "cds_filename",
                "gff_filename",
                "gbff_filename",
                "genome_filename",
                "cds_url_template",
                "gff_url_template",
                "genome_url_template",
                "local_cds_path",
                "local_gff_path",
                "local_gbff_path",
                "local_genome_path",
            ]
        )
        + "\n"
        + "\t".join(
            [
                "direct",
                "Allium_cepa_PlantGARDEN_v1.2",
                "Allium_cepa",
                "https://example.org/allium.cds.fa.gz",
                "https://example.org/allium.gff3.gz",
                "",
                "https://example.org/allium.genome.fa.gz",
                "",
                "",
                "",
                "",
                "allium.cds.fa.gz",
                "allium.gff3.gz",
                "",
                "allium.genome.fa.gz",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    empty_input = tmp_path / "empty_input"
    empty_input.mkdir()
    out = tmp_path / "download_manifest.xlsx"
    completed = run_script(
        "--provider",
        "all",
        "--input-dir",
        str(empty_input),
        "--output",
        str(out),
        "--direct-catalog-manifest",
        str(direct_manifest),
    )
    assert completed.returncode == 0, completed.stderr + "\n" + completed.stdout
    workbook = load_workbook(out, data_only=False)
    try:
        assert "_direct_catalog" in workbook.sheetnames
        assert workbook["_direct_catalog"]["A2"].value == "Allium_cepa_PlantGARDEN_v1.2 (Allium cepa)"
    finally:
        workbook.close()


def test_build_download_manifest_local_provider_from_phytozome_like_fixture(tmp_path):
    input_root = tmp_path / "dataset"
    local_species_dir = input_root / "Local" / "species_wise_original" / "Hydrocotyle_leucocephala_HAP1v2.1"
    local_species_dir.mkdir(parents=True)

    source_dir = SMALL_DATASET_ROOT / "Phytozome" / "species_wise_original" / "Hydrocotyle_leucocephala_HAP1v2.1"
    for name in (
        "HleucocephalaHAP1_768_v2.1.cds_primaryTranscriptOnly.fa",
        "HleucocephalaHAP1_768_v2.1.gene.gff3",
    ):
        shutil.copy2(source_dir / name, local_species_dir / name)

    out = tmp_path / "download_manifest_local.xlsx"
    completed = run_script(
        "--provider",
        "local",
        "--input-dir",
        str(input_root),
        "--output",
        str(out),
    )
    assert completed.returncode == 0, completed.stderr + "\n" + completed.stdout
    rows = read_manifest(out)
    assert len(rows) == 1
    assert rows[0]["provider"] == "local"
    assert rows[0]["id"] == "Hydrocotyle_leucocephala_HAP1v2.1"
    assert rows[0]["cds_filename"].endswith(".cds_primaryTranscriptOnly.fa")
    assert rows[0]["gff_filename"].endswith(".gene.gff3")


def test_build_download_manifest_all_rows_keep_local_last(tmp_path):
    input_root = tmp_path / "dataset"

    ncbi_species_dir = input_root / "NCBI_Genome" / "species_wise_original" / "GCF_000001405.40"
    ncbi_species_dir.mkdir(parents=True, exist_ok=True)
    (ncbi_species_dir / "GCF_000001405.40_GRCh38.p14_cds_from_genomic.fna.gz").write_bytes(b"dummy")
    (ncbi_species_dir / "GCF_000001405.40_GRCh38.p14_genomic.gff.gz").write_bytes(b"dummy")
    (ncbi_species_dir / "GCF_000001405.40_GRCh38.p14_genomic.fna.gz").write_bytes(b"dummy")

    local_species_dir = input_root / "Local" / "species_wise_original" / "Hydrocotyle_leucocephala_HAP1v2.1"
    local_species_dir.mkdir(parents=True, exist_ok=True)
    (local_species_dir / "HleucocephalaHAP1_768_v2.1.cds_primaryTranscriptOnly.fa").write_text(">g1\nATG\n", encoding="utf-8")
    (local_species_dir / "HleucocephalaHAP1_768_v2.1.gene.gff3").write_text(
        "chr1\tsrc\tgene\t1\t3\t.\t+\t.\tID=g1\n",
        encoding="utf-8",
    )

    out = tmp_path / "download_manifest.tsv"
    completed = run_script(
        "--provider",
        "all",
        "--input-dir",
        str(input_root),
        "--output",
        str(out),
    )
    assert completed.returncode == 0, completed.stderr + "\n" + completed.stdout

    rows = read_manifest(out)
    assert [row["provider"] for row in rows] == ["ncbi", "local"]
