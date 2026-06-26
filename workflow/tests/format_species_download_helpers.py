import bz2
import csv
import gzip
import io
import json
import os
import socket
import subprocess
import sys
import tarfile
import threading
import time
import zipfile
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

from openpyxl import Workbook

SCRIPT_PATH = Path(__file__).resolve().parents[1] / "support" / "format_species_inputs.py"
SMALL_DATASET_ROOT = Path(__file__).resolve().parent / "data" / "small_gfe_dataset"
TEST_COMMAND_TIMEOUT = 60


def run_script(*args, env=None, timeout=TEST_COMMAND_TIMEOUT):
    return subprocess.run(
        [sys.executable, str(SCRIPT_PATH), *args],
        capture_output=True,
        text=True,
        check=False,
        env=env,
        timeout=timeout,
    )


def make_manifest(path, rows):
    with open(path, "wt", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "provider",
                "id",
                "species_key",
                "cds_url",
                "gff_url",
                "gbff_url",
                "genome_url",
                "cds_archive_member",
                "gff_archive_member",
                "gbff_archive_member",
                "genome_archive_member",
                "cds_filename",
                "gff_filename",
                "gbff_filename",
                "genome_filename",
                "cds_url_template",
                "gff_url_template",
                "genome_url_template",
                "local_cds_path",
                "local_gff_path",
                "local_gbff_path",
                "local_genome_path",
                "fernbase_confidence_mode",
            ],
            delimiter="\t",
        )
        writer.writeheader()
        for row in rows:
            payload = dict(row)
            if str(payload.get("id", "") or "").strip() == "":
                payload["id"] = str(payload.get("species_key", "") or "").strip()
            writer.writerow(payload)


def make_manifest_xlsx(path, headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "download_plan"
    sheet.append(list(headers))
    for row in rows:
        sheet.append([row.get(key, "") for key in headers])
    workbook.save(path)
    workbook.close()


def make_manifest_xlsx_with_direct_catalog(path, manifest_headers, manifest_rows, catalog_headers, catalog_rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "download_plan"
    sheet.append(list(manifest_headers))
    for row in manifest_rows:
        sheet.append([row.get(key, "") for key in manifest_headers])

    catalog_sheet = workbook.create_sheet("_direct_catalog")
    catalog_sheet.append(list(catalog_headers))
    for row in catalog_rows:
        catalog_sheet.append([row.get(key, "") for key in catalog_headers])

    workbook.save(path)
    workbook.close()


def to_file_url(path):
    return path.resolve().as_uri()


def _current_boot_id():
    boot_id_path = Path("/proc/sys/kernel/random/boot_id")
    try:
        if boot_id_path.is_file():
            return boot_id_path.read_text(encoding="utf-8").strip()
    except OSError:
        pass
    completed = subprocess.run(
        ["sysctl", "-n", "kern.bootsessionuuid"],
        capture_output=True,
        text=True,
        check=False,
        timeout=TEST_COMMAND_TIMEOUT,
    )
    if completed.returncode != 0:
        return ""
    return completed.stdout.strip()


class _NcbiFixtureHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, root_dir=None, **kwargs):
        self._root_dir = root_dir
        super().__init__(*args, directory=str(root_dir), **kwargs)

    def _send_json(self, payload):
        body = json.dumps(payload).encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        if self.path.startswith("/eutils/esearch.fcgi"):
            self._send_json(
                {
                    "header": {"type": "esearch", "version": "0.3"},
                    "esearchresult": {"idlist": ["12345"]},
                }
            )
            return
        if self.path.startswith("/eutils/esummary.fcgi"):
            self._send_json(
                {
                    "header": {"type": "esummary", "version": "0.3"},
                    "result": {
                        "uids": ["12345"],
                        "12345": {
                            "ftppath_refseq": "ftp://ftp.ncbi.nlm.nih.gov/genomes/all/GCF/000/001/405/GCF_000001405.40_GRCh38.p14",
                            "organism": "Homo sapiens (human)",
                            "speciesname": "Homo sapiens",
                        },
                    },
                }
            )
            return
        super().do_GET()


class _VEuPathDbFixtureHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, root_dir=None, **kwargs):
        self._root_dir = root_dir
        super().__init__(*args, directory=str(root_dir), **kwargs)

    def _send_json(self, payload):
        body = json.dumps(payload).encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        if self.path.startswith("/veupathdb/service/record-types/organism/searches/GenomeDataTypes/reports/standard"):
            base = "http://127.0.0.1:{}".format(self.server.server_port)
            self._send_json(
                {
                    "records": [
                        {
                            "attributes": {
                                "primary_key": "Entamoeba nuttalli P19",
                                "species": "Entamoeba nuttalli",
                                "project_id": "AmoebaDB",
                                "URLGenomeFasta": (
                                    base
                                    + "/common/downloads/Current_Release/EnuttalliP19/fasta/data/"
                                    "AmoebaDB-68_EnuttalliP19_Genome.fasta"
                                ),
                                "URLgff": (
                                    base
                                    + "/common/downloads/Current_Release/EnuttalliP19/gff/data/"
                                    "AmoebaDB-68_EnuttalliP19.gff"
                                ),
                                "URLproteinFasta": (
                                    base
                                    + "/common/downloads/Current_Release/EnuttalliP19/fasta/data/"
                                    "AmoebaDB-68_EnuttalliP19_AnnotatedProteins.fasta"
                                ),
                            }
                        }
                    ]
                }
            )
            return
        super().do_GET()


class _InsectBaseFixtureHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, root_dir=None, **kwargs):
        self._root_dir = root_dir
        super().__init__(*args, directory=str(root_dir), **kwargs)

    def _send_json(self, payload):
        body = json.dumps(payload).encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        if self.path.startswith("/api/genome/genomes/IBG_00001/"):
            self._send_json(
                {
                    "ibg_id": "IBG_00001",
                    "species": "Abrostola tripartita",
                }
            )
            return
        super().do_GET()


def write_tar_bz2(path, members):
    with tarfile.open(path, "w:bz2") as archive:
        for name, content in members.items():
            payload = str(content).encode("utf-8")
            info = tarfile.TarInfo(name=name)
            info.size = len(payload)
            archive.addfile(info, io.BytesIO(payload))


__all__ = (
    "Path",
    "bz2",
    "csv",
    "gzip",
    "io",
    "json",
    "os",
    "socket",
    "subprocess",
    "sys",
    "tarfile",
    "threading",
    "time",
    "parse_qs",
    "urlparse",
    "zipfile",
    "SimpleHTTPRequestHandler",
    "ThreadingHTTPServer",
    "Workbook",
    "SCRIPT_PATH",
    "SMALL_DATASET_ROOT",
    "TEST_COMMAND_TIMEOUT",
    "run_script",
    "make_manifest",
    "make_manifest_xlsx",
    "make_manifest_xlsx_with_direct_catalog",
    "to_file_url",
    "_current_boot_id",
    "_NcbiFixtureHandler",
    "_VEuPathDbFixtureHandler",
    "_InsectBaseFixtureHandler",
    "write_tar_bz2",
)
